"""주간 훈련기관 교육실적 취합 App — 온라인 포털

specs/자동화_흐름도.md 의 흐름을 웹에서 돌린다.

  · 훈련기관: 화면에서 직접 입력하거나 엑셀 업로드로 제출 (둘 다 같은 저장소로)
  · 관리자  : 제출·승인 현황 확인, 승인/반려, 주간 리포트, 엑셀 다운로드
  · 저장소  : data/ 폴더에 주차별 JSON — 전주·작년 자료를 자동으로 불러옴

실행:
    python app.py          (http://localhost:8000)
    python app.py 9000     (포트 지정)
"""

import email
import html
import io
import json
import re
import sys
import time
import webbrowser
from calendar import monthrange
from collections import Counter
from datetime import date
from email.policy import default as email_policy
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from tempfile import TemporaryDirectory
from urllib.parse import parse_qs, quote, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import auth
import calendar_store
import chatbot
import gb_issues
import kosis_stats
import storage
from feature1_aggregate import SHEETS, TARGET_SHEET, process_rows, read_rows, read_targets
from report_engine import OUTLIER_THRESHOLD, build_report, read_notes, read_schedule, today_todos

PORT = 8000

# 연습용 샘플 (저장소에 자료가 없을 때 요약 화면들이 이걸로 돈다)
SAMPLE_DIR = Path(__file__).parent / "inputs"
SAMPLE_PREFIX = "sample-training-data-"


def sample_for(week):
    """그 주차의 연습용 샘플 엑셀 경로. 없으면 None."""
    if not week:
        return None
    p = SAMPLE_DIR / f"{SAMPLE_PREFIX}{week}.xlsx"
    return p if p.is_file() else None


def sample_weeks():
    """inputs/ 에 있는 샘플 주차 키 목록 (오래된 것부터)."""
    keys = []
    for p in SAMPLE_DIR.glob(f"{SAMPLE_PREFIX}*.xlsx"):
        try:
            keys.append((storage.parse_week(p.stem[len(SAMPLE_PREFIX):]), p.stem[len(SAMPLE_PREFIX):]))
        except Exception:
            continue  # 주차 형식이 아닌 파일은 목록에서 뺀다
    return [k for _, k in sorted(keys)]


def prev_sample_week(week):
    """샘플 파일들 중 이 주차 바로 직전 주차. storage.prev_week_key 의 샘플판."""
    y, w = storage.parse_week(week)
    earlier = [k for k in sample_weeks() if storage.parse_week(k) < (y, w)]
    return earlier[-1] if earlier else None


# ── 표시 보조 ────────────────────────────────────────────────────


def e(s):
    return html.escape("" if s is None else str(s))


def pct(v):
    """지표 표시. 터미널(feature1.pct)과 같은 말로 보여준다.

    문자열 상수(`검증 필요`·`계산 불가`)는 그대로, 값이 없으면 '입력 없음'.
    """
    if isinstance(v, str):
        return f'<span class="dim">{e(v)}</span>'
    return '<span class="dim">입력 없음</span>' if v is None else f"{v * 100:.1f}%"


def delta(v):
    if v is None:
        return '<span class="dim">—</span>'
    if abs(v) < 0.05:
        return '<span class="dim">― 0.0%p</span>'
    cls, mark = ("up", "▲") if v > 0 else ("down", "▼")
    return f'<span class="{cls}">{mark} {abs(v):.1f}%p</span>'


def delta_chip(v, base="전주 대비"):
    """지표 카드 아래에 놓는 증감 칩. delta() 와 같은 값을 알약 모양으로 보여준다."""
    if v is None:
        return f'<span class="chip n">—</span> <span>{e(base)} 자료 없음</span>'
    if abs(v) < 0.05:
        return f'<span class="chip n">― 0.0%p</span> <span>{e(base)}</span>'
    # 색은 기존 delta() 와 같게 둔다 — 증가는 빨강, 감소는 파랑.
    # (이수율은 오르는 게 좋고 탈락률은 내리는 게 좋아, 좋음·나쁨으로 칠하지 않는다)
    cls, mark = ("u", "▲") if v > 0 else ("d", "▼")
    return f'<span class="chip {cls}">{mark} {abs(v):.1f}%p</span> <span>{e(base)}</span>'


def phead(eyebrow, title, sub="", actions=""):
    """화면 머리말 — 작은 눈썹글 + 큰 제목, 오른쪽에 버튼 자리."""
    옆 = f'<div class="acts">{actions}</div>' if actions else ""
    아래 = f'<p class="sub" style="margin:0">{sub}</p>' if sub else ""
    return f'<div class="phead"><div class="tt"><p class="eyebrow">{eyebrow}</p><h1>{title}</h1>{아래}</div>{옆}</div>'


def stat(ic, lab, val, foot=""):
    """지표 카드 한 장 — 아이콘 상자 · 이름 · 큰 숫자 · 아래 보조 문구."""
    return (f'<div><div class="ic">{icon(ic)}</div><div class="lab">{lab}</div>'
            f'<div class="val">{val}</div><div class="ft">{foot}</div></div>')


def cell(v, raw=None):
    if v is not None:
        return e(v)
    return f'<span class="bad">{e(raw)}</span>' if raw not in (None, "") else '<span class="dim">—</span>'


WEEK_NOTE = "실적 기준일은 각 주차의 일요일"


def week_label(key, note=False):
    """주차 키를 사람 말로 덧붙여 보여준다 — '2026-W30(26년 7월 4주차)'.

    괄호 안의 '몇 월 몇 주차'는 그 주의 **일요일**(실적 기준일)이 속한 달을 기준으로 센다.
    주가 달을 걸치면 기준일 쪽 달을 따른다.
    """
    try:
        y, w = storage.parse_week(key)
        일요일 = date.fromisocalendar(y, w, 7)
    except Exception:
        return e(key)
    lab = f"{e(key)}({일요일.year % 100}년 {일요일.month}월 {(일요일.day - 1) // 7 + 1}주차)"
    if note:
        lab += f' <span class="dim" style="font-size:12px;font-weight:400">{WEEK_NOTE}</span>'
    return lab


def month_week_sunday(year, month, nth):
    """그 달의 nth 번째 일요일(= 실적 기준일). 없는 주차면 None."""
    try:
        첫날 = date(int(year), int(month), 1)
    except (ValueError, TypeError):
        return None
    첫일요일 = 1 + (6 - 첫날.weekday()) % 7  # weekday(): 월=0 … 일=6
    day = 첫일요일 + 7 * (int(nth) - 1)
    try:
        return date(int(year), int(month), day)
    except ValueError:
        return None  # 그 달에 nth 번째 일요일이 없다


def month_week_to_key(year, month, nth):
    """'몇 월 몇 주차' → 저장용 주차 키. 기준일(일요일)이 속한 ISO 주차로 잡는다."""
    일요일 = month_week_sunday(year, month, nth)
    if 일요일 is None:
        return None, None
    iso = 일요일.isocalendar()
    return storage.week_key(iso[0], iso[1]), 일요일


def key_to_month_week(key):
    """주차 키 → (월, 그 달의 몇 주차). week_label 과 같은 기준(일요일)을 쓴다."""
    try:
        y, w = storage.parse_week(key)
        일요일 = date.fromisocalendar(y, w, 7)
    except Exception:
        오늘 = date.today()
        return 오늘.month, (오늘.day - 1) // 7 + 1
    return 일요일.month, (일요일.day - 1) // 7 + 1


def default_week():
    y, w, _ = date.today().isocalendar()
    return storage.week_key(y, w)


def attachment(filename):
    """Content-Disposition 값. HTTP 헤더는 latin-1 만 담을 수 있어
    한글 파일명은 RFC 5987 방식(filename*)으로 따로 붙인다."""
    ascii_name = filename.encode("ascii", "replace").decode("ascii").replace('"', "_")
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(filename)}"


STYLE = """
*{box-sizing:border-box}
:root{--bg:#f6f7f9;--line:#e8ebf0;--soft:#f2f4f7;--ink:#101828;--mut:#667085;--fade:#98a2b3;
 --acc:#2563eb;--accbg:#eef4ff;--ok:#027a48;--okbg:#ecfdf3;--bad:#b42318;--badbg:#fef3f2}
body{font-family:'Malgun Gothic','맑은 고딕',system-ui,sans-serif;margin:0;background:var(--bg);color:var(--ink);line-height:1.6}
/* 좌측 사이드바 + 본문 (참고 디자인의 2단 구성) */
.app{display:flex;min-height:100vh;align-items:stretch}
.side{flex:0 0 234px;width:234px;background:#fff;border-right:1px solid var(--line);
 padding:16px 12px 18px;display:flex;flex-direction:column;position:sticky;top:0;height:100vh;overflow-y:auto}
.side .brand{display:flex;align-items:center;gap:9px;padding:4px 8px 6px;font-size:14px;font-weight:800;letter-spacing:-.01em}
.side .brand .mk{flex:0 0 28px;height:28px;border-radius:9px;background:linear-gradient(135deg,#2563eb,#7aa7f7);
 color:#fff;display:grid;place-items:center;font-size:13px;font-weight:800}
.side .brand span{white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.side .grp{font-size:11px;font-weight:700;color:var(--fade);letter-spacing:.04em;margin:16px 9px 5px}
.side a{display:flex;align-items:center;gap:10px;padding:8px 10px;border-radius:9px;
 color:#475467;text-decoration:none;font-size:13px;font-weight:500}
.side a svg{flex:0 0 18px;opacity:.8}
.side a:hover{background:#f5f7fa;color:var(--ink)}
.side a.on{background:var(--accbg);color:var(--acc);font-weight:700}
.side a.on svg{opacity:1}
.side .me{margin-top:auto;padding-top:13px;border-top:1px solid var(--line);display:flex;align-items:center;gap:9px}
.side .me .av{flex:0 0 32px;height:32px;border-radius:50%;background:var(--accbg);color:var(--acc);
 display:grid;place-items:center;font-weight:800;font-size:13px}
.side .me .who{min-width:0;font-size:12px;line-height:1.35}
.side .me .who b{display:block;font-size:13px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.side .me .who span{color:var(--mut)}
.side .me a{padding:6px;border-radius:8px}
.main{flex:1;min-width:0}
.wrap{max-width:1180px;margin:0 auto;padding:26px 28px 80px}
/* 화면 머리말 — 작은 눈썹글 + 큰 제목 (참고 디자인의 Insight/Overview) */
.phead{display:flex;justify-content:space-between;align-items:flex-end;gap:14px;flex-wrap:wrap;margin-bottom:20px}
.phead .tt{min-width:0}
.phead .acts{display:flex;gap:8px;align-items:center;flex-wrap:wrap}
.eyebrow{font-size:12px;font-weight:600;color:var(--mut);margin:0 0 1px}
h1{font-size:27px;line-height:1.25;letter-spacing:-.025em;font-weight:800;margin:0 0 6px}
h2{font-size:17px;margin:0 0 4px;display:flex;align-items:center;gap:8px;flex-wrap:wrap;letter-spacing:-.01em}
h3{font-size:14px;margin:20px 0 8px;letter-spacing:-.01em}
.sub{color:var(--mut);font-size:13px;margin:0 0 18px}
.card{background:#fff;border:1px solid var(--line);border-radius:14px;padding:20px 22px;margin-bottom:16px;
 box-shadow:0 1px 2px rgba(16,24,40,.04)}
.card>.hd{border-bottom:1px solid var(--soft);padding-bottom:13px;margin-bottom:15px}
label{display:block;font-weight:700;font-size:13px;margin:14px 0 6px}
input[type=text],input[type=number],input[type=date],textarea,select,input[type=file]{width:100%;padding:9px 11px;
 border:1px solid #d0d5dd;border-radius:9px;font:inherit;background:#fff}
input:focus,textarea:focus,select:focus{outline:2px solid #bfd3fb;outline-offset:0;border-color:var(--acc)}
textarea{min-height:96px;resize:vertical}
.hint{color:var(--mut);font-size:12px;font-weight:400;margin-top:3px}
button{background:var(--acc);color:#fff;border:0;border-radius:10px;padding:9px 17px;font:inherit;font-size:13px;
 font-weight:700;cursor:pointer;box-shadow:0 1px 2px rgba(16,24,40,.08)}
button:hover{background:#1d4ed8}
button.ghost{background:#fff;color:#344054;border:1px solid #d0d5dd}
button.ghost:hover{background:#f9fafb}
button.sm{padding:5px 11px;font-size:12px;border-radius:8px}
button.danger{background:#fff;color:var(--bad);border:1px solid #fecaca}
table{border-collapse:collapse;width:100%;font-size:13px}
th,td{padding:11px 12px;text-align:left;white-space:nowrap}
th{background:#fff;font-size:11.5px;font-weight:700;color:var(--fade);letter-spacing:.03em;border-bottom:1px solid var(--line)}
td{border-bottom:1px solid var(--soft)}
tbody tr:hover{background:#fafbfc}
td.n,th.n{text-align:right;font-variant-numeric:tabular-nums}
tr.err{background:#fffaf5}
tr.total{background:#f9fafb;font-weight:700;border-top:1px solid #d0d5dd}
tr.total:hover{background:#f9fafb}
.scroll{overflow-x:auto}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:10px}
.tag{display:inline-block;padding:2px 9px;border-radius:999px;font-size:11px;font-weight:700;white-space:nowrap}
.t-ok{background:var(--okbg);color:var(--ok)}.t-wait{background:#fffaeb;color:#b54708}
.t-rej{background:var(--badbg);color:var(--bad)}.t-none{background:var(--soft);color:#475467}
.up{color:#dc2626;font-weight:700}.down{color:var(--acc);font-weight:700}
.dim{color:var(--fade)}.bad{color:#b45309;font-weight:700}
.eye{background:var(--okbg);color:var(--ok);border:1px solid #a6f4c5;padding:1px 8px;border-radius:999px;font-size:11px;font-weight:700}
.note{background:#f9fafb;border-left:3px solid #d0d5dd;padding:10px 14px;font-size:13px;color:#475467;margin:12px 0 0;border-radius:0 8px 8px 0}
.warn{background:#fffcf5;border-left-color:#f79009;color:#b54708}
.ok{background:#f6fef9;border-left-color:#12b76a;color:var(--ok)}
/* 지표 카드 — 아이콘 상자 + 큰 숫자 + 증감 칩 */
.kpi{display:grid;grid-template-columns:repeat(auto-fit,minmax(198px,1fr));gap:14px;margin-bottom:16px}
.kpi>div{background:#fff;border:1px solid var(--line);border-radius:14px;padding:16px 18px;
 box-shadow:0 1px 2px rgba(16,24,40,.04)}
.kpi .ic{width:34px;height:34px;border-radius:10px;background:var(--accbg);color:var(--acc);
 display:grid;place-items:center;margin-bottom:12px}
.kpi .lab{font-size:12.5px;color:var(--mut)}
.kpi .val{font-size:26px;font-weight:800;letter-spacing:-.025em;font-variant-numeric:tabular-nums;line-height:1.25}
.kpi .ft{margin-top:12px;padding-top:10px;border-top:1px solid var(--soft);font-size:12px;color:var(--mut);
 display:flex;align-items:center;gap:7px;flex-wrap:wrap;min-height:22px}
.chip{display:inline-flex;align-items:center;gap:3px;padding:2px 8px;border-radius:999px;font-size:11.5px;font-weight:700}
.chip.u{background:var(--badbg);color:#dc2626}.chip.d{background:#eff8ff;color:#175cd3}
.chip.g{background:var(--okbg);color:var(--ok)}.chip.r{background:var(--badbg);color:var(--bad)}
.chip.n{background:var(--soft);color:#475467}
.bar{display:flex;align-items:center;gap:11px;margin:9px 0;font-size:13px}
.bar .nm{width:150px;text-align:right;color:#344054;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.bar .tr{flex:1;background:var(--soft);border-radius:999px;height:8px;min-width:80px}
.bar .fl{height:100%;border-radius:999px;background:var(--acc)}
.bar .fl.low{background:#f04438}
.bar .vl{width:150px;font-variant-numeric:tabular-nums;white-space:nowrap}
.flag{border:1px solid #fecaca;background:#fffbfa;border-radius:12px;padding:13px 16px;margin-bottom:9px}
ul.notes{list-style:none;padding:0;margin:0}
ul.notes li{padding:8px 0;border-bottom:1px solid #f3f4f6;font-size:13px}
ul.notes li:last-child{border-bottom:0}
.srcs{margin-top:5px;display:flex;flex-direction:column;gap:2px}
.srcs a{font-size:12px;color:#2563eb}
.legend{display:flex;gap:16px;flex-wrap:wrap;font-size:12px;color:#374151;margin:2px 0 6px}
details>summary{cursor:pointer;font-size:12px;color:#6b7280;margin-top:8px}
svg text{font-family:inherit}
/* 선 그래프 호버: 시점 띠에 마우스를 올리면 그 시점 값 말풍선이 뜬다 */
.hb .tip{opacity:0;pointer-events:none;transition:opacity .08s}
.hb:hover .tip{opacity:1}
@media print{.hb .tip{display:none}}
.chk{background:#f0fdf4;border:1px solid #86efac;border-radius:9px;padding:12px 15px;margin-top:14px;font-size:13px}
.chk label{display:flex;gap:8px;align-items:flex-start;font-weight:400;margin:5px 0}
.rowbox{border:1px solid #e5e7eb;border-radius:9px;padding:12px;margin-bottom:9px;background:#fcfcfd;position:relative}
.rowbox .del{position:absolute;top:9px;right:9px}
.rowbox label{margin:6px 0 4px}
.inline{display:flex;gap:9px;flex-wrap:wrap;align-items:flex-end}
.inline>div{flex:1;min-width:120px}
a.btn{text-decoration:none}
/* 챗봇 위젯 — 화면 우측 아래 떠 있는 말풍선 버튼과 채팅창 */
.cbtn{position:fixed;right:22px;bottom:22px;z-index:60;width:56px;height:56px;border-radius:50%;
 background:var(--acc);color:#fff;border:0;display:grid;place-items:center;cursor:pointer;
 box-shadow:0 8px 22px rgba(37,99,235,.4);transition:transform .12s}
.cbtn:hover{transform:scale(1.06);background:#1d4ed8}
.cbox{position:fixed;right:22px;bottom:88px;z-index:60;width:360px;max-width:calc(100vw - 32px);
 height:520px;max-height:calc(100vh - 130px);background:#fff;border:1px solid var(--line);
 border-radius:16px;box-shadow:0 18px 48px rgba(16,24,40,.22);display:none;flex-direction:column;overflow:hidden}
.cbox.open{display:flex}
.chd{background:var(--ink);color:#fff;padding:13px 16px;display:flex;align-items:center;gap:9px}
.chd .mk{flex:0 0 26px;height:26px;border-radius:8px;background:linear-gradient(135deg,#2563eb,#7aa7f7);
 display:grid;place-items:center;font-size:12px;font-weight:800}
.chd b{font-size:14px}.chd .st{font-size:11px;color:#9aa4b2}
.chd .x{margin-left:auto;background:transparent;box-shadow:none;padding:4px;color:#9aa4b2;border-radius:7px}
.chd .x:hover{background:#1f2937;color:#fff}
.cmsgs{flex:1;overflow-y:auto;padding:15px;display:flex;flex-direction:column;gap:10px;background:var(--bg)}
.cmsg{max-width:82%;padding:9px 13px;border-radius:14px;font-size:13px;line-height:1.5;white-space:pre-wrap;word-break:break-word}
.cmsg.bot{background:#fff;border:1px solid var(--line);border-bottom-left-radius:4px;align-self:flex-start}
.cmsg.me{background:var(--acc);color:#fff;border-bottom-right-radius:4px;align-self:flex-end}
.cmsg.err{background:var(--badbg);color:var(--bad);border:1px solid #fecaca;align-self:flex-start}
.cmsg.typing{color:var(--mut);align-self:flex-start}
.cfoot{border-top:1px solid var(--line);padding:10px;display:flex;gap:8px;background:#fff}
.cfoot textarea{flex:1;min-height:0;height:40px;max-height:96px;resize:none;padding:9px 11px;font-size:13px}
.cfoot button{padding:0 15px;border-radius:9px}
.cfoot button:disabled{background:#9db8ef;cursor:default}
@media print{.cbtn,.cbox{display:none!important}}
@media (max-width:720px){.cbox{right:12px;left:12px;width:auto;bottom:80px}.cbtn{right:16px;bottom:16px}}
@media print{body{background:#fff}.side,nav,.noprint{display:none!important}
 .card{break-inside:avoid;border-color:#ccc;box-shadow:none}}
/* 중간 화면 — 사이드바를 아이콘만 남기고 좁힌다 */
@media (max-width:1080px){
 .side{flex-basis:62px;width:62px;padding:14px 8px}
 .side .brand span,.side .grp,.side a span,.side .me .who{display:none}
 .side a{justify-content:center;padding:9px 0}
 .side .me{justify-content:center}
}
/* 좁은 화면(휴대폰) — 사이드바를 위쪽 가로 띠로 눕힌다 */
@media (max-width:720px){
 .app{display:block}
 .side{position:static;flex-basis:auto;width:auto;height:auto;border-right:0;border-bottom:1px solid var(--line);
  flex-direction:row;align-items:center;gap:4px;overflow-x:auto;padding:9px 10px}
 .side .brand{padding:0 8px 0 0}
 .side .brand span{display:none}
 .side a{padding:8px 10px;white-space:nowrap}
 .side .me{margin-top:0;margin-left:auto;padding-top:0;border-top:0}
 .wrap{padding:16px 13px 90px}
 h1{font-size:21px} h2{font-size:16px}
 .card{padding:15px 16px;border-radius:12px}
 .phead{align-items:flex-start}
 .inline{flex-direction:column;align-items:stretch}
 .inline>div{max-width:none!important}
 button{width:100%;padding:12px 16px}
 button.sm{width:auto}
 table{font-size:12px}
 th,td{padding:9px 8px}
 .bar{flex-wrap:wrap} .bar .nm{width:100%;text-align:left} .bar .vl{width:auto}
}
/* 월별 달력 (일정 조사) */
.cals{display:grid;grid-template-columns:repeat(auto-fit,minmax(322px,1fr));gap:16px;margin:6px 0 4px}
table.cal{width:100%;border-collapse:collapse;table-layout:fixed}
table.cal caption{caption-side:top;text-align:left;font-weight:700;font-size:14px;padding:0 0 7px}
table.cal th{background:#f9fafb;border:1px solid #e5e7eb;padding:5px 0;font-size:12px;text-align:center;color:#6b7280}
table.cal td{border:1px solid #e5e7eb;vertical-align:top;padding:4px;height:34px;background:#fff}
table.cal td.out{background:#fafafa;border-color:#f3f4f6}
table.cal td.off{background:#fafafa}
table.cal .dnum{font-size:12px;font-weight:700;color:#374151;line-height:1.3}
table.cal td.off .dnum{color:#d1d5db;font-weight:400}
table.cal .sun .dnum,table.cal th.sun{color:#b91c1c}
table.cal .sat .dnum,table.cal th.sat{color:#2563eb}
table.cal td.pickday{height:88px}
table.cal td.cnt{height:56px}
/* 날짜 칸이 좁아서 가로로 늘어놓으면 글자가 깨진다 — 세로로 쌓는다 */
.pick{display:flex;flex-direction:column;gap:2px;margin-top:4px}
.pick label{margin:0;font-weight:400}
.pick input{position:absolute;opacity:0;width:0;height:0}
.pick span{display:block;text-align:center;font-size:11px;padding:2px 0;border:1px solid #e5e7eb;
 border-radius:5px;color:#9ca3af;background:#fff;cursor:pointer;user-select:none;line-height:1.4}
.pick input:focus-visible+span{outline:2px solid #2563eb;outline-offset:1px}
.pick .ok input:checked+span{background:#166534;border-color:#166534;color:#fff}
.pick .no input:checked+span{background:#b91c1c;border-color:#b91c1c;color:#fff}
.pick .maybe input:checked+span{background:#6b7280;border-color:#6b7280;color:#fff}
.cntbox{margin-top:4px;font-size:11px;color:#6b7280;line-height:1.4}
.cntbox b{font-size:13px;color:#166534;font-variant-numeric:tabular-nums}
.cntbox .zero{color:#9ca3af;font-weight:400}
.cntbox .best{display:inline-block;background:#dcfce7;color:#166534;border-radius:4px;padding:0 4px;font-weight:700}
/* 엑셀처럼 쓰는 입력 표 */
/* 카드용 .grid{display:grid} 가 이 표에도 먹으면 머리글·입력칸 열이 어긋난다 — 표 레이아웃으로 되돌린다. */
table.grid{display:table;grid-template-columns:none;gap:0}
table.grid td{padding:4px 5px;white-space:nowrap}
table.grid input,table.grid select{padding:6px 7px;font-size:13px;min-width:92px}
table.grid input[type=text]{min-width:150px}
table.grid input[type=number]{min-width:78px;text-align:right}
table.grid td:last-child{width:1%}
table.grid button{width:auto}
/* 모바일 전용 화면(/m) 요소 */
.mtop{display:flex;align-items:center;gap:9px;background:#fff;border-bottom:1px solid var(--line);padding:11px 14px;font-size:13px}
.mtop .mk{flex:0 0 26px;height:26px;border-radius:8px;background:linear-gradient(135deg,#2563eb,#7aa7f7);
 color:#fff;display:grid;place-items:center;font-size:12px;font-weight:800}
.mtop .who{margin-left:auto;font-size:12px;color:var(--mut)}
.mcard{display:block;border:1px solid var(--line);border-radius:13px;padding:15px;margin-bottom:10px;background:#fff;text-decoration:none;color:inherit;box-shadow:0 1px 2px rgba(16,24,40,.04)}
.mcard b{display:block;font-size:15px;margin-bottom:3px}
.mcard .big{font-size:26px;font-weight:700;font-variant-numeric:tabular-nums}
.mrow{display:flex;justify-content:space-between;align-items:center;gap:10px;font-size:13px}
.mlist li{list-style:none;border-bottom:1px solid #f3f4f6;padding:10px 0;font-size:13px}
.mlist{padding:0;margin:0}
.mnav{display:flex;gap:8px;flex-wrap:wrap;margin:14px 0}
.mnav a{flex:1 1 46%;text-align:center;background:#111827;color:#fff;border-radius:9px;padding:12px 8px;text-decoration:none;font-size:13px;font-weight:700}
.mnav a.ghost{background:#fff;color:#374151;border:1px solid #d1d5db}
"""


# 사이드바 아이콘 — 선(stroke) 방식 18px SVG. 색은 글자색을 따라간다.
ICONS = {
    "home": "M3 9.4 10 4l7 5.4V16a1 1 0 0 1-1 1h-3.2v-4.6H7.2V17H4a1 1 0 0 1-1-1z",
    "pencil": "M4 13.4 13.1 4.3a1.5 1.5 0 0 1 2.1 0l.5.5a1.5 1.5 0 0 1 0 2.1L6.6 16H4z",
    "upload": "M10 12.5v-9m0 0L6.7 6.8M10 3.5l3.3 3.3M3.5 13v2A1.5 1.5 0 0 0 5 16.5h10a1.5 1.5 0 0 0 1.5-1.5v-2",
    "check": "M10 2.8a7.2 7.2 0 1 0 0 14.4 7.2 7.2 0 0 0 0-14.4zm-3.1 7.4 2.2 2.2 4.1-4.4",
    "chart": "M4 16.2V9.4m4 6.8V4.2m4 12v-4.6m4 4.6V7.2",
    "star": "m10 3.2 2 4.2 4.5.6-3.3 3.1.8 4.5-4-2.2-4 2.2.8-4.5L3.5 8l4.5-.6z",
    "calendar": "M4 6.6A1.6 1.6 0 0 1 5.6 5h8.8A1.6 1.6 0 0 1 16 6.6v7.8a1.6 1.6 0 0 1-1.6 1.6H5.6A1.6 1.6 0 0 1 4 14.4zM4 8.8h12M7.4 3.4v3M12.6 3.4v3",
    "news": "M4.2 5.6h8.6v10.8H5.4a1.2 1.2 0 0 1-1.2-1.2zm8.6 2.6h2.9v6.9a1.5 1.5 0 0 1-2.9 0zM6.4 8.2h4.2M6.4 11h4.2M6.4 13.6h2.6",
    "bulb": "M7.8 15.4h4.4M8.6 17.4h2.8M10 2.9a4.6 4.6 0 0 1 2.7 8.3c-.4.3-.7.8-.7 1.4H8c0-.6-.3-1.1-.7-1.4A4.6 4.6 0 0 1 10 2.9z",
    "trend": "M3.4 13.6 8 9l3 3 5.6-5.6M12.4 6.4h4.2v4.2",
    "users": "M12.8 16.4v-1.6a3 3 0 0 0-3-3H6.2a3 3 0 0 0-3 3v1.6M8 9.2a2.9 2.9 0 1 0 0-5.8 2.9 2.9 0 0 0 0 5.8m8.8 7.2v-1.6a3 3 0 0 0-2.2-2.9M13 3.6a3 3 0 0 1 0 5.6",
    "list": "M4 5.6h12M4 10h12M4 14.4h8",
    "phone": "M7 2.6h6A1.5 1.5 0 0 1 14.5 4.1v11.8A1.5 1.5 0 0 1 13 17.4H7a1.5 1.5 0 0 1-1.5-1.5V4.1A1.5 1.5 0 0 1 7 2.6zM8.8 15h2.4",
    "logout": "M12.4 6.2V4.6A1.6 1.6 0 0 0 10.8 3H5.6A1.6 1.6 0 0 0 4 4.6v10.8A1.6 1.6 0 0 0 5.6 17h5.2a1.6 1.6 0 0 0 1.6-1.6v-1.6M8.6 10h8m0 0-2.6-2.6M16.6 10 14 12.6",
}


def icon(name):
    d = ICONS.get(name)
    return (f'<svg width="18" height="18" viewBox="0 0 20 20" fill="none" stroke="currentColor" '
            f'stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"><path d="{d}"/></svg>'
            if d else "")


def side_nav(sess, active):
    """사이드바 메뉴. 구역(주 메뉴 / 훈련실적 / 참고자료 / 설정)으로 묶는다.

    가로 드롭다운을 세로 구역으로 바꾼 것뿐이라, 어떤 화면이 어디 속하는지는 그대로다 —
    리포트·요약은 훈련실적 안에, 종합 제언은 경북 이슈 바로 아래에 둔다.
    """
    관리자 = sess["role"] == "admin"
    구역 = [
        ("주 메뉴", [("/", "홈", "home", "home")]
         + ([] if 관리자 else [("/mine", "내 제출 내역", "mine", "list")])
         + [("/calendar", "캘린더", "calendar", "calendar")]),
        ("훈련실적", [("/submit", "① 직접 입력", "submit", "pencil"),
                      ("/upload", "① 엑셀 업로드", "upload", "upload")]
         + ([("/admin", "② 취합·승인", "admin", "check"),
             ("/report", "③ 주간 리포트", "report", "chart"),
             ("/brief", "③-1 핵심 요약", "brief", "star")] if 관리자 else [])),
        ("참고자료", [("/issues", "경북 이슈", "issues", "news"),
                      ("/advice", "종합 제언", "advice", "bulb"),
                      ("/stats", "고용 통계", "stats", "trend")]),
        ("설정", [("/users", "계정 관리", "users", "users")] if 관리자 else []),
    ]
    out = ""
    for 이름, items in 구역:
        if not items:
            continue
        out += f'<div class="grp">{이름}</div>'
        for href, label, key, ic in items:
            on = " class=\"on\"" if active == key else ""
            out += f'<a href="{href}"{on} title="{label}">{icon(ic)}<span>{label}</span></a>'
    return out


CHAT_WIDGET = """
<button class="cbtn" id="cbtn" title="무엇이든 물어보세요" aria-label="도움 챗봇 열기" onclick="cbToggle()">
<svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"
 stroke-linecap="round" stroke-linejoin="round"><path d="M21 11.5a8.5 8.5 0 0 1-12.2 7.6L3 20.5l1.4-5.3A8.5 8.5 0 1 1 21 11.5z"/>
<path d="M8.5 11.5h.01M12 11.5h.01M15.5 11.5h.01"/></svg></button>
<div class="cbox" id="cbox" role="dialog" aria-label="도움 챗봇">
<div class="chd"><span class="mk">?</span><div><b>무엇이든 물어보세요</b><br>
<span class="st">이 서비스에 대해 안내해 드립니다</span></div>
<button class="x" onclick="cbToggle()" aria-label="닫기">✕</button></div>
<div class="cmsgs" id="cmsgs"></div>
<div class="cfoot"><textarea id="cin" rows="1" placeholder="예: 실적은 어떻게 제출하나요?"
 onkeydown="cbKey(event)" aria-label="질문 입력"></textarea>
<button id="csend" onclick="cbSend()">보내기</button></div>
</div>
<script>
var cbHist=[], cbBusy=false, cbGreeted=false;
function cbToggle(){
 var box=document.getElementById('cbox'); box.classList.toggle('open');
 if(box.classList.contains('open')){
  if(!cbGreeted){cbAdd('bot','안녕하세요. 이 서비스에 대해 궁금한 점을 물어봐 주세요. 예를 들어 "실적은 어떻게 내나요?" 처럼요.');cbGreeted=true;}
  document.getElementById('cin').focus();
 }
}
function cbAdd(cls,text){
 var d=document.createElement('div'); d.className='cmsg '+cls; d.textContent=text;
 var box=document.getElementById('cmsgs'); box.appendChild(d); box.scrollTop=box.scrollHeight; return d;
}
function cbKey(ev){ if(ev.key==='Enter'&&!ev.shiftKey){ev.preventDefault();cbSend();} }
function cbSend(){
 if(cbBusy) return;
 var inp=document.getElementById('cin'), q=inp.value.trim(); if(!q) return;
 inp.value=''; cbAdd('me',q); cbHist.push({role:'user',content:q});
 cbBusy=true; document.getElementById('csend').disabled=true;
 var typing=cbAdd('typing','답변을 작성하고 있습니다…');
 fetch('/chat',{method:'POST',headers:{'Content-Type':'application/json'},
  body:JSON.stringify({messages:cbHist})})
 .then(function(r){return r.json();})
 .then(function(j){
  typing.remove();
  if(j.reply){cbAdd('bot',j.reply); cbHist.push({role:'assistant',content:j.reply});}
  else{cbAdd('err',j.error||'답변을 가져오지 못했습니다.');}
 })
 .catch(function(){typing.remove();cbAdd('err','연결에 문제가 있습니다. 잠시 뒤 다시 시도해 주세요.');})
 .finally(function(){cbBusy=false;document.getElementById('csend').disabled=false;document.getElementById('cin').focus();});
}
</script>"""


def page(title, body, active="", sess=None):
    if not sess:
        return f"""<!doctype html><html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>{e(title)}</title><style>{STYLE}</style></head><body>
<div class="wrap">{body}</div></body></html>"""

    역할 = "관리자" if sess["role"] == "admin" else (sess["org"] or "기관")
    아바타 = e((sess["org"] or sess["id"] or "?")[0])
    return f"""<!doctype html><html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>{e(title)}</title><style>{STYLE}</style></head><body>
<div class="app">
<aside class="side">
<div class="brand"><span class="mk">훈</span><span>주간 교육실적 취합</span></div>
{side_nav(sess, active)}
<div class="me"><span class="av">{아바타}</span>
<span class="who"><b>{e(sess["id"])}</b><span>{e(역할)}</span></span>
<a href="/view?to=m" title="휴대폰용 화면">{icon("phone")}</a>
<a href="/logout" title="로그아웃">{icon("logout")}</a></div>
</aside>
<div class="main"><div class="wrap">{body}</div></div></div>
{CHAT_WIDGET}</body></html>"""


# ── 모바일 전용 화면 ─────────────────────────────────────────────

MOBILE_UA = ("iphone", "android", "ipad", "ipod", "windows phone", "mobile")


def is_mobile(headers):
    """휴대폰 접속인지. 쿠키(view=pc|m)가 있으면 그 선택이 우선한다."""
    for chunk in (headers.get("Cookie") or "").split(";"):
        k, _, v = chunk.strip().partition("=")
        if k == "view":
            return v == "m"
    ua = (headers.get("User-Agent") or "").lower()
    return any(x in ua for x in MOBILE_UA)


def m_page(title, body, sess):
    """모바일 전용 뼈대. 가로 메뉴 대신 큰 버튼과 세로 카드만 쓴다."""
    who = f'{e(sess["id"])} · {"관리자" if sess["role"] == "admin" else e(sess["org"] or "기관")}'
    return f"""<!doctype html><html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>{e(title)}</title><style>{STYLE}</style></head><body>
<div class="mtop"><span class="mk">훈</span><b>주간 교육실적 취합</b><span class="who">{who}</span></div>
<div class="wrap" style="max-width:560px">{body}
<p class="hint" style="margin-top:22px;text-align:center">
<a href="/view?to=pc">PC 화면으로 보기</a> · <a href="/logout">로그아웃</a></p></div>
{CHAT_WIDGET}</body></html>"""


def m_admin(sess):
    """관리자 모바일 홈 — 이번 주 상태와 오늘 할 일만 세로로."""
    rep, week, _, sample = week_report(None, None, "")
    if rep is None:
        return m_page("모바일", '<div class="card"><p class="note warn">아직 자료가 없습니다. PC 화면에서 제출을 먼저 받아 주세요.</p></div>', sess)

    t = rep["실적"]["합계_전체"][0]
    board = rep["현황판"]
    승인 = sum(1 for b in board if b["상태"] == "승인")
    막힘 = [b for b in board if b["상태"] in ("미제출", "반려 대상")]
    todos = rep["오늘할일"]
    급함 = [x for x in todos if x["우선순위"] == "높음"]

    지표 = f"""<div class="card"><div class="mrow"><b>{week_label(week)}</b>
<span class="dim">{"연습용 샘플" if sample else "저장소 자료"}</span></div>
<div class="mrow" style="margin-top:10px"><span>실시율(목표대비)</span><span class="big">{pct(t['실시율'])}</span></div>
<div class="mrow"><span>전체 탈락률</span><span class="big">{pct(t['탈락률'])}</span></div>
<div class="mrow"><span>제출·승인</span><span class="big">{승인}/{len(board)}</span></div></div>"""

    할일 = "".join(
        f'<li><span class="tag {"t-rej" if x["우선순위"] == "높음" else "t-wait"}">{e(x["우선순위"])}</span> '
        f'<b>{e(x["대상"])}</b><br><span class="dim">{e(x["할일"])}</span></li>'
        for x in todos[:8]
    )
    할일카드 = f"""<div class="card"><b>오늘 할 일 {len(todos)}건</b>
<p class="hint" style="margin:2px 0 8px">먼저 볼 것 {len(급함)}건{" · 8건까지만 표시" if len(todos) > 8 else ""}</p>
<ul class="mlist">{할일 or "<li>처리할 항목이 없습니다.</li>"}</ul></div>"""

    막힘카드 = (
        f'<div class="card"><b>발행 보류 {len(막힘)}곳</b><ul class="mlist">'
        + "".join(f'<li>{e(b["기관명"])} <span class="dim">{e(b["상태"])}</span></li>' for b in 막힘)
        + "</ul></div>"
        if 막힘
        else '<div class="card"><b>발행 가능</b><p class="hint" style="margin:2px 0 0">미제출·반려가 없습니다.</p></div>'
    )

    return m_page(
        "모바일 요약",
        f"""<h1 style="font-size:19px">오늘 확인할 것</h1>
{지표}{할일카드}{막힘카드}
<div class="mnav">
<a href="/brief?week={quote(week)}">핵심 요약</a>
<a href="/report?week={quote(week)}">주간 리포트</a>
<a href="/admin?week={quote(week)}" class="ghost">취합·승인</a>
<a href="/calendar" class="ghost">캘린더</a>
<a href="/advice" class="ghost">종합 제언</a></div>""",
        sess,
    )


def m_org(sess):
    """기관 모바일 홈 — 내 제출 상태와 제출 버튼."""
    org = sess["org"]
    조사 = calendar_store.open_for(org, storage.load_roster())
    items = ""
    for k in storage.list_weeks()[:6]:
        s = (storage.load_week(k) or {}).get("제출", {}).get(org)
        if not s:
            continue
        tags = {"승인": "t-ok", "제출": "t-wait", "반려": "t-rej"}
        items += (
            f'<li><div class="mrow"><b>{week_label(k)}</b>'
            f'<span class="tag {tags.get(s["상태"], "t-wait")}">{e(s["상태"])}</span></div>'
            f'<span class="dim">과정 {len(s["실적"])}개 · {e(s["제출시각"])}</span></li>'
        )
    return m_page(
        "내 제출",
        f"""<h1 style="font-size:19px">내 제출 내역</h1>
<p class="sub">{e(org)} · 최근 6개 주차</p>
{f'<div class="card"><b>응답할 일정 조사 {len(조사)}건</b><ul class="mlist">' + "".join(f'<li><a href="/calendar?id={x["번호"]}">{e(x["제목"])}</a><br><span class="dim">{e(x["시작일"])} ~ {e(x["종료일"])}</span></li>' for x in 조사) + "</ul></div>" if 조사 else ""}
<div class="card"><ul class="mlist">{items or "<li>아직 제출한 내역이 없습니다.</li>"}</ul></div>
<div class="mnav"><a href="/submit">실적 입력하기</a>
<a href="/upload" class="ghost">엑셀로 올리기</a>
<a href="/calendar" class="ghost">캘린더</a></div>
<p class="hint">반려된 주차는 수정해서 다시 제출하면 상태가 '제출' 로 돌아갑니다.</p>""",
        sess,
    )


# ── 화면: 로그인 ─────────────────────────────────────────────────


def login_page(msg="", user_id=""):
    return f"""<!doctype html><html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>로그인</title><style>{STYLE}</style></head><body>
<div class="wrap" style="max-width:420px;padding-top:70px">
<h1 style="text-align:center">주간 훈련기관 교육실적 취합</h1>
<p class="sub" style="text-align:center">관리자가 발급한 아이디로 접속하세요.</p>
<form class="card" method="post" action="/login">
{msg}
<label>아이디</label><input type="text" name="id" required autofocus value="{e(user_id)}">
<label>비밀번호</label><input type="password" name="pw" required>
<p style="margin-top:20px"><button type="submit" style="width:100%">로그인</button></p>
</form>
<p class="hint" style="text-align:center">계정 발급은 관리자에게 요청하세요. 자율 가입은 없습니다.</p>
</div></body></html>"""


# ── 화면: 계정 관리 (관리자) ─────────────────────────────────────


def users_page(sess, msg=""):
    users = auth.load_users()
    roster = storage.load_roster()
    rows = ""
    for uid, u in users.items():
        role = "관리자" if u["권한"] == "admin" else "기관"
        tag = "t-ok" if u["권한"] == "admin" else "t-wait"
        rows += f"""<tr><td><b>{e(uid)}</b></td><td><span class="tag {tag}">{role}</span></td>
<td>{e(u.get("기관명")) or '<span class="dim">—</span>'}</td><td class="dim">{e(u.get("발급시각",""))}</td>
<td class="noprint"><form method="post" action="/users" style="display:flex;gap:5px;align-items:center">
<input type="hidden" name="do" value="reset"><input type="hidden" name="id" value="{e(uid)}">
<input type="text" name="pw" placeholder="새 비밀번호" style="width:130px;padding:4px 8px;font-size:12px" required>
<button class="sm ghost">변경</button></form></td>
<td class="noprint"><form method="post" action="/users" onsubmit="return confirm('{e(uid)} 계정을 삭제할까요?')">
<input type="hidden" name="do" value="delete"><input type="hidden" name="id" value="{e(uid)}">
<button class="sm danger">삭제</button></form></td></tr>"""

    opts = "".join(f"<option>{e(n)}</option>" for n in roster)
    org_field = (
        f'<select name="org"><option value="">— 기관 선택 —</option>{opts}</select>'
        if roster
        else '<input type="text" name="org" placeholder="가온직업전문학교">'
    )
    return page(
        "계정 관리",
        f"""<h1>계정 관리</h1>
<p class="sub">관리자가 아이디·비밀번호를 발급해 각 기관에 배포합니다. 자율 가입은 없습니다.</p>
{msg}
<div class="card"><div class="hd"><h2>발급된 계정 {len(users)}개</h2></div>
<div class="scroll"><table><thead><tr><th>아이디</th><th>권한</th><th>소속 기관</th><th>발급</th>
<th class="noprint">비밀번호 변경</th><th class="noprint"></th></tr></thead><tbody>{rows}</tbody></table></div>
<p class="note">비밀번호는 해시로만 저장되어 <b>다시 볼 수 없습니다.</b> 잊었다면 여기서 새로 지정해 전달하세요.</p></div>

<div class="card"><div class="hd"><h2>새 계정 발급</h2></div>
<form method="post" action="/users"><input type="hidden" name="do" value="create">
<div class="inline">
<div><label>아이디 *</label><input type="text" name="id" required placeholder="gaon"></div>
<div><label>비밀번호 *</label><input type="text" name="pw" required placeholder="전달할 초기 비밀번호"></div>
<div style="max-width:150px"><label>권한 *</label><select name="role"><option value="org">기관</option><option value="admin">관리자</option></select></div>
<div><label>소속 기관 <span class="hint" style="font-weight:400">(기관 계정만)</span></label>{org_field}</div>
</div><p style="margin-top:18px"><button type="submit">발급</button></p></form>
<p class="note">기관 계정은 <b>자기 기관 제출·조회만</b> 할 수 있고, 취합·승인·리포트 화면에는 들어올 수 없습니다.</p></div>""",
        "users",
        sess,
    )


# ── 화면: 경북 산업·직업훈련 이슈 ────────────────────────────────


def issues_page(sess, msg=""):
    cache = gb_issues.load_cache()
    blocks = ""
    for block in gb_issues.ordered(cache):
        items = ""
        출처있음 = 0
        for it in block["이슈"]:
            links = "".join(
                f'<a href="{e(s["url"])}" target="_blank" rel="noopener noreferrer">{e(s["제목"])}</a>'
                for s in it["출처"]
            )
            출처있음 += bool(links)
            links = (f'<div class="srcs">{links}</div>' if links
                     else '<div class="srcs"><span class="tag t-rej">출처 없음 — 확인 전 사용 금지</span></div>')
            items += f'<li>{e(it["내용"])}{links}</li>'
        # 한 섹터가 통째로 출처 없이 오면 검색이 돌지 않고 모델이 지어낸 것이다. 그대로 쓰면 안 된다.
        경고 = ('<p class="note warn">이 섹터는 <b>모든 이슈에 출처가 없습니다.</b> 웹 검색이 돌지 않고 '
                '모델이 지어냈을 수 있으니 쓰지 마시고, 이 섹터만 다시 수집해 주세요.</p>'
                if block["이슈"] and not 출처있음 else "")
        blocks += f"""<h3>{e(block["주제"])} <span class="dim" style="font-weight:400">({len(block["이슈"])}건 ·
출처 있음 {출처있음}건 · {e(block["수집시각"])})</span></h3>
{경고}<ul class="notes">{items}</ul>"""

    if not blocks:
        blocks = '<p class="note">아직 수집한 이슈가 없습니다. 위 버튼으로 수집을 시작하세요.</p>'

    topics = ", ".join(t for t, _ in gb_issues.TOPICS)
    last = gb_issues.last_updated(cache)
    return page(
        "경북 산업·직업훈련 이슈",
        phead("참고자료", "경북 산업·직업훈련 이슈",
              f"경상북도 주력 산업(섹터)과 직업훈련 분야에서 뉴스에 자주 언급되는 이슈를 "
              f"섹터마다 {gb_issues.PER_TOPIC}건씩 찾아 출처 링크와 함께 모읍니다.")
        + f"""
{msg}
<div class="card"><div class="hd"><h2>섹터 {len(gb_issues.TOPICS)}개 × {gb_issues.PER_TOPIC}건 · 최종 수집 {e(last or "없음")}</h2>
<p class="sub" style="margin:0">모델 <code>{e(gb_issues.MODEL)}</code> · {e(topics)}</p></div>
<form method="post" action="/issues" style="display:flex;gap:9px;flex-wrap:wrap;align-items:flex-end">
<div style="min-width:170px"><label>섹터</label>
<select name="topic"><option value="">전체 ({len(gb_issues.TOPICS)}개 섹터)</option>
{"".join(f'<option value="{e(t)}">{e(t)}</option>' for t, _ in gb_issues.TOPICS)}</select></div>
<button type="submit"
 onclick="return confirm('웹 검색을 새로 돌립니다. API 요금이 발생합니다. 계속할까요?')">새로 수집</button>
</form>
<p class="hint" style="margin-top:10px">수집 결과는 <code>data/gb_issues.json</code> 에 저장되고, 버튼을 누를 때만 다시 검색합니다.
링크는 검색 결과에 실제로 딸려 온 주소만 씁니다 — 지어낸 주소가 섞이지 않게 한 것입니다.</p></div>

<div class="card"><div class="hd"><h2>수집된 이슈</h2>
<p class="sub" style="margin:0">각 이슈 아래의 링크로 원문을 확인하세요. 보고서에 쓰기 전 반드시 원문을 한 번 보십시오.</p></div>
{blocks}</div>""",
        "issues",
        sess,
    )


# ── 화면: 고용 통계 (KOSIS) ──────────────────────────────────────

# 색은 dataviz 팔레트 검증을 통과한 4개만 쓴다 (돌려쓰지 않는다).
SERIES_COLORS = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100"]


def line_chart(periods, series, unit, color_from=0, legend_box=True, height=250):
    """시계열 선 그래프(SVG). series = {지역명: [값|None, ...]}, 길이는 periods 와 같다."""
    vals = [v for vs in series.values() for v in vs if v is not None]
    if not vals:
        return '<p class="note">그릴 값이 없습니다.</p>'

    W, H, L, R, T, B = 760, height, 54, 112, 14, 28
    lo, hi = min(vals), max(vals)
    if hi == lo:
        lo, hi = lo - 1, hi + 1
    pad = (hi - lo) * 0.12
    lo, hi = lo - pad, hi + pad
    if min(vals) >= 0:
        lo = max(lo, 0)  # 인구·비율은 음수가 없다. 축에 음수 눈금을 만들지 않는다.
    n = len(periods)

    def x(i):
        return L + (W - L - R) * (i / (n - 1) if n > 1 else 0.5)

    def y(v):
        return T + (H - T - B) * (1 - (v - lo) / (hi - lo))

    def fmt(v):
        return f"{v:,.1f}" if abs(v) < 1000 else f"{v:,.0f}"

    # 눈금 (가로선 5개) — 축·격자는 뒤로 물러나게
    grid = ""
    for k in range(5):
        v = lo + (hi - lo) * k / 4
        gy = y(v)
        grid += (
            f'<line x1="{L}" y1="{gy:.1f}" x2="{W - R}" y2="{gy:.1f}" stroke="#eef0f3" stroke-width="1"/>'
            f'<text x="{L - 8}" y="{gy + 4:.1f}" text-anchor="end" font-size="11" fill="#9ca3af">{fmt(v)}</text>'
        )

    # 시점 라벨 6개쯤
    step = max(1, n // 6)
    for i in range(0, n, step):
        grid += (
            f'<text x="{x(i):.1f}" y="{H - 8}" text-anchor="middle" font-size="11" fill="#9ca3af">'
            f"{e(kosis_stats.fmt_period(periods[i]))}</text>"
        )

    lines, legend, ends = "", "", []
    for idx, (name, values) in enumerate(series.items()):
        color = SERIES_COLORS[(idx + color_from) % len(SERIES_COLORS)]
        d, pen_down = "", False
        for i, v in enumerate(values):
            if v is None:  # 결측은 선을 끊는다 (0 으로 잇지 않는다)
                pen_down = False
                continue
            d += f"{'L' if pen_down else 'M'}{x(i):.1f},{y(v):.1f} "
            pen_down = True
        lines += f'<path d="{d.strip()}" fill="none" stroke="{color}" stroke-width="2" stroke-linejoin="round"/>'

        last = next(((i, v) for i, v in reversed(list(enumerate(values))) if v is not None), None)
        if last:
            i, v = last
            ends.append({"x": x(i), "y": y(v), "글자y": y(v), "색": color, "글": f"{name} {fmt(v)}"})
        legend += (
            f'<span style="display:inline-flex;align-items:center;gap:5px">'
            f'<span style="width:11px;height:11px;border-radius:3px;background:{color}"></span>{e(name)}</span>'
        )

    # 선 끝 라벨: 값이 가까우면 글자가 겹치므로 위에서부터 최소 간격만큼 밀어낸다.
    labels = ""
    ends.sort(key=lambda d: d["글자y"])
    바닥 = T + 6
    for end in ends:
        end["글자y"] = 바닥 = max(end["글자y"], 바닥)
        바닥 += 14
    for end in ends:
        연결 = ""
        if abs(end["글자y"] - end["y"]) > 2:  # 점에서 밀려난 만큼 가는 선으로 이어 준다
            연결 = (
                f'<line x1="{end["x"] + 5:.1f}" y1="{end["y"]:.1f}" x2="{end["x"] + 9:.1f}"'
                f' y2="{end["글자y"]:.1f}" stroke="{end["색"]}" stroke-width="1" opacity="0.5"/>'
            )
        labels += (
            f'<circle cx="{end["x"]:.1f}" cy="{end["y"]:.1f}" r="4" fill="{end["색"]}" stroke="#fff" stroke-width="2"/>'
            f'{연결}<text x="{end["x"] + 11:.1f}" y="{end["글자y"] + 4:.1f}" font-size="11" fill="#374151">'
            f'{e(end["글"])}</text>'
        )

    # 마우스를 올린 시점의 값을 전부 보여주는 말풍선. 투명 띠 위에 CSS(:hover)로만 켠다.
    hover = ""
    band = (W - L - R) / max(1, n - 1) if n > 1 else W - L - R
    for i, period in enumerate(periods):
        머리 = kosis_stats.fmt_period(period)
        붙임 = "" if unit == "%" else " "  # '65.1 %' 처럼 뜨지 않게
        줄 = [(nm, f"{fmt(vs[i])}{붙임}{unit}" if vs[i] is not None else "값 없음",
               SERIES_COLORS[(k + color_from) % len(SERIES_COLORS)])
              for k, (nm, vs) in enumerate(series.items())]
        폭 = max([len(머리)] + [len(nm) + len(v) + 2 for nm, v, _ in 줄]) * 7 + 26
        높이 = 20 + 15 * len(줄) + 8
        왼쪽 = x(i) > (L + W - R) / 2  # 오른쪽 절반이면 말풍선을 왼쪽에 띄운다
        bx = x(i) - 폭 - 10 if 왼쪽 else x(i) + 10
        bx = min(max(bx, 4), W - 폭 - 4)
        by = min(max(T, T + 6), H - B - 높이)

        점 = "".join(
            f'<circle cx="{x(i):.1f}" cy="{y(vs[i]):.1f}" r="4" fill="{SERIES_COLORS[(k + color_from) % len(SERIES_COLORS)]}"'
            ' stroke="#fff" stroke-width="2"/>'
            for k, vs in enumerate(series.values()) if vs[i] is not None
        )
        글 = f'<text x="{bx + 10:.1f}" y="{by + 16:.1f}" font-size="11" font-weight="700" fill="#111827">{e(머리)}</text>'
        for k, (nm, v, color) in enumerate(줄):
            gy = by + 33 + 15 * k
            글 += (
                f'<rect x="{bx + 10:.1f}" y="{gy - 8:.1f}" width="8" height="8" rx="2" fill="{color}"/>'
                f'<text x="{bx + 23:.1f}" y="{gy:.1f}" font-size="11" fill="#374151">{e(nm)} {e(v)}</text>'
            )
        hover += (
            f'<g class="hb"><rect x="{x(i) - band / 2:.1f}" y="{T}" width="{band:.1f}" height="{H - T - B}" fill="transparent"/>'
            f'<g class="tip"><line x1="{x(i):.1f}" y1="{T}" x2="{x(i):.1f}" y2="{H - B}" stroke="#9ca3af"'
            f' stroke-width="1" stroke-dasharray="3 3"/>{점}'
            f'<rect x="{bx:.1f}" y="{by:.1f}" width="{폭}" height="{높이}" rx="7" fill="#fff" stroke="#d1d5db"/>{글}</g></g>'
        )

    # 계열이 하나면 제목이 곧 범례라 범례 상자를 두지 않는다.
    상단 = f'<div class="legend">{legend}<span class="dim">단위 {e(unit)}</span></div>' if legend_box else ""
    return f"""{상단}<div class="scroll"><svg viewBox="0 0 {W} {H}" width="100%" height="{H}" role="img"
 aria-label="시계열 선 그래프">{grid}{lines}{labels}{hover}</svg></div>
<p class="hint noprint">선 위(같은 세로 구간)에 마우스를 올리면 그 시점의 값이 모두 표시됩니다.</p>"""


def chart_block(periods, series, unit):
    """한 그래프에 넣을지, 지역별로 쪼갤지 고른다.

    지역 간 값 차이가 너무 크면(생산가능인구의 전국 4.6천만 vs 시도 200만대) 한 축에
    같이 그릴 때 작은 쪽이 바닥에 붙어 증감이 안 보인다. 그럴 땐 지역별 작은 그래프로 나눈다.
    """
    최대값 = [max((v for v in vs if v is not None), default=0) for vs in series.values()]
    최대값 = [m for m in 최대값 if m > 0]
    쪼갬 = len(series) > 1 and 최대값 and max(최대값) / min(최대값) > 4

    if not 쪼갬:
        return line_chart(periods, series, unit)

    blocks = ""
    for idx, (name, values) in enumerate(series.items()):
        blocks += (
            f'<div><h3 style="margin:14px 0 0">{e(name)}</h3>'
            + line_chart(periods, {name: values}, unit, color_from=idx, legend_box=False, height=170)
            + "</div>"
        )
    return (
        '<p class="hint">지역 간 규모 차이가 커서 한 축에 같이 그리면 작은 쪽 변화가 보이지 않습니다. '
        f"지역별로 나눠 그렸습니다 (단위 {e(unit)}).</p>{blocks}"
    )


def stats_table(periods, series, unit):
    """그래프 아래에 같이 두는 수치 표. 접지 않고 항상 펼쳐 둔다.

    그래프의 각 선 색을 표의 지역명 앞에 같이 찍어 어느 선인지 눈으로 잇게 한다.
    """
    head = "".join(f'<th class="n">{e(kosis_stats.fmt_period(p))}</th>' for p in periods)
    rows = ""
    for idx, (nm, vs) in enumerate(series.items()):
        color = SERIES_COLORS[idx % len(SERIES_COLORS)]
        칩 = f'<span style="display:inline-block;width:9px;height:9px;border-radius:2px;background:{color};margin-right:6px"></span>'
        rows += (
            f"<tr><td>{칩}<b>{e(nm)}</b></td>"
            + "".join(f'<td class="n">{"-" if v is None else f"{v:,.1f}"}</td>' for v in vs)
            + "</tr>"
        )
    return f"""<h3>수치 표 <span class="dim" style="font-weight:400;font-size:12px">단위 {e(unit)} · 그래프와 같은 값</span></h3>
<div class="scroll"><table><thead><tr><th>지역</th>{head}</tr></thead><tbody>{rows}</tbody></table></div>"""


def stats_page(sess, msg=""):
    data = kosis_stats.load_cache()
    picked = list(data.get("지표", {}).get("고용률", {}).get("지역", {})) or kosis_stats.DEFAULT_REGIONS

    charts = ""
    for label, block in data.get("지표", {}).items():
        unit = block["단위"]
        charts += f"""<div class="card"><div class="hd"><h2>{e(label)}</h2>
<p class="sub" style="margin:0">전국·시도별 월간 추이 · 단위 {e(unit)}</p></div>
{chart_block(data["시점"], block["지역"], unit)}
{stats_table(data["시점"], block["지역"], unit)}</div>"""

    if not charts:
        charts = '<div class="card"><p class="note">아직 받아온 통계가 없습니다. 위에서 조건을 고르고 불러오세요.</p></div>'

    checks = "".join(
        f'<label style="display:inline-flex;gap:5px;align-items:center;font-weight:400;margin:0 12px 4px 0">'
        f'<input type="checkbox" name="region" value="{e(r)}" {"checked" if r in picked else ""}>{e(r)}</label>'
        for r in kosis_stats.ALL_REGIONS
    )
    months = "".join(
        f'<option value="{m}" {"selected" if m == len(data.get("시점", [])) else ""}>최근 {m}개월</option>'
        for m in (12, 24, 36, 60)
    )
    src = data.get("출처", {})
    받은기간 = (
        f'{kosis_stats.fmt_period(data["시점"][0])} ~ {kosis_stats.fmt_period(data["시점"][-1])}'
        if data.get("시점")
        else "없음"
    )

    return page(
        "고용 통계",
        phead("참고자료", "고용 통계",
              "전국·시도별 고용률, 실업률, 생산가능인구(15세이상인구)를 월별로 봅니다. 출처는 KOSIS 국가통계입니다.")
        + f"""
{msg}
<div class="card"><div class="hd"><h2>불러올 조건</h2>
<p class="sub" style="margin:0">받은 기간 {e(받은기간)} · 최종 수집 {e(data.get("수집시각", "없음"))}</p></div>
<form method="post" action="/stats">
<label>지역 <span class="dim" style="font-weight:400">(최대 {kosis_stats.MAX_SERIES}개 — 선 색이 겹치지 않게 제한합니다)</span></label>
<div style="margin:4px 0 12px">{checks}</div>
<div class="inline"><div style="max-width:200px"><label>기간</label><select name="months">{months}</select></div>
<div><button type="submit">KOSIS 에서 불러오기</button></div></div>
</form>
<p class="hint" style="margin-top:10px">받은 값은 <code>data/kosis_stats.json</code> 에 저장되고, 버튼을 누를 때만 다시 받습니다.</p></div>
{charts}
{f'<p class="hint">출처: <a href="{e(src["url"])}" target="_blank" rel="noopener noreferrer">{e(src["이름"])}</a></p>' if src else ""}""",
        "stats",
        sess,
    )


# ── 화면: 종합 제언 (경북 이슈 + 고용 통계) ──────────────────────
#
# 이 화면의 문장은 「경북 이슈」와 「고용 통계」 두 자료에서만 뽑는다.
# 주차 실적은 쓰지 않는다 — 실적 판단은 ③ 주간 리포트 쪽 몫이다.

ADVICE_TOPICS = [
    ("산업", "경북 산업 섹터에서 지금 움직이는 것 — 「경북 이슈」 수집분"),
    ("고용", "고용률·실업률·생산가능인구 — 「고용 통계」 수집분"),
    ("직업훈련", "위 두 자료에서 읽히는 훈련 수요"),
]
훈련섹터 = "직업훈련·고용"  # gb_issues.TOPICS 중 산업이 아닌 주제
확장신호 = ("투자", "증설", "신설", "착공", "준공", "유치", "협약", "채용", "공장", "일자리", "인력")
조정신호 = ("전환", "감산", "침체", "위기", "폐업", "철수", "감원", "구조조정", "부진", "감소")

# 낱말 빈도용 — 형태소 분석 없이 자주 붙는 조사만 떼고 센다(긴 것부터).
JOSA = ("으로써", "에서는", "에게는", "으로는", "까지도", "으로", "에서", "에게", "까지", "부터",
        "보다", "처럼", "마다", "이나", "와의", "과의")
짧은조사 = ("의", "을", "를", "이", "가", "은", "는", "에", "도", "와", "과", "로", "만")
STOPWORDS = {
    "경북", "경상북", "경상북도", "지역", "지난", "최근", "올해", "내년", "작년", "이번", "현재",
    "관련", "대한", "통해", "위해", "따라", "대비", "규모", "계획", "예정", "발표", "추진", "진행",
    "마련", "실시", "개최", "밝혔", "밝혔다", "했다", "한다", "있다", "있는", "이다", "된다",
    "전망", "기대", "그리고", "또한", "하지만", "이후", "오는", "지난해", "가운데", "대해", "위한",
    "산업", "사업", "경우", "최대", "최소", "이상", "이하", "각각", "포함", "가장", "모두", "다양한",
}
WORD_RE = re.compile(r"[가-힣A-Za-z][가-힣A-Za-z0-9]+")


def stem(w):
    """낱말 끝의 조사를 뗀다. 떼고 나서 두 글자가 안 되면 세지 않는다(None)."""
    for j in JOSA:
        if w.endswith(j):
            return w[: -len(j)] if len(w) - len(j) >= 2 else None
    if w[-1] in 짧은조사:
        # '원을'·'일부터' 처럼 한 글자 낱말에 조사가 붙은 것은 뜻이 없어 버린다.
        return w[:-1] if len(w) >= 3 else None
    return w


def keywords(texts, n=6):
    """이슈 본문에서 자주 나온 낱말 상위 n개. 두 번 이상 나온 것만.

    형태소 분석 없이 조사만 떼고 세는 기계적 빈도다 — 정확한 분석이 아니라
    '어떤 말이 반복해서 나오는지' 를 보는 용도로만 쓴다.
    """
    cnt = Counter()
    for t in texts:
        for w in WORD_RE.findall(t):
            if w.endswith("니다") or w.endswith("했다") or w.endswith("된다"):
                continue  # 서술어는 빈도가 높아도 뜻을 주지 않는다
            w = stem(w)
            if not w or len(w) < 2 or w in STOPWORDS:
                continue
            if w[-1] in ("하", "되"):
                continue  # '지원하(는)' 처럼 어간만 남은 것은 버린다
            cnt[w] += 1
    return [(w, c) for w, c in cnt.most_common(n) if c > 1]


def kw_text(texts, n=6):
    kws = keywords(texts, n)
    return ", ".join(f"{w}({c}회)" for w, c in kws) if kws else ""


def 이가(w):
    """섹터 이름 뒤에 붙일 '이/가' 를 받침 보고 고른다."""
    ch = w[-1] if w else ""
    return "이" if "가" <= ch <= "힣" and (ord(ch) - 0xAC00) % 28 else "가"


def issue_texts(topics):
    return [it["내용"] for t in topics for it in t.get("이슈", [])]


def signal_count(texts, words):
    """신호어가 하나라도 들어간 이슈 수."""
    return sum(1 for t in texts if any(w in t for w in words))


def series_facts(cache):
    """고용 통계 캐시에서 지표별 (최근값, 직전값, 1년 전 값, 시점)을 뽑는다."""
    out = {}
    시점 = cache.get("시점") or []
    for name, block in (cache.get("지표") or {}).items():
        for region, vs in (block.get("지역") or {}).items():
            값 = [(i, v) for i, v in enumerate(vs) if v is not None]
            if not 값:
                continue
            i, 최근 = 값[-1]
            직전 = 값[-2][1] if len(값) > 1 else None
            일년 = next((v for j, v in 값 if j == i - 12), None)
            out[name] = {"지역": region, "단위": block.get("단위", ""), "최근": 최근, "직전": 직전,
                         "일년": 일년, "시점": kosis_stats.fmt_period(시점[i]) if i < len(시점) else ""}
            break  # 지역이 여럿이면 첫 지역(기본 경상북도) 기준으로 본다
    return out


def move(cur, base):
    """증감 문구. 비교값이 없으면 None."""
    if cur is None or base is None:
        return None
    d = cur - base
    return f"{'▲' if d > 0 else ('▼' if d < 0 else '―')} {abs(d):.1f}%p"


def 방향(f):
    """비교 기준을 1년 전 → 직전 달 순으로 찾는다. 둘 다 없으면 비교하지 않는다."""
    if f["일년"] is not None:
        return f["최근"] - f["일년"], "1년 전"
    if f["직전"] is not None:
        return f["최근"] - f["직전"], "직전 달"
    return None, None


def advice_blocks(facts, issues):
    """주제별 (근거, 중점, 전략) 목록. 근거가 없는 주제는 '자료 없음'을 그대로 남긴다."""
    고용, 실업, 인구 = facts.get("고용률"), facts.get("실업률"), facts.get("생산가능인구")
    산업섹터 = [t for t in issues if t.get("주제") != 훈련섹터]
    훈련이슈 = [t for t in issues if t.get("주제") == 훈련섹터]
    산업문장, 훈련문장 = issue_texts(산업섹터), issue_texts(훈련이슈)
    확장수, 조정수 = signal_count(산업문장, 확장신호), signal_count(산업문장, 조정신호)
    blocks = {k: {"근거": [], "중점": [], "전략": []} for k, _ in ADVICE_TOPICS}

    # ── 산업 (경북 이슈)
    if 산업문장:
        for t in sorted(산업섹터, key=lambda x: -len(x.get("이슈", []))):
            말 = kw_text(issue_texts([t]), 4)
            blocks["산업"]["근거"].append(
                f"[{t['주제']}] 이슈 {len(t.get('이슈', []))}건" + (f" · 자주 나온 말: {말}" if 말 else "")
            )
        전체말 = kw_text(산업문장, 8)
        if 전체말:
            blocks["산업"]["근거"].append(f"산업 이슈 전체에서 자주 나온 말 — {전체말}")

        최다 = max(산업섹터, key=lambda x: len(x.get("이슈", [])))
        blocks["산업"]["중점"].append(
            f"수집된 산업 이슈 {len(산업문장)}건 가운데 '{최다['주제']}'{이가(최다['주제'])} "
            f"{len(최다.get('이슈', []))}건으로 가장 많이 잡혔습니다 — "
            "인력 수요 변화를 가장 먼저 확인할 섹터입니다."
        )
        if 확장수:
            blocks["산업"]["중점"].append(
                f"투자·증설·채용 같은 확장 표현이 들어간 이슈가 {확장수}건입니다 — 신규 인력 수요가 걸린 건입니다."
            )
        if 조정수:
            blocks["산업"]["중점"].append(
                f"전환·감산·부진 같은 조정 표현이 들어간 이슈가 {조정수}건입니다 — 재직자 전환 수요가 걸린 건입니다."
            )
        if 확장수 >= 조정수:
            blocks["산업"]["전략"].append(
                "확장 신호가 우세하므로, 해당 섹터의 채용 규모·시점이 원문에 적힌 건만 골라 신규 양성 과정 후보로 올립니다."
            )
        if 조정수:
            blocks["산업"]["전략"].append(
                "조정 신호가 잡힌 섹터는 신규 양성보다 재직자 향상·전직 훈련에 무게를 둡니다."
            )
        blocks["산업"]["전략"].append(
            "자주 나온 말에 오른 업종·기업명을 훈련 직종 후보 목록으로 옮겨, 다음 분기 과정 개편 때 먼저 검토합니다."
        )
    else:
        blocks["산업"]["근거"].append("수집된 산업 이슈가 없습니다. 「경북 이슈」 화면에서 먼저 수집해 주세요.")

    # ── 고용 (KOSIS 통계)
    for f, 이름 in ((고용, "고용률"), (실업, "실업률")):
        if not f:
            continue
        mv = move(f["최근"], f["일년"]) or move(f["최근"], f["직전"])
        blocks["고용"]["근거"].append(
            f"{f['지역']} {이름} {f['최근']}{f['단위']} ({f['시점']} 기준"
            + (f", 1년 전 대비 {mv}" if f["일년"] is not None and mv else "") + ")"
        )
    if 인구:
        d, 기준 = 방향(인구)
        blocks["고용"]["근거"].append(
            f"{인구['지역']} 생산가능인구 {인구['최근']:,}{인구['단위']} ({인구['시점']} 기준"
            + (f", {기준} 대비 {d:+,.1f}{인구['단위']}" if d is not None else "") + ")"
        )

    if 고용:
        d, 기준 = 방향(고용)
        if d is None:
            blocks["고용"]["중점"].append("고용률 비교 시점이 없어 추세를 판단하지 않았습니다.")
            blocks["고용"]["전략"].append("「고용 통계」에서 기간을 넉넉히 받아 1년 전과 견줄 수 있게 한 뒤 다시 봅니다.")
        elif d < 0:
            blocks["고용"]["중점"].append(f"고용률이 {기준}보다 {abs(d):.1f}%p 낮습니다 — 취업 연계가 약해진 구간입니다.")
            blocks["고용"]["전략"].append("훈련 종료 후 취업 연계(채용 설명회·기업 매칭)를 과정 일정 안에 넣어 운영합니다.")
        else:
            blocks["고용"]["중점"].append(f"고용률이 {기준}보다 {d:.1f}%p 높습니다 — 급격한 악화는 확인되지 않습니다.")
            blocks["고용"]["전략"].append("물량을 늘리기보다 미충원 직종 쪽으로 자원을 옮깁니다.")
    if 실업:
        d, 기준 = 방향(실업)
        if d is not None and d > 0:
            blocks["고용"]["중점"].append(f"실업률이 {기준}보다 {d:.1f}%p 높습니다.")
            blocks["고용"]["전략"].append("단기 재취업 과정 비중을 늘리고 구직급여 수급자 안내를 강화합니다.")
        elif d is not None:
            blocks["고용"]["중점"].append(f"실업률은 {기준}보다 {abs(d):.1f}%p 낮습니다.")
    if 인구:
        d, 기준 = 방향(인구)
        if d is not None and d <= -0.05:  # 반올림해서 0으로 보일 만큼 작은 변화는 추세로 읽지 않는다
            blocks["고용"]["중점"].append(
                f"생산가능인구가 {기준} 대비 {abs(d):,.1f}{인구['단위']} 줄었습니다 — 훈련 대상 모수 자체가 줄어드는 방향입니다."
            )
            blocks["고용"]["전략"].append("모집 인원을 늘려 잡기보다, 중장년·여성 등 미참여 층을 겨냥한 과정 설계로 방향을 돌립니다.")
    if not blocks["고용"]["근거"]:
        blocks["고용"]["근거"].append("고용 통계 캐시가 비어 있습니다. 「고용 통계」 화면에서 먼저 받아 주세요.")

    # ── 직업훈련 (훈련·고용 섹터 이슈 + 위 두 자료의 연결)
    if 훈련문장:
        blocks["직업훈련"]["근거"].append(f"'{훈련섹터}' 섹터 이슈 {len(훈련문장)}건")
        말 = kw_text(훈련문장, 6)
        if 말:
            blocks["직업훈련"]["근거"].append(f"훈련·고용 이슈에서 자주 나온 말 — {말}")
        for t in 훈련이슈:
            for it in t.get("이슈", [])[:3]:
                blocks["직업훈련"]["근거"].append(it["내용"])
        blocks["직업훈련"]["중점"].append(
            f"훈련·고용 이슈 {len(훈련문장)}건은 사업 공고·인력양성 협약이 섞여 있습니다 — 우리 기관 과정과 겹치는 건을 먼저 가려냅니다."
        )
        blocks["직업훈련"]["전략"].append("훈련 관련 이슈는 원문에서 대상·기간·수행기관을 확인해, 중복되는 과정은 정원 조정으로 대응합니다.")
    else:
        blocks["직업훈련"]["근거"].append(f"'{훈련섹터}' 섹터에서 수집된 이슈가 없습니다.")

    if 확장수 or 조정수:
        blocks["직업훈련"]["중점"].append(
            f"산업 이슈에서 확장 신호 {확장수}건 · 조정 신호 {조정수}건이 잡혔습니다 — "
            + ("신규 양성 쪽 수요가 더 큽니다." if 확장수 > 조정수 else
               ("전환·향상 훈련 쪽 수요가 더 큽니다." if 조정수 > 확장수 else "두 방향의 수요가 비슷합니다."))
        )
    if 고용 and 실업:
        blocks["직업훈련"]["근거"].append(
            f"같은 기간 고용 지표 — 고용률 {고용['최근']}{고용['단위']} · 실업률 {실업['최근']}{실업['단위']} ({고용['시점']} 기준)"
        )
        blocks["직업훈련"]["전략"].append("이슈에서 고른 직종 후보를 고용 지표 추이와 같은 기간으로 놓고 봐서, 지표와 어긋나는 후보는 뒤로 미룹니다.")
    if not blocks["직업훈련"]["전략"]:
        blocks["직업훈련"]["전략"].append("근거 자료가 모자라 전략을 적지 않았습니다. 「경북 이슈」·「고용 통계」를 먼저 수집해 주세요.")
    return blocks


def advice_page(sess):
    stats = kosis_stats.load_cache()
    issues = gb_issues.ordered(gb_issues.load_cache())
    facts = series_facts(stats)
    blocks = advice_blocks(facts, issues)

    cards = ""
    for idx, (name, desc) in enumerate(ADVICE_TOPICS, 1):
        b = blocks[name]
        근거 = "".join(f"<li>{e(x)}</li>" for x in b["근거"]) or "<li>자료 없음</li>"
        중점 = "".join(f"<li>{e(x)}</li>" for x in b["중점"]) or "<li>근거 자료가 없어 적지 않았습니다.</li>"
        전략 = "".join(f"<li>{e(x)}</li>" for x in b["전략"]) or "<li>근거 자료가 없어 적지 않았습니다.</li>"
        cards += f"""<div class="card"><div class="hd"><h2>{idx}. {e(name)}</h2>
<p class="sub" style="margin:0">{e(desc)}</p></div>
<h3>근거 (수집된 자료)</h3><ul class="notes">{근거}</ul>
<h3>주요 중점사항</h3><ul class="notes">{중점}</ul>
<h3>향후 전략</h3><ul class="notes">{전략}</ul></div>"""

    이슈수 = sum(len(t.get("이슈", [])) for t in issues)
    수집 = " · ".join(
        x for x in [
            f"경북 이슈 {이슈수}건 (섹터 {len(issues)}개)" if issues else "",
            f"고용 통계 {e(stats.get('수집시각'))}" if stats.get("수집시각") else "",
        ] if x
    ) or "수집된 자료 없음"

    고용, 실업, 인구 = facts.get("고용률"), facts.get("실업률"), facts.get("생산가능인구")
    산업문장 = issue_texts([t for t in issues if t.get("주제") != 훈련섹터])
    확장수, 조정수 = signal_count(산업문장, 확장신호), signal_count(산업문장, 조정신호)

    종합 = []
    if 고용 and 실업:
        d, 기준 = 방향(고용)
        종합.append(
            f"{고용['지역']}는 {고용['시점']} 기준 고용률 {고용['최근']}{고용['단위']}, 실업률 {실업['최근']}{실업['단위']} 로"
            + (f", {기준} 대비 고용률이 {abs(d):.1f}%p {'낮습니다' if d < 0 else '높습니다'}." if d is not None else " 비교 시점이 없습니다.")
        )
    if 산업문장:
        종합.append(
            f"산업 쪽은 수집된 이슈 {len(산업문장)}건 가운데 확장 신호 {확장수}건 · 조정 신호 {조정수}건으로, "
            + ("신규 인력 수요 쪽 이야기가 더 많습니다." if 확장수 > 조정수 else
               ("전환·조정 쪽 이야기가 더 많습니다." if 조정수 > 확장수 else "두 방향이 비슷하게 섞여 있습니다."))
        )
    if 인구:
        d, 기준 = 방향(인구)
        if d is not None and d <= -0.05:
            종합.append(
                f"생산가능인구는 {기준} 대비 {abs(d):,.1f}{인구['단위']} 줄어, 인원을 늘려 채우는 방식은 한계가 있습니다."
            )
    종합.append(
        "따라서 과정 수를 늘리는 방향보다, 확장 신호가 뚜렷한 섹터로 정원을 옮기고 "
        "조정 신호가 잡힌 섹터는 재직자 전환 훈련으로 성격을 바꾸는 편이 낫습니다."
    )
    종합.append(
        "이 화면의 문장은 「경북 이슈」와 「고용 통계」 두 자료에서만 뽑은 것이고, 낱말 빈도는 형태소 분석 없이 "
        "기계적으로 센 값입니다. 지역 현장 사정은 반영되어 있지 않으니 정책 판단 전에 원문과 담당자 확인을 거치십시오."
    )

    return page(
        "종합 제언",
        phead("참고자료", "종합 제언",
              "「경북 이슈」와 「고용 통계」에서만 뽑아, 산업·고용·직업훈련 관점의 중점사항과 향후 전략으로 정리했습니다. "
              "주차 실적은 쓰지 않습니다 — 실적은 ③ 주간 리포트에서 봅니다.")
        + f"""
<div class="card noprint"><p class="note">근거 자료: {수집}. 자료를 새로 받으려면
<a href="/issues">경북 이슈</a> · <a href="/stats">고용 통계</a> 화면에서 갱신하세요.</p></div>
{cards}
<div class="card" style="border-color:#c7d2fe;background:#f8faff"><div class="hd"><h2>종합 의견</h2>
<p class="sub" style="margin:0">위 세 관점을 한데 놓고 본 결론입니다.</p></div>
<ul class="notes">{"".join(f"<li>{e(x)}</li>" for x in 종합)}</ul></div>""",
        "advice",
        sess,
    )


# ── 화면: 캘린더(일정 조사) ──────────────────────────────────────

CAL_JS = """
// 응답 표의 라디오를 한 번에 고른다 (전체 가능 / 전체 불가 / 전체 미정)
function calAll(v){
 var box=document.getElementById('calform');
 if(!box) return;
 box.querySelectorAll('input[type=radio][value="'+v+'"]').forEach(function(r){r.checked=true;});
}
"""

CAL_MARK = {"가능": '<span style="color:#166534;font-weight:700">○</span>',
            "불가": '<span style="color:#b91c1c;font-weight:700">✕</span>'}

CAL_HEAD = ("일", "월", "화", "수", "목", "금", "토")  # 달력은 일요일 시작
PICK_CLASS = {"가능": "ok", "불가": "no", "미정": "maybe"}


def cal_mark(v):
    return CAL_MARK.get(v, '<span class="dim">–</span>')


def month_grid(days, cell, td_class=""):
    """조사 기간을 월별 달력 표로 그린다.

    days = 조사에 포함된 날짜(ISO 문자열) 목록, cell(iso) = 그 날 칸에 넣을 HTML.
    기간에 없는 날은 회색으로 남겨 두어, 물어본 범위가 달력 위에서 바로 보이게 한다.
    """
    포함 = set(days)
    달 = []
    for iso in days:
        d = calendar_store.parse_day(iso)
        if d and (d.year, d.month) not in 달:
            달.append((d.year, d.month))
    if not 달:
        return ""

    def 요일칸(i):
        return "sun" if i == 0 else ("sat" if i == 6 else "")

    표들 = ""
    for y, m in 달:
        앞칸 = (date(y, m, 1).weekday() + 1) % 7  # 월요일=0 → 일요일 시작으로 옮긴다
        칸 = [None] * 앞칸 + [date(y, m, i + 1) for i in range(monthrange(y, m)[1])]
        칸 += [None] * (-len(칸) % 7)
        줄 = ""
        for w in range(0, len(칸), 7):
            tds = ""
            for i, d in enumerate(칸[w : w + 7]):
                if d is None:
                    tds += '<td class="out"></td>'
                    continue
                iso = d.isoformat()
                안에 = iso in 포함
                cls = " ".join(x for x in (요일칸(i), "" if 안에 else "off", td_class if 안에 else "") if x)
                tds += f'<td class="{cls}"><div class="dnum">{d.day}</div>{cell(iso) if 안에 else ""}</td>'
            줄 += f"<tr>{tds}</tr>"
        머리 = "".join(f'<th class="{요일칸(i)}">{h}</th>' for i, h in enumerate(CAL_HEAD))
        표들 += (f'<table class="cal"><caption>{y}년 {m}월</caption>'
                 f"<thead><tr>{머리}</tr></thead><tbody>{줄}</tbody></table>")
    return f'<div class="cals">{표들}</div>'


def cal_pick_grid(days, mine):
    """기관 응답용 달력 — 날짜 칸마다 가능·불가·미정 중 하나를 고른다."""
    def cell(iso):
        고름 = mine.get(iso, "미정")
        return '<div class="pick">' + "".join(
            f'<label class="{PICK_CLASS[v]}"><input type="radio" name="a_{e(iso)}" value="{v}"'
            f'{" checked" if 고름 == v else ""}><span>{v}</span></label>'
            for v in calendar_store.ANSWERS
        ) + "</div>"

    return month_grid(days, cell, "pickday")


def cal_count_grid(ev, days, 대상수):
    """응답 현황용 달력 — 날짜 칸마다 '가능' 인원을 보여준다."""
    표 = calendar_store.tally(ev)
    최다 = calendar_store.best_days(ev)

    def cell(iso):
        n = 표.get(iso, {}).get("가능", 0)
        불가 = 표.get(iso, {}).get("불가", 0)
        수 = (f'<span class="best">{n}</span>' if iso in 최다 else
              (f"<b>{n}</b>" if n else '<span class="zero">0</span>'))
        뒤 = f"/{대상수}" if 대상수 else ""
        return f'<div class="cntbox">가능 {수}{뒤}<br>불가 {불가}</div>'

    return month_grid(days, cell, "cnt")


def cal_state(ev, sess, roster=None):
    """목록에 붙일 상태 태그 — 마감/응답 여부."""
    if calendar_store.is_closed(ev):
        return '<span class="tag t-none">마감</span>'
    if sess["role"] == "admin":
        return '<span class="tag t-wait">진행 중</span>'
    if not calendar_store.is_target(ev, sess["org"], roster):
        return '<span class="tag t-none">대상 아님</span>'
    답함 = sess["org"] in ev.get("응답", {})
    return f'<span class="tag {"t-ok" if 답함 else "t-wait"}">{"응답 완료" if 답함 else "응답 필요"}</span>'


def cal_period(ev):
    return f'{e(ev["시작일"])} ~ {e(ev["종료일"])}'


def calendar_page(sess, msg=""):
    """일정 조사 목록. 관리자는 여기서 새 조사를 낸다."""
    roster = storage.load_roster()
    rows = ""
    for ev in calendar_store.list_events():
        if not calendar_store.can_open(ev, sess["role"], sess["org"], roster):
            continue  # 대상도 아니고 전체 공개도 아닌 조사는 보이지 않는다
        대상 = calendar_store.targets(ev, roster)
        답수 = len(ev.get("응답", {}))
        공개 = '<span class="eye">전체 공개</span>' if ev.get("공개") else '<span class="tag t-none">관리자만</span>'
        rows += (
            f'<tr><td><b>{e(ev["제목"])}</b><br><span class="dim" style="font-size:12px">{cal_period(ev)}'
            f'{" · 마감 " + e(ev["마감일"]) if ev.get("마감일") else ""}</span></td>'
            f'<td class="n">{답수}/{len(대상) or 답수}</td><td>{cal_state(ev, sess, roster)} {공개}</td>'
            f'<td><a href="/calendar?id={ev["번호"]}"><button class="sm ghost">'
            f'{"응답 보기" if sess["role"] == "admin" or not calendar_store.is_target(ev, sess["org"], roster) else "응답하기"}'
            f'</button></a></td></tr>'
        )
    table = (
        f'<div class="scroll"><table><thead><tr><th>제목 · 기간</th><th class="n">응답</th>'
        f'<th>상태</th><th></th></tr></thead><tbody>{rows}</tbody></table></div>'
        if rows
        else '<p class="note">아직 등록된 일정 조사가 없습니다.</p>'
    )

    만들기 = ""
    if sess["role"] == "admin":
        대상칸 = (
            '<div class="grid">'
            + "".join(
                f'<label style="font-weight:400;display:flex;gap:7px;align-items:center">'
                f'<input type="checkbox" name="target" value="{e(n)}" style="width:auto"> {e(n)}</label>'
                for n in roster
            )
            + "</div>"
            if roster
            else '<p class="hint">대상 명단이 비어 있습니다. 취합·승인 화면에서 명단을 등록하면 여기서 고를 수 있습니다.</p>'
        )
        오늘 = date.today().isoformat()
        만들기 = f"""<div class="card"><div class="hd"><h2>새 일정 조사</h2>
<p class="sub" style="margin:0">기간을 정해 물으면 기관이 날짜마다 가능·불가로 답합니다. 응답은 기본적으로 관리자만 봅니다.</p></div>
<form method="post" action="/calendar"><input type="hidden" name="do" value="create">
<div class="inline">
<div style="min-width:240px"><label>제목 *</label><input type="text" name="title" required placeholder="10월 2주차 워크숍 참여 가능일"></div>
<div style="max-width:180px"><label>시작일 *</label><input type="date" name="start" required value="{오늘}"></div>
<div style="max-width:180px"><label>종료일 *</label><input type="date" name="end" required value="{오늘}"></div>
<div style="max-width:180px"><label>응답 마감일</label><input type="date" name="due"></div>
<div style="max-width:170px"><label>공개 범위</label><select name="public">
<option value="N">관리자만 보기</option><option value="Y">전체 공개</option></select></div>
</div>
<label>안내 문구</label><input type="text" name="body" placeholder="워크숍 참석이 가능한 날짜를 모두 표시해 주세요.">
<label>대상 기관 <span class="hint" style="font-weight:400">(고르지 않으면 명단 전체)</span></label>{대상칸}
<p style="margin-top:18px"><button type="submit">조사 만들기</button></p></form>
<p class="note">기간은 최대 {calendar_store.MAX_DAYS}일까지 물을 수 있습니다. 공개 범위는 만든 뒤에도 바꿀 수 있습니다.</p></div>"""

    안내 = (
        "기관이 답한 가능·불가는 관리자만 봅니다. 조사마다 '전체 공개' 를 켜면 참여 기관끼리도 서로 확인할 수 있습니다."
        if sess["role"] == "admin"
        else "관리자가 요청한 기간에 가능한 날짜를 표시해 주세요. 내 응답은 관리자만 보며, 전체 공개된 조사만 다른 기관의 답이 함께 보입니다."
    )
    return page(
        "캘린더",
        phead("일정", "캘린더", 안내)
        + f"""
{msg}
<div class="card"><div class="hd"><h2>일정 조사 {len(calendar_store.list_events())}건</h2></div>{table}</div>
{만들기}""",
        "calendar",
        sess,
    )


def cal_answer_table(ev, roster):
    """기관 × 날짜 응답 표 (관리자 또는 전체 공개일 때만 보여준다)."""
    days = calendar_store.days_of(ev)
    응답 = ev.get("응답", {})
    빈칸 = '<span class="dim">—</span>'
    머리 = "".join(
        f'<th class="n">{e(calendar_store.day_label(d))}</th>' for d in days
    )
    본문 = ""
    for org, a in 응답.items():
        칸 = "".join(f'<td class="n">{cal_mark((a.get("날짜별") or {}).get(d))}</td>' for d in days)
        본문 += (f'<tr><td><b>{e(org)}</b><br><span class="dim" style="font-size:12px">{e(a.get("응답시각",""))}</span></td>'
                 f'{칸}<td>{e(a.get("메모")) or 빈칸}</td></tr>')
    for org in calendar_store.pending(ev, roster):
        본문 += (f'<tr><td>{e(org)}</td>' + "".join('<td class="n"><span class="dim">–</span></td>' for _ in days)
                 + '<td><span class="tag t-wait">미응답</span></td></tr>')

    표 = calendar_store.tally(ev)
    합계 = "".join(f'<td class="n">{표[d]["가능"]}</td>' for d in days)
    본문 += f'<tr class="total"><td>가능 인원</td>{합계}<td></td></tr>'
    return (f'<div class="scroll"><table><thead><tr><th>기관</th>{머리}<th>메모</th></tr></thead>'
            f'<tbody>{본문}</tbody></table></div>')


def cal_event_page(sess, ev, msg=""):
    """일정 조사 한 건 — 기관은 응답 입력, 관리자는 응답 현황·설정."""
    roster = storage.load_roster()
    days = calendar_store.days_of(ev)
    대상 = calendar_store.targets(ev, roster)
    응답 = ev.get("응답", {})
    닫힘 = calendar_store.is_closed(ev)
    볼수있음 = calendar_store.can_see_answers(ev, sess["role"], sess["org"], roster)

    공개표시 = ('<span class="eye">전체 공개</span>' if ev.get("공개")
                else '<span class="tag t-none">관리자만 보기</span>')
    개요 = f"""<div class="card"><div class="hd"><h2>{e(ev["제목"])} {공개표시}
{'<span class="tag t-none">마감</span>' if 닫힘 else ""}</h2>
<p class="sub" style="margin:0">{cal_period(ev)} · {len(days)}일
{" · 응답 마감 " + e(ev["마감일"]) if ev.get("마감일") else " · 마감일 없음"}
· 요청 {e(ev.get("작성자") or "관리자")} ({e(ev.get("생성시각",""))})</p></div>
{f'<p class="note">{e(ev["설명"])}</p>' if ev.get("설명") else ""}
<p class="hint" style="margin-top:10px">응답 {len(응답)}/{len(대상) or len(응답)}곳
{" · 미응답 " + e(", ".join(calendar_store.pending(ev, roster))) if calendar_store.pending(ev, roster) else ""}</p></div>"""

    # 내 응답 (기관 계정)
    내응답 = ""
    if sess["role"] != "admin":
        mine = (응답.get(sess["org"]) or {}).get("날짜별", {})
        메모 = (응답.get(sess["org"]) or {}).get("메모", "")
        if not calendar_store.is_target(ev, sess["org"], roster):
            내응답 = ('<div class="card"><p class="note">이 조사의 대상 기관이 아니어서 응답 칸이 없습니다. '
                      '전체 공개된 조사라 내용만 확인할 수 있습니다.</p></div>')
        elif 닫힘:
            내응답 = ('<div class="card"><p class="note warn">응답 마감일이 지나 더 이상 고칠 수 없습니다. '
                      '수정이 필요하면 관리자에게 문의해 주세요.</p></div>')
        else:
            내응답 = f"""<div class="card"><div class="hd"><h2>내 응답 — {e(sess["org"])}</h2>
<p class="sub" style="margin:0">달력에서 날짜마다 가능·불가를 골라 주세요. 다시 제출하면 이전 응답을 덮어씁니다.</p></div>
<form method="post" action="/calendar" id="calform"><input type="hidden" name="do" value="answer">
<input type="hidden" name="id" value="{ev["번호"]}">
<p class="noprint"><button type="button" class="sm ghost" onclick="calAll('가능')">전체 가능</button>
<button type="button" class="sm ghost" onclick="calAll('불가')">전체 불가</button>
<button type="button" class="sm ghost" onclick="calAll('미정')">전체 미정</button></p>
{cal_pick_grid(days, mine)}
<p class="hint">회색 날짜는 이번 조사에서 묻지 않는 날입니다. 고르지 않은 날은 '미정' 으로 저장됩니다.</p>
<label>메모 <span class="hint" style="font-weight:400">(선택)</span></label>
<input type="text" name="memo" value="{e(메모)}" placeholder="오전만 가능합니다">
<p style="margin-top:16px"><button type="submit">응답 제출</button></p></form></div>"""

    # 응답 현황
    if 볼수있음:
        현황 = f"""<div class="card"><div class="hd"><h2>응답 현황</h2>
<p class="sub" style="margin:0">달력은 날짜별 '가능' 인원, 아래 표는 기관별 응답입니다 — ○ 가능 · ✕ 불가 · – 미정(또는 미응답)</p></div>
{cal_count_grid(ev, days, len(대상)) if days else ""}
{cal_answer_table(ev, roster) if days else '<p class="note warn">기간이 잘못 저장되어 있습니다.</p>'}
{f'<p class="note ok">가장 많이 가능한 날 — <b>{e(", ".join(calendar_store.day_label(d) for d in calendar_store.best_days(ev)))}</b></p>' if calendar_store.best_days(ev) else '<p class="note">아직 가능으로 표시된 날짜가 없습니다.</p>'}</div>"""
    else:
        현황 = ('<div class="card"><div class="hd"><h2>응답 현황</h2></div>'
                '<p class="note">이 조사는 <b>관리자만 보기</b>로 설정되어 있어 다른 기관의 응답은 보이지 않습니다. '
                '관리자가 전체 공개로 바꾸면 여기에 함께 표시됩니다.</p></div>')

    # 관리자 설정
    설정 = ""
    if sess["role"] == "admin":
        sel = lambda v: " selected" if bool(ev.get("공개")) == v else ""
        설정 = f"""<div class="card noprint"><div class="hd"><h2>공개 범위 · 마감일</h2>
<p class="sub" style="margin:0">전체 공개로 바꾸면 대상 기관들이 서로의 응답을 볼 수 있습니다.</p></div>
<form method="post" action="/calendar"><input type="hidden" name="do" value="options">
<input type="hidden" name="id" value="{ev["번호"]}">
<div class="inline">
<div style="max-width:200px"><label>공개 범위</label><select name="public">
<option value="N"{sel(False)}>관리자만 보기</option><option value="Y"{sel(True)}>전체 공개</option></select></div>
<div style="max-width:200px"><label>응답 마감일</label><input type="date" name="due" value="{e(ev.get("마감일",""))}"></div>
<div><button type="submit">저장</button></div>
</div></form>
<form method="post" action="/calendar" style="margin-top:16px"
 onsubmit="return confirm('이 조사와 응답을 모두 지울까요?')">
<input type="hidden" name="do" value="delete"><input type="hidden" name="id" value="{ev["번호"]}">
<button class="sm danger">조사 삭제</button></form></div>"""

    return page(
        ev["제목"],
        f"""<h1>{e(ev["제목"])}</h1>
<p class="sub"><a href="/calendar">← 일정 조사 목록</a></p>
{msg}{개요}{내응답}{현황}{설정}
<script>{CAL_JS}</script>""",
        "calendar",
        sess,
    )


# ── 화면: 기관 본인 제출 내역 ────────────────────────────────────


def mine_page(sess):
    org = sess["org"]
    rows = ""
    for k in storage.list_weeks():
        d = storage.load_week(k)
        s = d["제출"].get(org)
        if not s:
            continue
        tags = {"승인": "t-ok", "제출": "t-wait", "반려": "t-rej"}
        rows += f"""<tr><td><b>{week_label(k)}</b></td><td class="n">{len(s['실적'])}</td><td>{e(s['출처'])}</td>
<td class="dim">{e(s['제출시각'])}</td><td><span class="tag {tags.get(s['상태'],'t-wait')}">{e(s['상태'])}</span></td>
<td><a href="/submit?week={quote(k)}"><button class="sm ghost">수정</button></a></td></tr>"""
    table = (
        f'<div class="scroll"><table><thead><tr><th>주차</th><th class="n">과정</th><th>제출 방식</th>'
        f'<th>제출 시각</th><th>상태</th><th></th></tr></thead><tbody>{rows}</tbody></table></div>'
        if rows
        else '<p class="note">아직 제출한 내역이 없습니다.</p>'
    )
    조사 = calendar_store.open_for(org, storage.load_roster())
    알림 = (
        '<p class="note warn">응답할 일정 조사가 있습니다 — '
        + ", ".join(f'<a href="/calendar?id={x["번호"]}">{e(x["제목"])}</a>' for x in 조사)
        + "</p>"
        if 조사
        else ""
    )
    return page(
        "내 제출 내역",
        f"""<h1>내 제출 내역 <span class="dim" style="font-size:15px">— {e(org)}</span></h1>
<p class="sub">제출한 주차와 승인 상태를 확인합니다. <b>반려</b>된 주차는 수정해서 다시 제출하세요.</p>
{알림}
<div class="card">{table}
<p style="margin-top:16px"><a href="/submit"><button>새로 입력하기</button></a>
<a href="/upload"><button class="ghost">엑셀로 올리기</button></a></p></div>""",
        "mine",
        sess,
    )


# ── 화면: 홈 ─────────────────────────────────────────────────────


def home_page(sess):
    weeks = storage.list_weeks()
    roster = storage.load_roster()
    rows = ""
    for k in weeks[:12]:
        d = storage.load_week(k)
        subs = d["제출"]
        appr = sum(1 for s in subs.values() if s["상태"] == "승인")
        courses = sum(len(s["실적"]) for s in subs.values())
        base = len(roster) or len(subs)
        rows += (
            f'<tr><td><b>{week_label(k)}</b></td>'
            f'<td class="n">{len(subs)}</td><td class="n">{courses}</td>'
            f'<td class="n">{appr}/{base}</td>'
            f'<td><a href="/report?week={quote(k)}"><button class="sm ghost">리포트</button></a> '
            f'<a href="/admin?week={quote(k)}"><button class="sm ghost">취합·승인</button></a> '
            f'<a href="/export?week={quote(k)}&kind=raw"><button class="sm ghost">입력값 엑셀</button></a></td></tr>'
        )
    table = (
        f'<div class="scroll"><table><thead><tr><th>주차</th><th class="n">제출 기관</th><th class="n">과정</th>'
        f'<th class="n">승인</th><th>바로가기</th></tr></thead><tbody>{rows}</tbody></table></div>'
        if rows
        else '<p class="note">아직 저장된 제출이 없습니다. <b>① 직접 입력</b> 또는 <b>① 엑셀 업로드</b>로 시작하세요.</p>'
    )
    return page(
        "주간 훈련기관 교육실적 취합",
        phead("대시보드", "주간 훈련기관 교육실적 취합·시각화",
              "기관이 직접 입력하거나 엑셀로 올리면 같은 저장소에 쌓이고, 관리자가 승인하면 주간 리포트가 만들어집니다.")
        + f"""

<div class="card"><div class="hd"><h2>저장된 주차</h2>
<p class="sub" style="margin:0">대상 명단 {len(roster)}곳 · 저장소 <code>data/</code></p></div>{table}</div>

<div class="card"><div class="hd"><h2>어떻게 쓰나요</h2></div>
<div class="grid">
<div class="rowbox"><b>① 훈련기관</b><p class="hint">화면에서 과정별로 직접 입력하거나, 정부 「지산맞」 양식을 그대로 업로드합니다. 제출 즉시 잠정 계산(실시율·수료율·탈락률)을 미리 볼 수 있습니다.</p></div>
<div class="rowbox"><b>② 관리자</b><p class="hint">제출·승인 현황판에서 기관별로 확인·승인·반려합니다. 대상 명단을 등록하면 미제출 기관이 잡힙니다.</p></div>
<div class="rowbox"><b>③ 리포트</b><p class="hint">전원 승인되면 취합·집계·전주 대비 비교·이상치 플래그·시각화가 자동으로 돌아갑니다. 엑셀·PDF로 내려받습니다.</p></div>
</div></div>""",
        "home",
        sess,
    )


# ── 화면: 직접 입력 ──────────────────────────────────────────────

JS_ROWS = """
function tpl(kind){
 if(kind==='p') return `<tr>
  <td><select name="p_type"><option>양성</option><option>향상</option></select></td>
  <td><select name="p_reg"><option>정기</option><option>수시</option></select></td>
  <td><input type="text" name="p_kind" placeholder="일반"></td>
  <td><input type="text" name="p_ncs" placeholder="기계"></td>
  <td><input type="text" name="p_keco" placeholder="금속가공 기계 조작원"></td>
  <td><input type="text" name="p_course" required placeholder="CNC밀링 조작"></td>
  <td><input type="number" name="p_goal" min="0" required></td>
  <td><input type="number" name="p_run" min="0" required></td>
  <td><input type="number" name="p_done" min="0" required></td>
  <td><input type="number" name="p_drop" min="0" required></td>
  <td><input type="number" name="p_ing" min="0"></td>
  <td><input type="number" name="p_emp" min="0"></td>
  <td><button type="button" class="sm danger" onclick="rm(this)">삭제</button></td></tr>`;
 if(kind==='n') return `<tr>
  <td><input type="text" name="n_course" placeholder="(선택) 과정명"></td>
  <td><select name="n_cat"><option>출결</option><option>시설</option><option>기타</option></select></td>
  <td><input type="text" name="n_body" placeholder="중도탈락 3명 중 2명 장기결석"></td>
  <td><select name="n_chk"><option>N</option><option>Y</option></select></td>
  <td><button type="button" class="sm danger" onclick="rm(this)">삭제</button></td></tr>`;
 return `<tr>
  <td><input type="date" name="s_date"></td>
  <td><select name="s_kind"><option>개강</option><option>종강</option><option>점검</option><option>기타</option></select></td>
  <td><input type="text" name="s_body" placeholder="스마트팩토리 4기 종강"></td>
  <td><button type="button" class="sm danger" onclick="rm(this)">삭제</button></td></tr>`;
}
function add(kind,box,n){for(var i=0;i<(n||1);i++)document.getElementById(box).insertAdjacentHTML('beforeend',tpl(kind));}
function rm(b){var r=b.closest('tr');var box=r.parentNode;if(box.children.length>1)r.remove();else alert('최소 1줄은 있어야 합니다.');}

// 붙여넣기: 엑셀에서 여러 줄을 복사해 첫 칸에 붙이면 표에 그대로 펼친다
function paste(ev,kind,box){
 var text=(ev.clipboardData||window.clipboardData).getData('text');
 if(!text || text.indexOf('	')<0 && text.indexOf('
')<0) return;
 ev.preventDefault();
 var rows=text.replace(/
/g,'').split('
').filter(function(x){return x.trim()!==''});
 var start=ev.target.closest('tr');
 rows.forEach(function(line,idx){
  var cells=line.split('	');
  var tr=start;
  if(idx>0){ start.insertAdjacentHTML('afterend',tpl(kind)); tr=start.nextElementSibling; start=tr; }
  var ins=tr.querySelectorAll('input,select');
  cells.forEach(function(v,i){ if(ins[i]) ins[i].value=v.trim(); });
 });
}
function bindPaste(box,kind){
 var el=document.getElementById(box);
 if(el) el.addEventListener('paste',function(ev){ if(ev.target.tagName==='INPUT') paste(ev,kind,box); });
}
window.addEventListener('DOMContentLoaded',function(){
 bindPaste('pbox','p'); bindPaste('nbox','n'); bindPaste('sbox','s');
 if(document.getElementById('wk_m')) syncWeek();
});

// 월 + 그 달의 몇 주차 → 연중 주차(ISO) 자동 계산
function isoKey(y,m,n){
 var first=new Date(y,m-1,1);
 var day=1+((7-first.getDay())%7)+7*(n-1);   // getDay(): 0=일요일
 var sun=new Date(y,m-1,day);
 if(sun.getMonth()!==m-1) return null;        // 그 달에 없는 주차
 var t=new Date(Date.UTC(sun.getFullYear(),sun.getMonth(),sun.getDate()));
 t.setUTCDate(t.getUTCDate()-((t.getUTCDay()+6)%7)+3);      // 그 주의 목요일
 var thu=new Date(Date.UTC(t.getUTCFullYear(),0,4));
 thu.setUTCDate(thu.getUTCDate()-((thu.getUTCDay()+6)%7)+3);
 var wk=1+Math.round((t-thu)/604800000);
 return {key:t.getUTCFullYear()+'-W'+String(wk).padStart(2,'0'),
         sun:sun.getFullYear()+'-'+String(sun.getMonth()+1).padStart(2,'0')+'-'+String(sun.getDate()).padStart(2,'0')};
}
function syncWeek(){
 var y=+document.getElementById('wk_y').value, m=+document.getElementById('wk_m').value, n=+document.getElementById('wk_n').value;
 var out=document.getElementById('wk_out'), hid=document.getElementById('wk_key');
 var r=isoKey(y,m,n);
 if(!r){ out.innerHTML='<b style="color:#b91c1c">'+m+'월에는 '+n+'주차가 없습니다</b>'; if(hid) hid.value=''; return; }
 out.innerHTML='연중 주차 <b>'+r.key+'</b> · 기준일(일요일) '+r.sun;
 if(hid) hid.value=r.key;
}
"""


def perf_row_html(p=None):
    p = p or {}
    v = lambda k: e(p.get(k, "") if p.get(k) is not None else "")
    sel = lambda cur, opt: " selected" if str(cur) == opt else ""
    구분 = "".join(f'<option{sel(p.get("구분", "양성"), o)}>{o}</option>' for o in ("양성", "향상"))
    정기 = "".join(f'<option{sel(p.get("정기수시", "정기"), o)}>{o}</option>' for o in ("정기", "수시"))
    필수숫자 = "".join(
        f'<td><input type="number" name="{n}" min="0" required value="{v(k)}"></td>'
        for n, k in [("p_goal", "훈련목표인원"), ("p_run", "훈련실시인원"),
                     ("p_done", "훈련수료인원"), ("p_drop", "중도탈락자")]
    )
    선택숫자 = "".join(
        f'<td><input type="number" name="{n}" min="0" value="{v(k)}"></td>'
        for n, k in [("p_ing", "훈련중"), ("p_emp", "취업인원")]
    )
    return f"""<tr><td><select name="p_type">{구분}</select></td>
<td><select name="p_reg">{정기}</select></td>
<td><input type="text" name="p_kind" value="{v('과정구분')}"></td>
<td><input type="text" name="p_ncs" value="{v('NCS대분류명')}"></td>
<td><input type="text" name="p_keco" value="{v('KECO세분류명')}"></td>
<td><input type="text" name="p_course" required value="{v('과정명')}"></td>
{필수숫자}{선택숫자}<td><button type="button" class="sm danger" onclick="rm(this)">삭제</button></td></tr>"""


def note_row_html(n=None):
    n = n or {}
    sel = lambda cur, opt: " selected" if str(cur) == opt else ""
    cat, chk = n.get("분류", "출결"), (n.get("확인필요") or "N").upper()
    return f"""<tr><td><input type="text" name="n_course" value="{e(n.get('과정명',''))}"></td>
<td><select name="n_cat">{''.join(f'<option{sel(cat,o)}>{o}</option>' for o in ('출결','시설','기타'))}</select></td>
<td><input type="text" name="n_body" value="{e(n.get('내용',''))}"></td>
<td><select name="n_chk">{''.join(f'<option{sel(chk,o)}>{o}</option>' for o in ('N','Y'))}</select></td>
<td><button type="button" class="sm danger" onclick="rm(this)">삭제</button></td></tr>"""


def plan_row_html(s=None):
    s = s or {}
    sel = lambda cur, opt: " selected" if str(cur) == opt else ""
    return f"""<tr><td><input type="date" name="s_date" value="{e(s.get('날짜',''))}"></td>
<td><select name="s_kind">{''.join(f'<option{sel(s.get("구분","개강"),o)}>{o}</option>' for o in ('개강','종강','점검','기타'))}</select></td>
<td><input type="text" name="s_body" value="{e(s.get('내용',''))}"></td>
<td><button type="button" class="sm danger" onclick="rm(this)">삭제</button></td></tr>"""


def submit_page(sess, week=None, org=None):
    week = week or default_week()
    if sess["role"] != "admin":
        org = sess["org"]  # 기관 계정은 자기 기관만
    y, w = storage.parse_week(week) or storage.parse_week(default_week())
    data = storage.load_week(week)
    sub = data["제출"].get(org) if (data and org) else None
    m월, n주차 = key_to_month_week(week)
    m월 = (data or {}).get("월") or m월
    자동키, 기준일 = month_week_to_key(y, m월, n주차)
    주차옵션 = "".join(
        f'<option value="{i}"{" selected" if i == n주차 else ""}>{i}주차</option>' for i in range(1, 6)
    )
    roster = storage.load_roster()

    perf = "".join(perf_row_html(p) for p in (sub["실적"] if sub else [None] * 5))
    notes = "".join(note_row_html(n) for n in (sub["특이사항"] if sub and sub["특이사항"] else [None, None]))
    plans = "".join(plan_row_html(s) for s in (sub["주요일정"] if sub and sub["주요일정"] else [None, None]))

    if sess["role"] != "admin":  # 기관 계정은 자기 기관으로 고정
        picker = f'<input type="text" value="{e(org)}" disabled><input type="hidden" name="org" value="{e(org)}">'
    elif roster:
        opts = "".join(f'<option{" selected" if n == org else ""}>{e(n)}</option>' for n in roster)
        picker = f'<select name="org" required><option value="">— 기관 선택 —</option>{opts}</select>'
    else:
        picker = f'<input type="text" name="org" required value="{e(org or "")}" placeholder="가온직업전문학교">'

    banner = ""
    if sub:
        banner = f'<p class="note ok">이미 제출한 내용을 불러왔습니다 ({e(sub["제출시각"])} · {e(sub["출처"])}). 고쳐서 다시 제출하면 <b>덮어쓰기(재입력)</b> 되고 승인 상태는 초기화됩니다.</p>'

    return page(
        "실적 직접 입력",
        f"""<h1>① 실적 직접 입력</h1>
<p class="sub">엑셀처럼 표에 그대로 입력합니다. 여러 줄을 한 번에 넣을 수 있고, 엑셀에서 복사해 붙여넣어도 됩니다. <span class="dim">기획서 3-1단계</span></p>
<form method="post" action="/submit">
<div class="card"><div class="hd"><h2>제출 정보</h2></div>{banner}
<div class="inline">
<div><label>훈련기관 *</label>{picker}</div>
<div style="max-width:110px"><label>연 *</label>
<input type="number" id="wk_y" name="year" required value="{y}" oninput="syncWeek()"></div>
<div style="max-width:110px"><label>월 *</label>
<input type="number" id="wk_m" name="month" min="1" max="12" required value="{m월}" oninput="syncWeek()"></div>
<div style="max-width:150px"><label>그 달의 몇 주차 *</label>
<select id="wk_n" name="mweek" onchange="syncWeek()">{주차옵션}</select></div>
</div>
<p class="hint" id="wk_out">연중 주차 <b>{e(자동키)}</b> · 기준일(일요일) {e(str(기준일) if 기준일 else "-")}</p>
<input type="hidden" id="wk_key" name="week" value="{e(자동키)}">
<p class="hint">{WEEK_NOTE}. 저장은 연중 주차(<b>연도-W주차번호</b>)로 하고, 작년 동기 비교도 이 번호로 맞춥니다.</p></div>

<div class="card"><div class="hd"><h2>실적 (과정별)</h2>
<p class="sub" style="margin:0">한 줄 = 한 과정 · 수료율 = 수료÷실시 · 탈락률 = 중도탈락÷실시 · 취업은 양성 과정만</p></div>
<div class="scroll"><table class="grid"><thead><tr>
<th>구분 *</th><th>정기/수시 *</th><th>과정구분</th><th>NCS대분류</th><th>KECO세분류</th><th>과정명 *</th>
<th class="n">목표(정원) *</th><th class="n">실시 *</th><th class="n">수료 *</th><th class="n">탈락 *</th>
<th class="n">훈련중</th><th class="n">취업</th><th></th></tr></thead>
<tbody id="pbox">{perf}</tbody></table></div>
<p style="margin-top:10px;display:flex;gap:8px;flex-wrap:wrap">
<button type="button" class="ghost sm" onclick="add('p','pbox')">+ 1줄 추가</button>
<button type="button" class="ghost sm" onclick="add('p','pbox',5)">+ 5줄 추가</button></p>
<p class="hint">엑셀에서 여러 줄을 복사해 첫 칸에 붙여넣으면 줄이 자동으로 늘어납니다(열 순서 그대로).</p></div>

<div class="card"><div class="hd"><h2>특이사항 <span class="dim" style="font-weight:400;font-size:13px">(선택)</span></h2></div>
<div class="scroll"><table class="grid"><thead><tr>
<th>과정명</th><th>분류</th><th>내용</th><th>확인 필요</th><th></th></tr></thead>
<tbody id="nbox">{notes}</tbody></table></div>
<p style="margin-top:10px"><button type="button" class="ghost sm" onclick="add('n','nbox')">+ 1줄 추가</button></p></div>

<div class="card"><div class="hd"><h2>주요일정 <span class="dim" style="font-weight:400;font-size:13px">(선택)</span></h2></div>
<div class="scroll"><table class="grid"><thead><tr>
<th>날짜</th><th>구분</th><th>내용</th><th></th></tr></thead>
<tbody id="sbox">{plans}</tbody></table></div>
<p style="margin-top:10px"><button type="button" class="ghost sm" onclick="add('s','sbox')">+ 1줄 추가</button></p></div>

<p><button type="submit">제출하고 잠정 계산 보기</button></p></form>
<script>{JS_ROWS}</script>""",
        "submit",
        sess,
    )


# ── 화면: 엑셀 업로드 ────────────────────────────────────────────


def upload_page(sess, msg=""):
    week = default_week()
    y, w = storage.parse_week(week)
    m월, n주차 = key_to_month_week(week)
    자동키, 기준일 = month_week_to_key(y, m월, n주차)
    주차옵션 = "".join(
        f'<option value="{i}"{" selected" if i == n주차 else ""}>{i}주차</option>' for i in range(1, 6)
    )
    return page(
        "엑셀 업로드",
        f"""<h1>① 엑셀 업로드</h1>
<p class="sub">정부 「지산맞 훈련실적」 양식(「양성훈련 현황」·「향상훈련 현황」 시트)을 그대로 올립니다.
받은 파일을 고치지 않고 올리면 됩니다. 직접 입력과 <b>같은 저장소</b>에 쌓입니다.</p>
{msg}
<div class="card"><div class="hd"><h2>양식 내려받기</h2>
<p class="sub" style="margin:0">정부 양식이 없을 때만 쓰는 간소 양식입니다. 정부 원본이 있으면 그대로 올리셔도 됩니다.</p></div>
<p><a class="btn" href="/template"><button type="button">간소 빈 양식 내려받기 (.xlsx)</button></a></p>
<p class="hint">「양성훈련 현황」·「향상훈련 현황」 두 시트에 과정별로 채웁니다. 열은 이름으로 찾으므로
정부 원본의 열 순서 그대로도 읽힙니다. 기관 연간 목표는 「교육실적」 시트에서 읽어 실시율·수료율(목표 대비)의 분모로 씁니다.</p></div>

<form class="card" method="post" action="/upload" enctype="multipart/form-data">
<div class="hd"><h2>파일 선택</h2></div>
<label>실적 엑셀 <span style="color:#dc2626">*필수</span>
<div class="hint">한 파일에 여러 기관이 섞여 있어도 <b>기관별로 나눠서</b> 저장합니다.</div></label>
<input type="file" name="file" accept=".xlsx,.xlsm" required>
<div class="inline">
<div><label>연 *</label><input type="number" id="wk_y" name="year" required value="{y}" oninput="syncWeek()"></div>
<div><label>월 *</label><input type="number" id="wk_m" name="month" min="1" max="12" required value="{m월}" oninput="syncWeek()"></div>
<div><label>그 달의 몇 주차 *</label><select id="wk_n" name="mweek" onchange="syncWeek()">{주차옵션}</select></div>
</div>
<p class="hint" id="wk_out">연중 주차 <b>{e(자동키)}</b> · 기준일(일요일) {e(str(기준일) if 기준일 else "-")}</p>
<input type="hidden" id="wk_key" name="week" value="{e(자동키)}">
<p class="hint">엑셀 양식에 주차 칸이 없으므로 여기서 지정합니다. {WEEK_NOTE}.</p>
<p style="margin-top:20px"><button type="submit">업로드하고 잠정 계산 보기</button></p></form>
<script>{JS_ROWS}</script>""",
        "upload",
        sess,
    )


# ── 화면: 제출 직후 잠정 계산 미리보기 (기획서 3-1단계) ──────────


def preview_page(sess, week, orgs, result, source):
    res = result
    body = ""
    for r in res["표"]:
        cls = ' class="err"' if r["_오류"] else ""
        cells = "".join(
            f'<td class="n">{cell(r[c], r["_원본"].get(c))}</td>'
            for c in ["훈련목표인원", "훈련실시인원", "훈련수료인원", "중도탈락자"]
        )
        body += (
            f'<tr{cls}><td class="n">{"⚠" if r["_오류"] else ""}</td><td>{e(r["구분"])}</td>'
            f'<td>{e(r["기관명"])}</td><td>{e(r["과정명"])}</td>{cells}'
            f'<td class="n">{pct(r["수료율"])}</td><td class="n">{pct(r["탈락률"])}</td></tr>'
        )
    t = res["합계_전체"][0]
    body += (
        f'<tr class="total"><td></td><td></td><td>합계</td><td>정상 {t["과정수"]}개'
        + (f' · 제외 {t["제외"]}개' if t["제외"] else "")
        + "</td>"
        + "".join(f'<td class="n">{t[c]}</td>' for c in ["훈련목표인원", "훈련실시인원", "훈련수료인원", "중도탈락자"])
        + f'<td class="n">{pct(t["수료율"])}</td><td class="n">{pct(t["탈락률"])}</td></tr>'
    )

    err = ""
    if res["오류"]:
        er = "".join(f'<tr><td class="n">{n}</td><td>{e(c)}</td><td>{e(m)}</td></tr>' for n, c, m in res["오류"])
        err = f"""<h3>입력 오류 {len(res['오류'])}건 — 고쳐서 다시 제출해 주세요</h3>
<div class="scroll"><table><thead><tr><th class="n">행</th><th>컬럼</th><th>사유</th></tr></thead><tbody>{er}</tbody></table></div>
<p class="note warn">오류가 있는 과정은 지표를 계산하지 않고(<b>검증 필요</b>) 합계에서도 제외했습니다.
관리자 승인 단계에서 <b>반려 대상</b>으로 표시됩니다.</p>"""

    edit = "".join(
        f'<a href="/submit?week={quote(week)}&org={quote(o)}"><button class="ghost sm">{e(o)} 수정</button></a> '
        for o in orgs
    )
    if sess["role"] == "admin":
        nav = (
            f'<a href="/admin?week={quote(week)}"><button class="ghost">취합·승인 화면</button></a>'
            f'<a href="/report?week={quote(week)}"><button class="ghost">주간 리포트</button></a>'
            f'<a href="/export?week={quote(week)}&kind=raw"><button>입력값 엑셀 받기</button></a>'
        )
    else:
        nav = '<a href="/mine"><button class="ghost">내 제출 내역</button></a>'
    return page(
        "제출 완료 — 잠정 계산",
        f"""<h1>제출되었습니다 <span class="dim" style="font-size:15px">— {week_label(week)}</span></h1>
<p class="sub">{e(', '.join(orgs))} · {e(source)} · 저장소에 기록됨</p>
<div class="card"><div class="hd"><h2>잠정 계산 미리보기</h2>
<p class="sub" style="margin:0">확정이 아닙니다. 관리자 승인 전 참고용입니다. <span class="dim">기획서 3-1단계</span></p></div>
<div class="scroll"><table><thead><tr><th></th><th>구분</th><th>기관</th><th>과정명</th>
<th class="n">목표</th><th class="n">실시</th><th class="n">수료</th><th class="n">탈락</th>
<th class="n">수료율</th><th class="n">탈락률</th></tr></thead><tbody>{body}</tbody></table></div>{err}
<p style="margin-top:20px;display:flex;gap:9px;flex-wrap:wrap">{edit}{nav}</p></div>""",
        "",
        sess,
    )


# ── 화면: 관리자 취합·승인 ───────────────────────────────────────


def admin_page(sess, week=None, msg=""):
    weeks = storage.list_weeks()
    week = week or (weeks[0] if weeks else default_week())
    data = storage.load_week(week)
    roster = storage.load_roster()

    opts = "".join(f'<option value="{e(k)}"{" selected" if k == week else ""}>{e(k)}</option>' for k in weeks)
    if week not in weeks:
        opts = f'<option value="{e(week)}" selected>{e(week)} (비어 있음)</option>' + opts

    picker = f"""<form method="get" action="/admin" class="inline" style="margin-bottom:4px">
<div style="max-width:220px"><label>주차</label><select name="week" onchange="this.form.submit()">{opts}</select></div>
</form>"""

    tags = {"승인": "t-ok", "제출": "t-wait", "반려": "t-rej"}
    rows, submitted = "", set()
    if data:
        for org, s in data["제출"].items():
            submitted.add(org)
            extra = "" if org in roster or not roster else ' <span class="tag t-none">명단 외</span>'
            rows += f"""<tr><td><b>{e(org)}</b>{extra}</td><td class="n">{len(s['실적'])}</td>
<td>{e(s['출처'])}</td><td class="dim">{e(s['제출시각'])}</td>
<td><span class="tag {tags.get(s['상태'],'t-wait')}">{e(s['상태'])}</span></td>
<td class="noprint"><form method="post" action="/status" style="display:flex;gap:5px">
<input type="hidden" name="week" value="{e(week)}"><input type="hidden" name="org" value="{e(org)}">
<button class="sm" name="status" value="승인">승인</button>
<button class="sm ghost" name="status" value="반려">반려</button>
<button class="sm danger" name="status" value="삭제" onclick="return confirm('{e(org)} 제출을 삭제할까요? 되돌릴 수 없습니다.')">삭제</button>
</form></td></tr>"""
    for name in roster:
        if name not in submitted:
            rows += f"""<tr><td>{e(name)}</td><td class="n dim">—</td><td class="dim">—</td><td class="dim">—</td>
<td><span class="tag t-none">미제출</span></td><td class="dim noprint">대기</td></tr>"""

    base = len(roster) or len(submitted)
    appr = sum(1 for s in (data["제출"].values() if data else []) if s["상태"] == "승인")
    rate = f"{appr}/{base} ({round(appr / base * 100) if base else 0}%)"
    blocked = base - appr
    status_note = (
        f'<p class="note ok">대상 {base}곳 전원 승인 완료 — 주간 리포트를 발행할 수 있습니다.</p>'
        if base and appr == base
        else f'<p class="note warn">아직 {blocked}곳이 승인 전입니다. 기획서 기준으로 <b>미제출·미승인 기관이 있으면 리포트를 마감(발행)하지 않습니다.</b> 지금 리포트를 열면 <b>승인 전 잠정 계산</b>입니다.</p>'
    )

    table = (
        f"""<div class="scroll"><table><thead><tr><th>훈련기관</th><th class="n">과정</th><th>제출 방식</th>
<th>제출 시각</th><th>상태</th><th class="noprint">처리</th></tr></thead><tbody>{rows}</tbody></table></div>{status_note}"""
        if rows
        else '<p class="note">이 주차에 제출된 내용이 없습니다.</p>'
    )

    return page(
        "취합·승인",
        phead("훈련실적", '② 취합·승인 <span class="eye">👁 4단계</span>',
              f"{week_label(week, True)} · 기관별 제출 내용을 확인하고 승인·반려합니다. 승인 완료율 <b>{rate}</b>")
        + f"""
{msg}
<div class="card">{picker}{table}
<p style="margin-top:16px;display:flex;gap:9px;flex-wrap:wrap">
<a href="/report?week={quote(week)}"><button>주간 리포트 보기</button></a>
<a href="/export?week={quote(week)}&kind=raw"><button class="ghost">입력값 엑셀 받기</button></a></p></div>

<div class="card"><div class="hd"><h2>대상 훈련기관 명단</h2>
<p class="sub" style="margin:0">한 줄에 하나. 등록하면 <b>미제출 기관</b>이 잡히고 승인 완료율의 분모가 됩니다.</p></div>
<form method="post" action="/roster">
<textarea name="roster" placeholder="가온직업전문학교&#10;새빛테크노교육원&#10;한들평생교육원">{e(chr(10).join(roster))}</textarea>
<input type="hidden" name="week" value="{e(week)}">
<p style="margin-top:12px"><button type="submit" class="ghost">명단 저장</button></p></form></div>""",
        "admin",
        sess,
    )


# ── 화면: 리포트 ─────────────────────────────────────────────────


TODO_SUB = "아래 리포트에서 조치가 필요한 것만 추린 목록입니다. 판단은 리포트 결과를 그대로 따릅니다."


def render_todos(rep, sub=TODO_SUB, todos=None):
    """요약 결과에서 뽑아낸 '오늘 할 일'만 따로 보여준다."""
    todos = rep["오늘할일"] if todos is None else todos
    hd = f"""<div class="hd"><h2>오늘 할 일 <span class="eye">👁 기준일 {e(rep['기준일'])}</span></h2>
<p class="sub" style="margin:0">{sub}</p></div>"""
    if not todos:
        return f'<div class="card">{hd}<p class="note ok">오늘 처리할 항목이 없습니다.</p></div>'

    tags = {"높음": "t-rej", "보통": "t-wait", "낮음": "t-none"}
    rows = "".join(
        f'<tr><td><span class="tag {tags[t["우선순위"]]}">{e(t["우선순위"])}</span></td>'
        f'<td>{e(t["구분"])}</td><td><b>{e(t["대상"])}</b></td><td style="white-space:normal">{e(t["할일"])}</td></tr>'
        for t in todos
    )
    high = sum(1 for t in todos if t["우선순위"] == "높음")
    checks = "".join(
        f'<label><input type="checkbox">{e(t["대상"])} — {e(t["구분"])}</label>' for t in todos
    )
    return f"""<div class="card">{hd}
<div class="scroll"><table><thead><tr><th>우선순위</th><th>구분</th><th>대상</th><th>할 일</th></tr></thead>
<tbody>{rows}</tbody></table></div>
<p class="note{' warn' if high else ''}">전체 {len(todos)}건 · 오늘 먼저 볼 것 {high}건</p>
<div class="chk noprint"><b>처리 체크</b>{checks}</div></div>"""


def apply_approval(rep, data):
    """저장소의 승인·반려 상태를 현황판에 반영하고 '오늘 할 일'을 다시 뽑는다.

    build_report 는 저장소의 승인 여부를 모르므로 승인된 기관도 '검토대기'로 둔다.
    현황판을 고친 뒤이므로 파생 결과인 오늘 할 일도 같이 다시 계산한다.
    """
    for b in rep["현황판"]:
        s = (data["제출"] if data else {}).get(b["기관명"])
        if s and s["상태"] == "승인" and b["상태"] == "검토대기":
            b["상태"], b["비고"] = "승인", s["제출시각"] + " 제출"
        elif s and s["상태"] == "반려":
            b["상태"] = "반려 대상"
    rep["오늘할일"] = today_todos(rep, rep["기준일"])
    return rep


def render_board(rep):
    board = rep["현황판"]
    tags = {"검토대기": "t-wait", "반려 대상": "t-rej", "미제출": "t-none", "명단 외": "t-none", "승인": "t-ok"}
    rows = "".join(
        f'<tr><td>{e(b["기관명"])}</td><td class="n">{"✅" if b["제출"] else "❌"}</td>'
        f'<td class="n">{"⚠ " + str(b["오류행"]) + "행" if b["오류행"] else ("⏳" if b["제출"] else "—")}</td>'
        f'<td><span class="tag {tags.get(b["상태"],"t-none")}">{e(b["상태"])}</span></td>'
        f'<td>{e(b["비고"]) or "—"}</td></tr>'
        for b in board
    )
    blocked = [b for b in board if b["상태"] in ("미제출", "반려 대상")]
    warn = ""
    if blocked:
        warn = (
            f'<p class="note warn"><b>리포트 발행 보류 대상:</b> {e(", ".join(b["기관명"] for b in blocked))} — '
            "아래 결과는 <b>승인 전 잠정 계산</b>입니다.</p>"
        )
    return f"""<div class="card"><div class="hd"><h2>① 제출·승인 현황판 <span class="eye">👁 4단계</span></h2>
<p class="sub" style="margin:0">대상 {len(board)}곳</p></div>
<div class="scroll"><table><thead><tr><th>훈련기관</th><th class="n">제출</th><th class="n">검토</th><th>상태</th><th>비고</th></tr></thead>
<tbody>{rows}</tbody></table></div>{warn}
<p class="noprint" style="margin-top:14px"><a href="/admin?week={quote(rep['주차'])}"><button class="ghost sm">승인 처리하러 가기</button></a></p></div>"""


def render_table(rep):
    res = rep["실적"]
    body = ""
    for r in res["표"]:
        cls = ' class="err"' if r["_오류"] else ""
        cells = "".join(
            f'<td class="n">{cell(r[c], r["_원본"].get(c))}</td>'
            for c in ["훈련목표인원", "훈련실시인원", "훈련수료인원", "중도탈락자"]
        )
        body += (
            f'<tr{cls}><td class="n">{"⚠" if r["_오류"] else ""}</td><td>{e(r["구분"])}</td>'
            f'<td>{e(r["기관명"])}</td><td>{e(r["과정명"])}</td>'
            f'{cells}<td class="n">{pct(r["수료율"])}</td><td class="n">{pct(r["탈락률"])}</td></tr>'
        )
    t = res["합계_전체"][0]
    sub = f'정상 {t["과정수"]}개' + (f' · 제외 {t["제외"]}개' if t["제외"] else "")
    body += (
        f'<tr class="total"><td></td><td></td><td>합계</td><td>{sub}</td>'
        + "".join(f'<td class="n">{t[c]}</td>' for c in ["훈련목표인원", "훈련실시인원", "훈련수료인원", "중도탈락자"])
        + f'<td class="n">{pct(t["수료율"])}</td><td class="n">{pct(t["탈락률"])}</td></tr>'
    )
    err = ""
    if res["오류"]:
        er = "".join(f'<tr><td class="n">{n}</td><td>{e(c)}</td><td>{e(m)}</td></tr>' for n, c, m in res["오류"])
        err = f"""<h3>입력 오류 {len(res['오류'])}건</h3>
<div class="scroll"><table><thead><tr><th class="n">행</th><th>컬럼</th><th>사유</th></tr></thead><tbody>{er}</tbody></table></div>
<p class="note">⚠ 행은 지표를 계산하지 않고(<b>검증 필요</b>) 합계에서도 제외했습니다.</p>"""
    return f"""<div class="card"><div class="hd"><h2>② 통합 교육실적표</h2>
<p class="sub" style="margin:0">과정 단위 · 수료율 = 수료÷실시 · 탈락률 = 중도탈락÷실시</p></div>
<div class="scroll"><table><thead><tr><th></th><th>구분</th><th>훈련기관</th><th>과정명</th>
<th class="n">목표</th><th class="n">실시</th><th class="n">수료</th><th class="n">탈락</th>
<th class="n">수료율</th><th class="n">탈락률</th></tr></thead><tbody>{body}</tbody></table></div>{err}</div>"""


def bars(groups, cmp_map=None):
    out = ""
    for g in groups:
        v = g["수료율"]
        vf = v if isinstance(v, float) else None  # 문자열 상수(계산 불가 등)는 막대 0
        w = 0 if vf is None else min(100, vf * 100)
        low = " low" if vf is not None and vf < 0.6 else ""
        d = " " + delta(cmp_map[g["구분"]]["수료율"]) if cmp_map and g["구분"] in cmp_map else ""
        ex = f' <span class="dim">제외 {g["제외"]}</span>' if g["제외"] else ""
        out += (
            f'<div class="bar"><div class="nm">{e(g["구분"])}</div>'
            f'<div class="tr"><div class="fl{low}" style="width:{w:.1f}%"></div></div>'
            f'<div class="vl">{pct(v)}{d}{ex}</div></div>'
        )
    return out


def render_charts(rep):
    res, cmp_ = rep["실적"], rep["비교"]
    t = res["합계_전체"][0]
    kpi = (
        '<div class="kpi">'
        + stat("chart", "실시율 (목표 대비)", pct(t["실시율"]), '<span class="chip n">실시 ÷ 연간목표</span>')
        + stat("star", "수료율 (실시 대비)", pct(t["수료율"]), delta_chip(cmp_["전체"]["수료율"] if cmp_ else None))
        + stat("trend", "전체 탈락률", pct(t["탈락률"]), delta_chip(cmp_["전체"]["탈락률"] if cmp_ else None))
        + stat("list", "집계 과정", t["과정수"],
               f'<span class="chip n">제외 {t["제외"]}개</span> <span>검증 오류</span>' if t["제외"]
               else '<span class="chip n">전 과정 집계</span>')
        + "</div>"
    )
    first = (
        '<p class="note">저장소에 직전 주차 데이터가 없어 <b>전주 대비 비교를 생략</b>하고 이번 주를 기준주로 삼았습니다. (흐름도 8단계 분기)</p>'
        if rep["첫주"]
        else ""
    )
    blocks = ""
    for title, key in [("기관별 수료율", "합계_기관별"), ("양성/향상/수시 수료율", "합계_구분별"),
                       ("NCS대분류별 수료율", "합계_NCS별"), ("KECO세분류별 수료율", "합계_KECO별")]:
        blocks += f'<h3>{title}</h3>{bars(res[key], cmp_["기관별"] if (cmp_ and key == "합계_기관별") else None)}'
    return f"""<div class="card"><div class="hd"><h2>③ 핵심 지표 시각화</h2>
<p class="sub" style="margin:0">막대는 수료율(실시 대비) · 기관별·양성/향상/수시·NCS·KECO 합계{' · 전주 대비 증감 표시' if cmp_ else ''}</p></div>
{kpi}{first}{blocks}</div>"""


def render_outliers(rep):
    o = rep["이상치"]
    if o is None:
        return """<div class="card"><div class="hd"><h2>④ 이상치 강조 <span class="eye">👁 12단계</span></h2></div>
<p class="note">저장소에 <b>작년 동기(같은 주차번호) 데이터가 없어</b> 이상치 플래그를 건너뛰었습니다.
작년 자료를 같은 주차로 올려두면 다음부터 자동으로 비교합니다.</p></div>"""
    if not o["플래그"]:
        return f"""<div class="card"><div class="hd"><h2>④ 이상치 강조 <span class="eye">👁 12단계</span></h2></div>
<p class="note">작년 동기와 매칭된 {o['비교된과정']}개 과정 모두 10%p 이내입니다. 플래그 없음.</p></div>"""
    items = "".join(
        f"""<div class="flag"><b>{"🔻" if f["방향"] == "급감" else "🔺"} {e(f["기관명"])} · {e(f["과정명"])}</b><br>
작년 동기 <b>{f['작년']*100:.1f}%</b> ({e(f['작년과정'])}) → 이번 <b>{f['이번']*100:.1f}%</b>
&nbsp;·&nbsp; 차이 <b>{f['차이']:+.1f}%p</b> ({e(f['방향'])})</div>"""
        for f in o["플래그"]
    )
    checks = "".join(
        f'<label><input type="checkbox">{e(f["기관명"])} {e(f["과정명"])} — 사유를 확인했습니다.</label>' for f in o["플래그"]
    )
    return f"""<div class="card"><div class="hd"><h2>④ 이상치 강조 <span class="eye">👁 12단계</span></h2>
<p class="sub" style="margin:0">작년 동기 대비 수료율 10%p 이상 차이 · 매칭 {o['비교된과정']}개 중 {len(o['플래그'])}건</p></div>
{items}<div class="chk noprint"><b>👁 이상치·증감 확인</b>{checks}</div></div>"""


def render_notes(rep):
    if not rep["특이사항"]:
        return '<div class="card"><div class="hd"><h2>⑤ 특이사항 요약</h2></div><p class="note">보고된 특이사항이 없습니다.</p></div>'
    blocks = ""
    for cat, items in rep["특이사항_분류"]:
        lis = "".join(
            f'<li>{"<b style=\'color:#b45309\'>[확인 필요]</b> " if n["확인필요"].upper() == "Y" else ""}'
            f'<b>{e(n["기관명"])}</b> {e(n["과정명"])} — {e(n["내용"])}</li>'
            for n in items
        )
        blocks += f'<h3>{e(cat)} <span class="dim" style="font-weight:400">({len(items)}건)</span></h3><ul class="notes">{lis}</ul>'
    return f"""<div class="card"><div class="hd"><h2>⑤ 특이사항 요약</h2>
<p class="sub" style="margin:0">전체 {len(rep['특이사항'])}건 · 확인 필요 {rep['확인필요수']}건</p></div>{blocks}</div>"""


def render_schedule(rep):
    if not rep["주요일정"]:
        return '<div class="card"><div class="hd"><h2>⑥ 주요일정</h2></div><p class="note">등록된 일정이 없습니다.</p></div>'
    rows = "".join(
        f'<tr><td>{e(s["날짜"])}</td><td>{e(s["기관명"])}</td><td>{e(s["구분"])}</td><td>{e(s["내용"])}</td></tr>'
        for s in rep["주요일정"]
    )
    return f"""<div class="card"><div class="hd"><h2>⑥ 주요일정 정리</h2></div>
<div class="scroll"><table><thead><tr><th>날짜</th><th>기관</th><th>구분</th><th>내용</th></tr></thead><tbody>{rows}</tbody></table></div></div>"""


def render_summary(rep, week):
    return f"""<div class="card"><div class="hd"><h2>요약 문구 확정 <span class="eye">👁 16단계</span></h2>
<p class="sub" style="margin:0">14단계에서 만든 초안입니다. 문구·톤을 고쳐 확정하세요.</p></div>
<textarea style="min-height:170px">{e(rep['요약초안'])}</textarea>
<p class="note noprint">이 칸의 편집 내용은 화면에만 남습니다(문구 저장 기능은 아직 없음). 실적·특이사항·일정은 저장소에 저장되어 있습니다.</p>
<p class="noprint" style="margin-top:16px;display:flex;gap:9px;flex-wrap:wrap">
<a href="/export?week={quote(week)}&kind=report"><button>리포트 엑셀 받기</button></a>
<a href="/export?week={quote(week)}&kind=raw"><button class="ghost">입력값 엑셀 받기</button></a>
<button type="button" class="ghost" onclick="window.print()">PDF로 저장 (인쇄)</button></p></div>"""


def yearly_totals(week):
    """선택 주차가 속한 해의 누적 실적을 기관별로 모은다.

    저장소의 같은 연도 주차를 그 주차까지 모아 한 번에 집계한다. 지표는 여기서
    새로 만들지 않고 feature1(process_rows)에 그대로 맡긴다.
    """
    year, wk = storage.parse_week(week)
    keys = [k for k in storage.list_weeks() if storage.parse_week(k)[0] == year and storage.parse_week(k)[1] <= wk]
    rows = []
    for k in sorted(keys, key=storage.parse_week):
        data = storage.load_week(k)
        if data:
            rows.extend(storage.to_rows(data))
    for i, r in enumerate(rows, 1):  # 여러 주차를 합치면 _행 이 겹쳐 오류 행 표기가 꼬인다
        r["_행"] = i
    return process_rows(rows), sorted(keys, key=storage.parse_week), year


def render_yearly(week):
    """기관별 당해연도 누적 실적 막대그래프 (누적 수료인원 기준)."""
    res, keys, year = yearly_totals(week)
    groups = [g for g in res["합계_기관별"] if g["구분"]]
    if not keys or not groups:
        return f"""<div class="card"><div class="hd"><h2>③-2 기관별 {year}년 누적 실적</h2></div>
<p class="note">{year}년에 저장된 주차가 없어 누적 실적을 만들지 못했습니다.</p></div>"""

    top = max(g["훈련수료인원"] for g in groups) or 1
    bars_html = ""
    for g in groups:
        w = g["훈련수료인원"] / top * 100
        목표 = g.get("목표훈련인원") or g["훈련목표인원"]
        bars_html += (
            f'<div class="bar"><div class="nm">{e(g["구분"])}</div>'
            f'<div class="tr"><div class="fl" style="width:{w:.1f}%"></div></div>'
            f'<div class="vl">{g["훈련수료인원"]:,}명 <span class="dim">/ 목표 {목표:,}</span> · {pct(g["수료율_목표"])}</div></div>'
        )

    rows = "".join(
        f'<tr><td>{e(g["구분"])}</td><td class="n">{g["과정수"]}개 과정</td>'
        f'<td class="n">{g["훈련목표인원"]:,} / {g["훈련실시인원"]:,} / {g["훈련수료인원"]:,}</td></tr>'
        for g in groups
    )
    t = res["합계_전체"][0]
    rows += (
        f'<tr class="total"><td>합계</td><td class="n">{t["과정수"]}개 과정</td>'
        f'<td class="n">{t["훈련목표인원"]:,} / {t["훈련실시인원"]:,} / {t["훈련수료인원"]:,}</td></tr>'
    )
    빠짐 = f' · 입력 오류로 제외 {t["제외"]}개 과정' if t["제외"] else ""
    return f"""<div class="card"><div class="hd"><h2>③-2 기관별 {year}년 누적 실적</h2>
<p class="sub" style="margin:0">{e(keys[0])} ~ {e(keys[-1])} · {len(keys)}개 주차를 합산 · 막대는 누적 수료인원{빠짐}</p></div>
{bars_html}
<h3>누적 인원 (목표 / 실시 / 수료)</h3>
<div class="scroll"><table><thead><tr><th>훈련기관</th><th class="n">누적 과정</th><th class="n">목표 / 실시 / 수료</th></tr></thead>
<tbody>{rows}</tbody></table></div></div>"""


def report_page(sess, week=None):
    weeks = storage.list_weeks()
    week = week or (weeks[0] if weeks else default_week())
    data = storage.load_week(week)
    if not data or not data["제출"]:
        return page(
            "주간 리포트",
            f"""<h1>③ 주간 리포트</h1>
<div class="card"><p class="note warn">{week_label(week)} 주차에 저장된 제출이 없습니다.
<b>① 직접 입력</b> 또는 <b>① 엑셀 업로드</b>로 먼저 제출해 주세요.</p>
<p><a href="/submit"><button>직접 입력하러 가기</button></a> <a href="/upload"><button class="ghost">엑셀 올리기</button></a></p></div>""",
            "report",
            sess,
        )

    prev = storage.prev_week_key(week)
    ly = storage.last_year_key(week)
    rep = apply_approval(
        build_report(
            data,
            storage.load_week(prev) if prev else None,
            storage.load_week(ly),
            storage.load_roster(),
            week,
        ),
        data,
    )

    src = f"""<div class="card noprint"><form method="get" action="/report" class="inline">
{week_field(week, weeks)}
</form><p class="hint" style="margin-top:10px">비교 자료: 전주 {e(prev) if prev else "없음"} · 작년 동기 {e(ly)}{"" if storage.load_week(ly) else " (저장소에 없음)"} — 저장소에서 자동으로 불러옵니다.</p></div>"""

    return page(
        f"주간 리포트 — {week}",
        phead("훈련실적", f"③ 주간 리포트 — {week_label(week, True)}",
              "흐름도 5~15단계 자동 처리 결과 · 👁 표시는 관리자가 확인하는 지점")
        + f"""{src}
{render_todos(rep)}{render_board(rep)}{render_table(rep)}{render_charts(rep)}{render_yearly(week)}{render_outliers(rep)}
{render_notes(rep)}{render_schedule(rep)}{render_summary(rep, week)}""",
        "report",
        sess,
    )


# ── 화면: 오늘 할 일 · 핵심 요약 ─────────────────────────────────


def week_report(week=None, today=None, src=""):
    """주차 리포트를 저장소에서 만든다. 제출이 없으면 연습용 샘플 엑셀로 대신 돌린다.

    샘플로 돌 때는 요청한 주차의 샘플 파일을 먼저 찾고, 없으면 가장 최근 샘플 주차로
    week 를 바꿔서 돌려준다 — 화면 제목과 실제 자료의 주차가 어긋나지 않게 하기 위함이다.

    반환: (rep, week, weeks, sample) — 샘플 파일까지 없으면 rep 이 None.
    """
    weeks = storage.list_weeks()
    today = today or date.today().isoformat()
    week = week or (weeks[0] if weeks else (sample_weeks() or [default_week()])[-1])
    data = storage.load_week(week)
    sample = src == "sample" or not (data and data["제출"])

    if sample:
        path = sample_for(week)
        if path is None:  # 요청 주차 샘플이 없으면 최신 샘플 주차로 맞춘다
            avail = sample_weeks()
            if not avail:
                return None, week, weeks, True
            week = avail[-1]
            path = sample_for(week)
        prev_p, ly_p = sample_for(prev_sample_week(week)), sample_for(storage.last_year_key(week))
        rep = build_report(
            str(path),
            str(prev_p) if prev_p else None,  # 샘플끼리도 전주 비교·이상치를 태운다
            str(ly_p) if ly_p else None,
            storage.load_roster(),
            week,
            today=today,
        )
    else:
        prev = storage.prev_week_key(week)
        rep = apply_approval(
            build_report(
                data,
                storage.load_week(prev) if prev else None,
                storage.load_week(storage.last_year_key(week)),
                storage.load_roster(),
                week,
                today=today,
            ),
            data,
        )
    return rep, week, weeks, sample


def origin_note(week, sample, what):
    if sample:
        p = sample_for(week)
        prev_w, ly_w = prev_sample_week(week), storage.last_year_key(week)
        extra = " · ".join(
            x for x in [f"전주 {prev_w}" if sample_for(prev_w) else "", f"작년 {ly_w}" if sample_for(ly_w) else ""] if x
        )
        return (
            f'<p class="note">저장된 제출이 없어 연습용 샘플 <code>{e(p.name) if p else "—"}</code> 로 계산한 화면입니다. '
            + (f"비교 자료도 샘플을 씁니다({e(extra)}). " if extra else "비교용 샘플(전주·작년)은 없습니다. ")
            + "실제 제출이 저장되면 저장소 자료로 자동 전환됩니다.</p>"
        )
    return f'<p class="note">저장소 <code>data/{e(week)}.json</code> · 승인·반려 처리까지 반영한 {what}입니다.</p>'


def read_fail_msg(filename, ex):
    """업로드 실패를 담당자가 알아볼 말로 바꾼다. 원인 문구는 맨 뒤에 짧게 남긴다."""
    name = filename or "올린 파일"
    if isinstance(ex, (KeyError, IndexError, StopIteration)):
        hint = "시트 구성이나 머리글이 양식과 다릅니다. 「빈 양식 내려받기」로 받은 파일에 값을 채워 다시 올려 주세요."
    elif "zip" in str(ex).lower() or isinstance(ex, ValueError):
        hint = "엑셀 파일(.xlsx)이 아니거나 파일이 손상됐습니다. 엑셀에서 다시 저장한 뒤 올려 주세요."
    else:
        hint = "파일을 여는 중 문제가 생겼습니다. 엑셀에서 열리는지 확인한 뒤 다시 올려 주세요."
    return f"{name} 을(를) 읽지 못했습니다. {hint} (원인: {type(ex).__name__})"


def week_field(week, weeks):
    """주차 입력 칸. 저장된 주차·샘플 주차를 제안하되 직접 입력도 받는다.

    목록에 없는 주차(아직 자료가 없는 주)도 열어볼 수 있어야 해서 select 가 아니다.
    """
    cand = list(dict.fromkeys(list(weeks) + sample_weeks()))
    opts = "".join(f'<option value="{e(k)}">' for k in cand)
    return f"""<div style="max-width:210px"><label>주차</label>
<input type="text" name="week" list="weeklist" value="{e(week)}" placeholder="2026-W31" onchange="this.form.submit()">
<datalist id="weeklist">{opts}</datalist>
<div class="hint">직접 입력해도 됩니다 (연도-W주차번호) · {WEEK_NOTE}</div></div>"""


def no_source_page(sess, title, active, week):
    return page(
        title,
        f"""<h1>{e(title)}</h1>
<div class="card"><p class="note warn">{week_label(week)} 주차에 저장된 제출이 없고, <code>inputs/</code> 에
연습용 샘플(<code>{e(SAMPLE_PREFIX)}주차.xlsx</code>)도 없습니다.</p>
<p><a href="/submit"><button>직접 입력하러 가기</button></a> <a href="/upload"><button class="ghost">엑셀 올리기</button></a></p></div>""",
        active,
        sess,
    )



def brief_page(sess, week=None, today=None, src=""):
    """리포트 결과에서 핵심 항목만 한 장으로 뽑아 보여준다.

    week_report 로 같은 자료를 쓰고를 쓰고, 여기서 새로 계산하지 않는다.
    각 줄은 '항목 · 값 · 비고' 세 칸이고 클릭하면 원래 화면으로 넘어간다.
    """
    rep, week, weeks, sample = week_report(week, today, src)
    if rep is None:
        return no_source_page(sess, "핵심 요약", "brief", week)
    today = rep["기준일"]

    t = rep["실적"]["합계_전체"][0]
    cmp_, board, todos = rep["비교"], rep["현황판"], rep["오늘할일"]
    roster = storage.load_roster()  # 비어 있으면 제출한 기관이 곧 대상이라 미제출이 잡히지 않는다
    출 = sum(1 for b in board if b["상태"] in ("승인", "검토대기", "반려 대상", "명단 외"))
    승인 = sum(1 for b in board if b["상태"] == "승인")
    막힘 = [b["기관명"] for b in board if b["상태"] in ("미제출", "반려 대상")]
    급함 = sum(1 for x in todos if x["우선순위"] == "높음")
    이상치 = len(rep["이상치"]["플래그"]) if rep["이상치"] else None
    오늘일정 = [s for s in rep["주요일정"] if str(s["날짜"])[:10] == today]

    kpi = (
        '<div class="kpi">'
        + stat("chart", "실시율 (목표 대비)", pct(t["실시율"]), '<span class="chip n">실시 ÷ 연간목표</span>')
        + stat("star", "수료율 (실시 대비)", pct(t["수료율"]), delta_chip(cmp_["전체"]["수료율"] if cmp_ else None))
        + stat("check", "제출·승인", f"{승인}/{len(board)}",
               f'<span>제출 {출}곳 · 대상 {len(board)}곳{"" if roster else " (명단 미등록)"}</span>')
        + stat("star", "먼저 볼 할 일", 급함, f'<span>전체 {len(todos)}건</span>')
        + "</div>"
    )

    def row(item, value, note, warn=False):
        cls = ' class="err"' if warn else ""
        return f'<tr{cls}><td>{item}</td><td class="n"><b>{value}</b></td><td>{note}</td></tr>'

    rows = row("집계 과정", f"{t['과정수']}개", f"검증 오류로 제외 {t['제외']}개" if t["제외"] else "전 과정 집계", bool(t["제외"]))
    rows += row(
        "리포트 발행",
        "보류" if 막힘 else "가능",
        (e(", ".join(막힘)) + " 미완료") if 막힘
        else ("미제출·반려 없음" if roster else '<a href="/admin">대상 명단이 없어 미제출을 잡지 못합니다</a>'),
        bool(막힘) or not roster,
    )
    rows += row(
        "작년 대비 이상치",
        "—" if 이상치 is None else f"{이상치}건",
        "작년 동기 자료 없음" if 이상치 is None else f"수료율 {int(OUTLIER_THRESHOLD * 100)}%p 이상 차이",
        bool(이상치),
    )
    rows += row(
        "확인 필요 특이사항",
        f"{rep['확인필요수']}건",
        f"보고된 특이사항 전체 {len(rep['특이사항'])}건",
        bool(rep["확인필요수"]),
    )
    rows += row(
        f"오늘({today}) 일정",
        f"{len(오늘일정)}건",
        e(" · ".join(f"{s['기관명']} {s['구분']}" for s in 오늘일정)) or "없음",
    )

    # 엑셀은 저장소 자료로만 만든다. 샘플로 보는 중에는 눌러도 실패하므로 버튼을 내린다.
    excel = (
        f'<a href="/export?week={quote(week)}&kind=report"><button class="sm ghost">리포트 엑셀 받기</button></a>'
        if not sample
        else '<span class="hint" style="align-self:center">엑셀 내려받기는 저장된 제출이 있어야 됩니다</span>'
    )
    picker = f"""<div class="card noprint"><form method="get" action="/brief" class="inline">
{week_field(week, weeks)}
<div style="max-width:200px"><label>기준일</label><input type="date" name="date" value="{e(today)}" onchange="this.form.submit()">
<div class="hint">이 날짜 기준으로 일정·할 일을 셉니다.</div></div>
<input type="hidden" name="src" value="{"sample" if sample else ""}">
</form>
<p style="margin-top:12px;display:flex;gap:8px;flex-wrap:wrap">
<a href="/report?week={quote(week)}"><button class="sm ghost">상세 리포트 보기</button></a>
{excel}</p>
{origin_note(week, sample, "요약")}</div>"""

    return page(
        f"핵심 요약 — {week}",
        phead("훈련실적", f'핵심 요약 <span class="eye">👁 {week_label(week, True)}</span>',
              "주간 리포트 결과에서 먼저 볼 항목만 한 장으로 추렸습니다.")
        + f"""{picker}
<div class="card"><div class="hd"><h2>핵심 항목 <span class="eye">👁 기준일 {e(today)}</span></h2>
<p class="sub" style="margin:0">값은 리포트 계산 결과 그대로이며 여기서 다시 계산하지 않습니다.</p></div>
{kpi}
<div class="scroll"><table><thead><tr><th>항목</th><th class="n">값</th><th>비고</th></tr></thead>
<tbody>{rows}</tbody></table></div>
<p class="note{' warn' if (막힘 or 급함) else ' ok'}">{
    f"조치가 필요한 항목이 있습니다 — 오늘 먼저 볼 것 {급함}건." if (막힘 or 급함) else "지금 막고 있는 항목은 없습니다."
}</p></div>""",
        "brief",
        sess,
    )


# ── 엑셀 내보내기 ────────────────────────────────────────────────


def _sheet(ws, headers, rows):
    hf, hl = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="4472C4")
    ws.append(headers)
    for c in ws[1]:
        c.font, c.fill = hf, hl
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = max(11, min(46, len(str(h)) * 2 + 7))


TEMPLATE_NAME = "훈련기관_실적제출_간소양식.xlsx"
TEMPLATE_ROWS = 300  # 입력 제한(드롭다운·정수 검사)을 걸어 둘 행 범위

# 정부 원본과 같은 머리글 이름 — read_rows(HEADER_MAP)가 이 이름으로 열을 찾는다.
양성_머리글 = ["순번", "훈련센터명", "훈련기관명", "정기/수시", "과정구분", "NCS대분류명", "KECO세분류명",
             "훈련과정명", "훈련인원", "실시인원", "중도탈락인원", "훈련중", "수료인원", "취업인원"]
향상_머리글 = ["순번", "훈련센터명", "훈련기관명", "정기/수시", "과정구분", "NCS대분류명", "KECO세분류명",
             "훈련과정명", "훈련인원", "실시인원", "중도탈락", "훈련중", "수료인원"]


def template_xlsx():
    """훈련기관이 채워서 올릴 수 있는 간소 빈 양식.

    정부 원본 파일이 있으면 그대로 올리면 되고, 없을 때만 이 파일을 쓴다.
    시트·머리글 이름은 read_rows(HEADER_MAP)·read_targets 가 찾는 이름 그대로다.
    """
    wb = Workbook()

    # ── 양성/향상 현황 (필수 중 하나 이상) ──
    for idx, (sheet, 머리글) in enumerate([("양성훈련 현황", 양성_머리글), ("향상훈련 현황", 향상_머리글)]):
        ws = wb.active if idx == 0 else wb.create_sheet(sheet)
        ws.title = sheet
        _sheet(ws, 머리글, [])
        # 정기/수시 드롭다운
        col = chr(ord("A") + 머리글.index("정기/수시"))
        d = DataValidation(type="list", formula1='"정기,수시"', allow_blank=True)
        ws.add_data_validation(d)
        d.add(f"{col}2:{col}{TEMPLATE_ROWS}")
        # 인원 칸(훈련인원~끝)은 0 이상 정수
        dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1=0, allow_blank=True)
        dv.error = "인원은 0 이상의 정수만 입력합니다."
        ws.add_data_validation(dv)
        start = chr(ord("A") + 머리글.index("훈련인원"))
        end = chr(ord("A") + len(머리글) - 1)
        dv.add(f"{start}2:{end}{TEMPLATE_ROWS}")

    # ── 교육실적 (기관 연간 목표) ── read_targets 가 라벨 오른쪽 숫자를 읽는다
    ws3 = wb.create_sheet(TARGET_SHEET)
    ws3.append(["지산맞인력양성사업 목표 훈련인원 (명)", None,
                "정기(채용예정자)", None, "정기(재직자향상)", None, "수시", None])
    ws3.append(["※ 위 빈칸(회색 오른쪽)에 목표 인원을 숫자로 채우세요. 비우면 정원 합으로 대체 계산합니다."])
    ws3.column_dimensions["A"].width = 34

    # ── 작성 방법 ──
    ws4 = wb.create_sheet("작성 방법")
    for row in [
        ["■ 정부 「지산맞 훈련실적」 원본이 있으면 이 파일 대신 원본을 그대로 올리셔도 됩니다."],
        [""],
        ["1. 과정별 실적은 「양성훈련 현황」·「향상훈련 현황」 두 시트에 채웁니다(해당 시트만 채워도 됩니다)."],
        ["2. 시트 이름과 머리글은 바꾸지 마세요. 열 순서는 바꿔도 되고, 열을 더 붙여도 됩니다(이름으로 찾습니다)."],
        ["3. 한 줄 = 한 과정. 훈련과정명·훈련센터명(기관)·훈련인원·실시인원·수료인원·중도탈락은 필수입니다."],
        ["4. 취업인원은 양성 과정만 채웁니다(향상은 칸이 없습니다)."],
        ["5. 주차는 엑셀에 적지 않습니다. 업로드 화면에서 연·월·주차번호를 고르세요."],
        [""],
        ["■ 「교육실적」 시트의 목표 훈련인원(정기 채용예정자=양성, 정기 재직자향상=향상, 수시)을 채우면"],
        ["  실시율·수료율(목표 대비)의 분모로 씁니다. 비우면 과정 정원(훈련인원) 합으로 대체합니다."],
        [""],
        ["■ 앱이 자동 계산합니다: 실시율=실시÷목표, 수료율(실시대비)=수료÷실시, 탈락률=중도탈락÷실시, 취업률=취업÷수료."],
        ["  비율을 엑셀에 직접 적지 마세요(쓰이지 않습니다)."],
        [""],
        ["■ 숫자가 맞아도 오류로 걸리는 경우: 수료인원>실시인원, 중도탈락>실시인원."],
    ]:
        ws4.append(row)
    ws4.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_raw(week):
    """입력된 값 그대로 — 양성·향상 과정 전체를 통합 세부내역으로."""
    data = storage.load_week(week)
    wb = Workbook()
    y, w = storage.parse_week(week)
    m = data.get("월") or ""

    rows = []
    for org, s in data["제출"].items():
        for p in s["실적"]:
            rows.append(
                [y, m, w, week, org, p.get("구분"), p.get("정기수시"), p.get("과정구분"),
                 p.get("NCS대분류명"), p.get("KECO세분류명"), p.get("과정명"),
                 p.get("훈련목표인원"), p.get("훈련실시인원"), p.get("중도탈락자"),
                 p.get("훈련중"), p.get("훈련수료인원"), p.get("취업인원"), s["출처"], s["제출시각"], s["상태"]]
            )
    _sheet(wb.active, ["연", "월", "주차번호", "주차", "기관명", "구분", "정기/수시", "과정구분",
                       "NCS대분류명", "KECO세분류명", "과정명",
                       "훈련목표인원", "훈련실시인원", "중도탈락자", "훈련중", "훈련수료인원", "취업인원",
                       "제출방식", "제출시각", "상태"], rows)
    wb.active.title = "세부내역"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_report(week):
    """정부 「기관별 합계」 양식의 결과표 — 총계(기관 소계+양성/향상/수시) + 세부실적."""
    data = storage.load_week(week)
    prev = storage.prev_week_key(week)
    rep = build_report(data, storage.load_week(prev) if prev else None,
                       storage.load_week(storage.last_year_key(week)), storage.load_roster(), week)
    res = rep["실적"]
    wb = Workbook()
    p100 = lambda v: round(v * 100, 1) if isinstance(v, float) else (v if isinstance(v, str) else None)

    # ── 훈련실적 총계 ── 전체 계 + 기관별 소계 + 그 기관의 양성/향상/수시
    구분별 = {g["구분"]: g for g in res["합계_기관구분별"]}  # "기관 · 양성" 형태
    총계행 = []

    def 지표행(라벨, g):
        return [라벨, g["과정수"], g["목표훈련인원"], g["훈련실시인원"], p100(g["실시율"]),
                g["중도탈락자"], g["훈련수료인원"], p100(g["수료율"]), g["취업인원"], p100(g["취업률"])]

    총계행.append(지표행("총계 · 계", res["합계_전체"][0]))
    for gi in res["합계_기관별"]:
        기관 = gi["구분"]
        총계행.append(지표행(f"{기관} · 소계", gi))
        for bucket in ("양성", "향상", "수시"):
            sub = 구분별.get(f"{기관} · {bucket}")
            if sub:
                총계행.append(지표행(f"　· {bucket}", sub))
    _sheet(wb.active, ["구분", "과정수", "목표인원", "실시인원", "실시율(%)",
                       "중도탈락", "수료인원", "수료율(%)", "취업인원", "취업률(%)"], 총계행)
    wb.active.title = "훈련실적 총계"

    # ── 세부실적 ── 과정 단위
    _sheet(wb.create_sheet("세부실적"),
           ["오류", "기관명", "구분", "정기/수시", "과정구분", "NCS대분류명", "KECO세분류명", "과정명",
            "목표(정원)", "실시인원", "중도탈락", "훈련중", "수료인원", "취업인원", "수료율(%)", "탈락률(%)"],
           [["오류" if r["_오류"] else "", r["기관명"], r["구분"], r["정기수시"], r["과정구분"],
             r["NCS대분류명"], r["KECO세분류명"], r["과정명"],
             r["훈련목표인원"], r["훈련실시인원"], r["중도탈락자"], r["훈련중"], r["훈련수료인원"], r["취업인원"],
             p100(r["수료율"]), p100(r["탈락률"])] for r in res["표"]])

    for title, key in [("NCS대분류별", "합계_NCS별"), ("KECO세분류별", "합계_KECO별")]:
        _sheet(wb.create_sheet(title),
               ["구분", "과정수", "제외", "목표인원", "실시인원", "수료인원", "중도탈락", "수료율(%)", "탈락률(%)"],
               [[g["구분"], g["과정수"], g["제외"], g["훈련목표인원"], g["훈련실시인원"], g["훈련수료인원"],
                 g["중도탈락자"], p100(g["수료율"]), p100(g["탈락률"])] for g in res[key]])

    if rep["이상치"] and rep["이상치"]["플래그"]:
        _sheet(wb.create_sheet("이상치"), ["기관명", "과정명", "이번(%)", "작년동기(%)", "차이(%p)", "방향"],
               [[f["기관명"], f["과정명"], round(f["이번"] * 100, 1), round(f["작년"] * 100, 1),
                 round(f["차이"], 1), f["방향"]] for f in rep["이상치"]["플래그"]])

    _sheet(wb.create_sheet("특이사항"), ["기관명", "과정명", "분류", "내용", "확인필요"],
           [[n["기관명"], n["과정명"], n["분류"], n["내용"], n["확인필요"]] for n in rep["특이사항"]])
    _sheet(wb.create_sheet("주요일정"), ["날짜", "기관명", "구분", "내용"],
           [[s["날짜"], s["기관명"], s["구분"], s["내용"]] for s in rep["주요일정"]])
    _sheet(wb.create_sheet("오류목록"), ["행", "컬럼", "사유"], [list(x) for x in res["오류"]])
    _sheet(wb.create_sheet("오늘할일"), ["우선순위", "구분", "대상", "할 일"],
           [[t["우선순위"], t["구분"], t["대상"], t["할일"]] for t in rep["오늘할일"]])

    ws = wb.create_sheet("요약초안")
    ws.append(["요약 초안"])
    ws["A1"].font = Font(bold=True)
    for line in rep["요약초안"].split("\n"):
        ws.append([line])
    ws.column_dimensions["A"].width = 120
    for row in ws.iter_rows():
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 폼 파싱 ──────────────────────────────────────────────────────


def parse_multipart(headers, body):
    raw = b"Content-Type: " + headers["Content-Type"].encode() + b"\r\n\r\n" + body
    msg = email.message_from_bytes(raw, policy=email_policy)
    out = {}
    for part in msg.iter_parts():
        name = part.get_param("name", header="content-disposition")
        if name:
            out[name] = (part.get_filename(), part.get_payload(decode=True) or b"")
    return out


def key_from_form(year, month, mweek, fallback=None):
    """폼의 연·월·(그 달의) 주차 → 저장용 주차 키.

    화면에서 자동 계산한 값(fallback)이 함께 오지만, 믿지 않고 서버에서 다시 계산한다.
    반환: (키, 오류문구) — 키가 None 이면 오류문구를 그대로 보여준다.
    """
    try:
        y, m, n = int(year), int(month), int(mweek)
    except (TypeError, ValueError):
        if fallback:
            try:
                storage.parse_week(fallback)
                return fallback, None
            except Exception:
                pass
        return None, "연·월·주차를 숫자로 입력해 주세요."
    key, 일요일 = month_week_to_key(y, m, n)
    if key is None:
        return None, f"{y}년 {m}월에는 {n}주차가 없습니다. 주차를 다시 고르세요."
    return key, None


def zip_rows(form, mapping, required):
    """같은 이름으로 여러 번 온 입력을 행 목록으로 묶는다."""
    cols = {dst: form.get(src, []) for dst, src in mapping.items()}
    n = max((len(v) for v in cols.values()), default=0)
    rows = []
    for i in range(n):
        row = {dst: (vals[i].strip() if i < len(vals) else "") for dst, vals in cols.items()}
        if not any(row.get(k) for k in required):
            continue
        rows.append(row)
    return rows


def to_num(s):
    try:
        return int(str(s).strip())
    except (ValueError, TypeError):
        return str(s).strip() or None


# ── 서버 ─────────────────────────────────────────────────────────

# 로그인 시도 제한 (외부 공개 시 비밀번호 대입 시도를 늦춘다). 메모리라 재시작하면 초기화.
LOGIN_LIMIT, LOGIN_WINDOW = 10, 600  # 10분에 10회
LOGIN_FAILS = {}  # ip -> [실패 시각, ...]


def login_quota(ip):
    """남은 시도 횟수. 창(10분)이 지난 기록은 버린다."""
    now = time.time()
    fails = [t for t in LOGIN_FAILS.get(ip, []) if now - t < LOGIN_WINDOW]
    LOGIN_FAILS[ip] = fails
    return LOGIN_LIMIT - len(fails)


def login_fail(ip):
    LOGIN_FAILS.setdefault(ip, []).append(time.time())


# 챗봇 호출 제한 (외부 공개 시 API 요금이 새지 않게). 메모리라 재시작하면 초기화.
CHAT_LIMIT, CHAT_WINDOW = 30, 600  # 10분에 30회
CHAT_HITS = {}  # ip -> [호출 시각, ...]


def chat_quota(ip):
    now = time.time()
    hits = [t for t in CHAT_HITS.get(ip, []) if now - t < CHAT_WINDOW]
    CHAT_HITS[ip] = hits
    return CHAT_LIMIT - len(hits)


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        print(f"  {self.command} {self.path}")

    def send(self, code, body, ctype="text/html; charset=utf-8", extra=None):
        data = body.encode("utf-8") if isinstance(body, str) else body
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        for k, v in (extra or {}).items():
            self.send_header(k, v)
        self.end_headers()
        self.wfile.write(data)

    def redirect(self, to):
        self.send_response(303)
        self.send_header("Location", to)
        self.end_headers()

    def error_page(self, msg, back="/"):
        self.send(400, page("오류", f"""<div class="card"><h2>처리할 수 없습니다</h2>
<p class="note warn">{e(msg)}</p><p><a href="{e(back)}"><button class="ghost">돌아가기</button></a></p></div>"""))

    def q(self):
        return {k: v[0] for k, v in parse_qs(urlparse(self.path).query).items()}

    def form(self):
        n = int(self.headers.get("Content-Length", 0))
        return parse_qs(self.rfile.read(n).decode("utf-8"), keep_blank_values=True)

    def json_body(self):
        n = int(self.headers.get("Content-Length", 0) or 0)
        if n <= 0 or n > 200_000:  # 챗 기록이 지나치게 크면 받지 않는다
            return {}
        try:
            return json.loads(self.rfile.read(n).decode("utf-8"))
        except (ValueError, UnicodeDecodeError):
            return {}

    def sess(self):
        return auth.session_of(self.headers.get("Cookie"))

    def deny(self):
        self.send(403, page("권한 없음", """<div class="card"><h2>접근 권한이 없습니다</h2>
<p class="note warn">이 화면은 관리자만 볼 수 있습니다. 기관 계정은 자기 기관의 입력·조회만 가능합니다.</p>
<p><a href="/mine"><button class="ghost">내 제출 내역으로</button></a></p></div>"""))

    # ── GET ──
    def do_GET(self):
        p = urlparse(self.path).path
        q = self.q()
        s = self.sess()

        if p == "/login":
            return self.send(200, login_page())
        if p == "/logout":
            for chunk in (self.headers.get("Cookie") or "").split(";"):
                k, _, v = chunk.strip().partition("=")
                if k == "sid":
                    auth.logout(v)
            return self.redirect("/login")
        if not s:
            return self.redirect("/login")

        admin_only = {"/admin", "/report", "/brief", "/users", "/export"}
        if p in admin_only and s["role"] != "admin":
            return self.deny()

        try:
            if p == "/m":  # 모바일 전용 화면
                return self.send(200, m_admin(s) if s["role"] == "admin" else m_org(s))
            if p == "/view":  # PC/모바일 화면 선택 (쿠키로 기억)
                to = "m" if q.get("to") == "m" else "pc"
                self.send_response(303)
                self.send_header("Location", "/m" if to == "m" else "/")
                self.send_header("Set-Cookie", f"view={to}; Path=/; Max-Age=31536000; SameSite=Lax")
                return self.end_headers()
            if p == "/":
                if is_mobile(self.headers):  # 휴대폰이면 전용 화면으로
                    return self.redirect("/m")
                return self.send(200, home_page(s) if s["role"] == "admin" else mine_page(s))
            if p == "/mine":
                return self.send(200, mine_page(s) if s["role"] != "admin" else home_page(s))
            if p == "/users":
                return self.send(200, users_page(s))
            if p == "/issues":
                return self.send(200, issues_page(s))
            if p == "/stats":
                return self.send(200, stats_page(s))
            if p == "/advice":
                return self.send(200, advice_page(s))
            if p == "/calendar":
                if not q.get("id"):
                    return self.send(200, calendar_page(s))
                ev = calendar_store.get(q.get("id"))
                if not ev:
                    return self.error_page("없는 일정 조사입니다.", "/calendar")
                if not calendar_store.can_open(ev, s["role"], s["org"], storage.load_roster()):
                    return self.deny()
                return self.send(200, cal_event_page(s, ev))
            if p == "/submit":
                return self.send(200, submit_page(s, q.get("week"), q.get("org")))
            if p == "/upload":
                return self.send(200, upload_page(s))
            if p == "/template":
                return self.send(200, template_xlsx(),
                                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 {"Content-Disposition": attachment(TEMPLATE_NAME)})
            if p == "/admin":
                return self.send(200, admin_page(s, q.get("week")))
            if p == "/report":
                return self.send(200, report_page(s, q.get("week")))
            if p == "/brief":
                return self.send(200, brief_page(s, q.get("week"), q.get("date"), q.get("src", "")))
            if p == "/export":
                week = q.get("week")
                if not storage.load_week(week):
                    return self.error_page("해당 주차에 저장된 자료가 없습니다.")
                kind = q.get("kind", "raw")
                data = export_raw(week) if kind == "raw" else export_report(week)
                fn = f"{week}-{'input' if kind == 'raw' else 'report'}.xlsx"
                return self.send(200, data,
                                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 {"Content-Disposition": attachment(fn)})
        except Exception as ex:
            return self.error_page(f"{type(ex).__name__}: {ex}")
        self.send(404, "not found", "text/plain; charset=utf-8")

    # ── POST ──
    def do_POST(self):
        p = urlparse(self.path).path
        if p == "/login":
            return self.post_login()
        s = self.sess()
        if not s:
            return self.redirect("/login")
        if p in {"/status", "/roster", "/users"} and s["role"] != "admin":
            return self.deny()
        try:
            if p == "/submit":
                return self.post_submit(s)
            if p == "/upload":
                return self.post_upload(s)
            if p == "/users":
                return self.post_users(s)
            if p == "/issues":
                return self.post_issues(s)
            if p == "/stats":
                return self.post_stats(s)
            if p == "/calendar":
                return self.post_calendar(s)
            if p == "/chat":
                return self.post_chat(s)
            if p == "/status":
                f = self.form()
                week, org, st = f["week"][0], f["org"][0], f["status"][0]
                storage.delete_submission(week, org) if st == "삭제" else storage.set_status(week, org, st)
                return self.redirect(f"/admin?week={quote(week)}")
            if p == "/roster":
                f = self.form()
                storage.save_roster(f.get("roster", [""])[0].splitlines())
                return self.redirect(f"/admin?week={quote(f.get('week', [default_week()])[0])}")
        except Exception as ex:
            return self.error_page(f"{type(ex).__name__}: {ex}")
        self.send(404, "not found", "text/plain; charset=utf-8")

    def client_ip(self):
        """프록시(HTTPS 터널) 뒤에서도 원래 주소를 본다."""
        fwd = self.headers.get("X-Forwarded-For", "")
        return fwd.split(",")[0].strip() or self.client_address[0]

    def post_login(self):
        ip = self.client_ip()
        남은 = login_quota(ip)
        if 남은 <= 0:  # 외부에 열어 두면 비밀번호를 찍어 보는 시도가 들어온다
            return self.send(429, login_page(
                '<p class="note warn">로그인 시도가 너무 많습니다. 잠시(10분) 뒤에 다시 시도해 주세요.</p>'))
        f = self.form()
        uid, pw = f.get("id", [""])[0].strip(), f.get("pw", [""])[0]
        user = auth.verify(uid, pw)
        if not user:
            login_fail(ip)
            return self.send(401, login_page(
                f'<p class="note warn">아이디 또는 비밀번호가 맞지 않습니다. (남은 시도 {남은 - 1}회)</p>', uid))
        LOGIN_FAILS.pop(ip, None)
        sid = auth.login(user)
        보안 = "; Secure" if self.headers.get("X-Forwarded-Proto") == "https" else ""
        self.send_response(303)
        self.send_header("Location", "/")
        self.send_header("Set-Cookie", f"sid={sid}; Path=/; HttpOnly; SameSite=Lax{보안}")
        self.end_headers()

    def post_issues(self, sess):
        topic = self.form().get("topic", [""])[0].strip()
        try:
            cache = gb_issues.collect([topic] if topic else None)
        except gb_issues.IssueError as ex:
            return self.send(200, issues_page(sess, f'<p class="note warn">{e(str(ex))}</p>'))
        건수 = sum(len(b["이슈"]) for b in gb_issues.ordered(cache))
        묶음 = topic or f"주제 {len(gb_issues.TOPICS)}개"
        return self.send(200, issues_page(sess, f'<p class="note ok">{e(묶음)} 수집 완료 — 지금 저장된 이슈 {건수}건.</p>'))

    def post_stats(self, sess):
        f = self.form()
        regions = [r for r in f.get("region", []) if r in kosis_stats.ALL_REGIONS]
        try:
            months = max(1, min(120, int(f.get("months", [""])[0])))
        except ValueError:
            months = kosis_stats.DEFAULT_MONTHS
        잘림 = ""
        if len(regions) > kosis_stats.MAX_SERIES:
            잘림 = f' 선택한 지역이 많아 앞의 {kosis_stats.MAX_SERIES}개만 그렸습니다.'
            regions = regions[: kosis_stats.MAX_SERIES]
        try:
            data = kosis_stats.collect(months, regions or None)
        except kosis_stats.StatError as ex:
            return self.send(200, stats_page(sess, f'<p class="note warn">{e(str(ex))}</p>'))
        기간 = f'{kosis_stats.fmt_period(data["시점"][0])} ~ {kosis_stats.fmt_period(data["시점"][-1])}'
        return self.send(200, stats_page(sess, f'<p class="note ok">KOSIS 에서 {e(기간)} 자료를 받았습니다.{e(잘림)}</p>'))

    def post_users(self, sess):
        f = self.form()
        do = f.get("do", [""])[0]
        uid, pw = f.get("id", [""])[0].strip(), f.get("pw", [""])[0]
        if do == "create":
            role, org = f.get("role", ["org"])[0], f.get("org", [""])[0]
            if role == "org" and not org.strip():
                return self.send(200, users_page(sess, '<p class="note warn">기관 계정에는 소속 기관이 필요합니다.</p>'))
            if not auth.create_user(uid, pw, role, org):
                return self.send(200, users_page(sess, f'<p class="note warn">이미 있는 아이디입니다 — {e(uid)}</p>'))
            msg = f'<p class="note ok">계정을 발급했습니다 — <b>{e(uid)}</b>. 비밀번호는 다시 볼 수 없으니 지금 전달하세요.</p>'
        elif do == "reset":
            auth.set_password(uid, pw)
            msg = f'<p class="note ok">{e(uid)} 비밀번호를 변경했습니다.</p>'
        elif do == "delete":
            ok = auth.delete_user(uid)
            msg = (
                f'<p class="note ok">{e(uid)} 계정을 삭제했습니다.</p>'
                if ok
                else '<p class="note warn">삭제할 수 없습니다. 마지막 관리자 계정은 지울 수 없습니다.</p>'
            )
        else:
            msg = ""
        self.send(200, users_page(sess, msg))

    def send_json(self, code, obj):
        self.send(code, json.dumps(obj, ensure_ascii=False), "application/json; charset=utf-8")

    def post_chat(self, sess):
        """서비스 안내 챗봇. 방문자의 질문을 chatbot.answer 로 넘겨 답을 만든다.

        답·오류를 모두 JSON 으로 돌려준다(위젯의 fetch 가 받는다). HTML 오류 페이지로
        새지 않게 여기서 예외를 잡는다.
        """
        if chat_quota(self.client_ip()) <= 0:
            return self.send_json(429, {"error": "질문이 잠시 많았습니다. 10분 뒤에 다시 물어봐 주세요."})
        body = self.json_body()
        history = body.get("messages") if isinstance(body, dict) else None
        if not isinstance(history, list):
            return self.send_json(400, {"error": "질문을 입력해 주세요."})
        CHAT_HITS.setdefault(self.client_ip(), []).append(time.time())
        try:
            reply = chatbot.answer(history)
            return self.send_json(200, {"reply": reply})
        except chatbot.ChatError as ex:
            return self.send_json(200, {"error": str(ex)})
        except Exception as ex:
            return self.send_json(500, {"error": f"답변 중 문제가 생겼습니다 — {type(ex).__name__}"})

    def post_calendar(self, sess):
        """일정 조사 — 만들기·설정·삭제는 관리자, 응답은 대상 기관."""
        f = self.form()
        do = f.get("do", [""])[0]
        관리자용 = {"create", "options", "delete"}
        if do in 관리자용 and sess["role"] != "admin":
            return self.deny()

        if do == "create":
            ev, 오류 = calendar_store.create(
                f.get("title", [""])[0], f.get("start", [""])[0], f.get("end", [""])[0],
                설명=f.get("body", [""])[0], 마감일=f.get("due", [""])[0].strip(),
                대상=f.get("target", []), 공개=f.get("public", ["N"])[0] == "Y", 작성자=sess["id"],
            )
            if not ev:
                return self.send(200, calendar_page(sess, f'<p class="note warn">{e(오류)}</p>'))
            return self.redirect(f"/calendar?id={ev['번호']}")

        번호 = f.get("id", [""])[0]
        ev = calendar_store.get(번호)
        if not ev:
            return self.error_page("없는 일정 조사입니다.", "/calendar")

        if do == "options":
            ev2, 오류 = calendar_store.set_options(
                번호, 공개=f.get("public", ["N"])[0] == "Y", 마감일=f.get("due", [""])[0].strip()
            )
            if not ev2:
                return self.send(200, cal_event_page(sess, ev, f'<p class="note warn">{e(오류)}</p>'))
            공개말 = "전체 공개" if ev2["공개"] else "관리자만 보기"
            return self.send(200, cal_event_page(sess, ev2, f'<p class="note ok">설정을 저장했습니다 — {공개말}.</p>'))

        if do == "delete":
            calendar_store.delete(번호)
            return self.send(200, calendar_page(sess, '<p class="note ok">일정 조사를 삭제했습니다.</p>'))

        if do == "answer":
            org = sess["org"] if sess["role"] != "admin" else ""
            if not calendar_store.is_target(ev, org, storage.load_roster()):
                return self.deny()  # 관리자·대상 아닌 기관은 대신 답할 수 없다
            날짜별 = {d: f.get(f"a_{d}", ["미정"])[0] for d in calendar_store.days_of(ev)}
            ev2, 오류 = calendar_store.answer(번호, org, 날짜별, f.get("memo", [""])[0], sess["id"])
            if not ev2:
                return self.send(200, cal_event_page(sess, ev, f'<p class="note warn">{e(오류)}</p>'))
            가능 = sum(1 for v in 날짜별.values() if v == "가능")
            return self.send(200, cal_event_page(
                sess, ev2, f'<p class="note ok">응답을 저장했습니다 — 가능 {가능}일. 관리자가 확인합니다.</p>'))

        return self.error_page("알 수 없는 요청입니다.", "/calendar")

    def post_submit(self, sess):
        f = self.form()
        org = f.get("org", [""])[0].strip()
        if sess["role"] != "admin":
            org = sess["org"]  # 기관 계정은 자기 기관으로 강제
        if not org:
            return self.error_page("훈련기관명이 비어 있습니다.", "/submit")
        month = f.get("month", [""])[0].strip() or None
        key, 오류 = key_from_form(f.get("year", [""])[0], month, f.get("mweek", [""])[0],
                                  f.get("week", [""])[0].strip())
        if not key:
            return self.error_page(오류, "/submit")

        perf = zip_rows(f, {"구분": "p_type", "정기수시": "p_reg", "과정구분": "p_kind",
                            "NCS대분류명": "p_ncs", "KECO세분류명": "p_keco", "과정명": "p_course",
                            "훈련목표인원": "p_goal", "훈련실시인원": "p_run",
                            "훈련수료인원": "p_done", "중도탈락자": "p_drop",
                            "훈련중": "p_ing", "취업인원": "p_emp"}, ["과정명"])
        for r in perf:
            for c in ("훈련목표인원", "훈련실시인원", "훈련수료인원", "중도탈락자", "훈련중", "취업인원"):
                r[c] = to_num(r[c])
        if not perf:
            return self.error_page("실적이 한 건도 입력되지 않았습니다.", "/submit")

        notes = zip_rows(f, {"과정명": "n_course", "분류": "n_cat", "내용": "n_body", "확인필요": "n_chk"}, ["내용"])
        plans = zip_rows(f, {"날짜": "s_date", "구분": "s_kind", "내용": "s_body"}, ["날짜", "내용"])

        # 직접 입력은 연간 목표를 따로 받지 않으므로, 집계기가 정원(훈련목표인원) 합으로 대체한다.
        storage.save_submission(key, org, month, perf, notes, plans, "직접 입력", targets=None)
        data = storage.load_week(key)
        only = {"제출": {org: data["제출"][org]}}
        self.send(200, preview_page(sess, key, [org], process_rows(storage.to_rows(only)), "직접 입력"))

    def post_upload(self, sess):
        n = int(self.headers.get("Content-Length", 0))
        fields = parse_multipart(self.headers, self.rfile.read(n))
        fn, blob = fields.get("file", (None, b""))
        if not blob:
            return self.error_page("파일이 없습니다.", "/upload")
        get = lambda k, d="": (fields.get(k, (None, b""))[1] or b"").decode("utf-8").strip() or d
        month = get("month") or None
        key, 오류 = key_from_form(get("year"), month, get("mweek"), get("week"))
        if not key:
            return self.error_page(오류, "/upload")

        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "up.xlsx"
            path.write_bytes(blob)
            try:
                rows = read_rows(path)                 # 양성·향상 현황 (기관 연간 목표 포함)
                targets = read_targets(load_workbook(path, data_only=True))  # 교육실적 목표
                notes, plans = read_notes(path), read_schedule(path)  # 이 양식엔 없으면 빈 목록
            except SystemExit as ex:
                return self.error_page(f"엑셀을 읽지 못했습니다 — {ex}", "/upload")
            except Exception as ex:
                return self.error_page(read_fail_msg(fn, ex), "/upload")

        by_org = {}
        for r in rows:
            org = (r.get("기관명") or "(기관명 미기재)").strip() or "(기관명 미기재)"
            by_org.setdefault(org, {"실적": [], "특이사항": [], "주요일정": []})
            by_org[org]["실적"].append({c: r.get(c) for c in storage.PERF_COLS})
        for x in notes:
            org = (x.get("기관명") or "(기관명 미기재)").strip() or "(기관명 미기재)"
            by_org.setdefault(org, {"실적": [], "특이사항": [], "주요일정": []})["특이사항"].append(
                {c: x.get(c, "") for c in storage.NOTE_COLS})
        for x in plans:
            org = (x.get("기관명") or "(기관명 미기재)").strip() or "(기관명 미기재)"
            by_org.setdefault(org, {"실적": [], "특이사항": [], "주요일정": []})["주요일정"].append(
                {c: x.get(c, "") for c in storage.PLAN_COLS})

        if not by_org:
            return self.error_page("양성·향상 현황 시트에서 읽을 과정이 없습니다.", "/upload")

        skipped = []
        if sess["role"] != "admin":  # 기관 계정은 자기 기관 행만 저장
            mine = sess["org"]
            skipped = [o for o in by_org if o != mine]
            by_org = {o: v for o, v in by_org.items() if o == mine}
            if not by_org:
                return self.error_page(
                    f"파일에 '{mine}' 기관(훈련센터명)의 과정이 없습니다. 다른 기관 자료는 올릴 수 없습니다"
                    + (f" (건너뜀: {', '.join(skipped)})." if skipped else "."),
                    "/upload",
                )

        # 기관 연간 목표는 파일 전체에서 하나로 읽힌다 — 파일의 기관이 하나일 때만 붙인다.
        for org, v in by_org.items():
            org_targets = targets if len(by_org) == 1 else None
            storage.save_submission(key, org, month, v["실적"], v["특이사항"], v["주요일정"],
                                    "엑셀 업로드", targets=org_targets)

        data = storage.load_week(key)
        only = {"제출": {o: data["제출"][o] for o in by_org}}
        src = f"엑셀 업로드 ({fn})" + (f" · 다른 기관 {len(skipped)}곳 제외됨" if skipped else "")
        self.send(200, preview_page(sess, key, list(by_org), process_rows(storage.to_rows(only)), src))


def lan_ips():
    """이 PC 가 사내망에서 받은 주소들. 외부 접속 안내에 쓴다."""
    out = []
    try:
        import socket as _s
        for info in _s.getaddrinfo(_s.gethostname(), None, _s.AF_INET):
            ip = info[4][0]
            if not ip.startswith("127.") and ip not in out:
                out.append(ip)
    except Exception:
        pass
    return out


if __name__ == "__main__":
    args = sys.argv[1:]
    공개 = any(a in ("--public", "--host", "-p") for a in args) or "0.0.0.0" in args
    브라우저 = "--no-browser" not in args
    포트인자 = [a for a in args if a.isdigit()]
    port = int(포트인자[0]) if 포트인자 else PORT
    host = "0.0.0.0" if 공개 else "127.0.0.1"  # 기본은 이 PC 에서만 접속
    url = f"http://localhost:{port}"
    storage.DATA_DIR.mkdir(exist_ok=True)
    first_pw = auth.ensure_admin()
    print(f"\n주간 훈련기관 교육실적 취합 App\n  {url}\n  저장소: {storage.DATA_DIR}")
    if 공개:
        print("\n  ┌─ 외부 접속 허용 모드 (--public) ──────────────")
        for ip in lan_ips():
            print(f"  │  같은 망에서: http://{ip}:{port}")
        print(
            f"  │  인터넷에서 : 공유기 {port}번 포트를 이 PC 로 넘기거나\n"
            "  │               터널(cloudflared·ngrok)을 쓰십시오.\n"
            "  │  ⚠ HTTP 라 비밀번호가 평문으로 오갑니다.\n"
            "  │    인터넷에 열 때는 반드시 HTTPS 를 앞에 두십시오.\n"
            "  │    자세한 절차: remote-access.md\n"
            "  └───────────────────────────────────────────────"
        )
    else:
        print(
            "\n  ┌─ 로컬 전용 모드 (기본값) ─────────────────────\n"
            f"  │  수신 주소 : 127.0.0.1:{port} — 이 PC 에서만 열립니다.\n"
            "  │  같은 망의 다른 PC·휴대폰에서는 접속되지 않습니다.\n"
            "  │  외부 접속이 필요해지면 python app.py --public\n"
            "  │  (절차·주의: remote-access.md)\n"
            "  └───────────────────────────────────────────────"
        )
    if first_pw:
        print(
            "\n  ┌─ 관리자 계정을 처음 만들었습니다 ─────────────\n"
            f"  │  아이디   : admin\n"
            f"  │  비밀번호 : {first_pw}\n"
            "  │  로그인 후 [계정 관리]에서 비밀번호를 바꾸고\n"
            "  │  각 기관 계정을 발급하세요. 이 비밀번호는\n"
            "  │  해시로만 저장되어 다시 볼 수 없습니다.\n"
            "  └───────────────────────────────────────────────"
        )
    print("\n  (Ctrl+C 로 종료)\n")
    if 브라우저:
        try:
            webbrowser.open(url)
        except Exception:
            pass
    ThreadingHTTPServer((host, port), Handler).serve_forever()
