"""주간 리포트 엔진 — specs/자동화_흐름도.md 의 처리 5~15단계

feature1_aggregate.py(검증된 1~3·7단계)를 그대로 재사용하고,
그 위에 취합·비교·이상치·특이사항·일정·요약초안을 얹는다.

사람이 판단하는 3곳(4·12·16단계)은 화면(app.py)에서 처리한다.
"""

import re
from datetime import date

from openpyxl import load_workbook

import storage
from feature1_aggregate import process, process_rows, read_rows, COUNT_COLS

OUTLIER_THRESHOLD = 0.10  # 작년 동기 대비 10%p


# ── 보조 시트 읽기 ───────────────────────────────────────────────


def read_sheet(path, name, cols):
    """선택 시트를 읽는다. 없으면 빈 목록."""
    wb = load_workbook(path, data_only=True)
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    idx = {c: header.index(c) for c in cols if c in header}
    out = []
    for values in rows_iter:
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        row = {}
        for c in cols:
            v = values[idx[c]] if c in idx and idx[c] < len(values) else None
            row[c] = "" if v is None else str(v).strip()
        out.append(row)
    return out


def read_notes(path):
    return read_sheet(path, "특이사항", ["기관명", "과정명", "분류", "내용", "확인필요"])


def read_schedule(path):
    return read_sheet(path, "주요일정", ["날짜", "기관명", "구분", "내용"])


# ── 5단계 취합 · 현황판 ──────────────────────────────────────────


def submission_board(result, roster):
    """대상 명단 대비 제출·검토 상태. (4단계 승인은 화면에서)

    roster 가 비어 있으면 업로드 파일에 등장한 기관을 대상으로 삼는다.
    """
    in_file = []
    for r in result["표"]:
        if r["기관명"] and r["기관명"] not in in_file:
            in_file.append(r["기관명"])

    targets = roster or in_file
    err_rows = {}
    for r in result["표"]:
        if r["_오류"]:
            err_rows[r["기관명"]] = err_rows.get(r["기관명"], 0) + 1

    board = []
    for name in targets:
        if name not in in_file:
            board.append({"기관명": name, "제출": False, "오류행": 0, "상태": "미제출", "비고": "제출 없음"})
        elif err_rows.get(name):
            n = err_rows[name]
            board.append(
                {"기관명": name, "제출": True, "오류행": n, "상태": "반려 대상", "비고": f"오류 {n}행 → 재입력 요청"}
            )
        else:
            board.append({"기관명": name, "제출": True, "오류행": 0, "상태": "검토대기", "비고": ""})

    # 명단에 없는데 제출한 기관
    for name in in_file:
        if name not in targets:
            board.append(
                {"기관명": name, "제출": True, "오류행": err_rows.get(name, 0), "상태": "명단 외", "비고": "대상 명단에 없음"}
            )
    return board


# ── 9단계 전주 대비 비교 ─────────────────────────────────────────


def _by_group(summary):
    return {g["구분"]: g for g in summary}


def compare_weeks(this, last):
    """전주 대비 이수율·탈락률 증감(%p). last 가 None 이면 첫 주."""
    if last is None:
        return None

    def diff(a, b):
        # 지표는 float | None | 문자열 상수 셋 중 하나다. 숫자끼리만 뺀다.
        return (a - b) * 100 if isinstance(a, float) and isinstance(b, float) else None

    t, l = this["합계_전체"][0], last["합계_전체"][0]
    out = {
        "전체": {
            "이수율": diff(t["이수율"], l["이수율"]),
            "탈락률": diff(t["탈락률"], l["탈락률"]),
            "지난_이수율": l["이수율"],
            "지난_탈락률": l["탈락률"],
        },
        "기관별": {},
    }
    lg = _by_group(last["합계_기관별"])
    for g in this["합계_기관별"]:
        prev = lg.get(g["구분"])
        out["기관별"][g["구분"]] = {
            "이수율": diff(g["이수율"], prev["이수율"]) if prev else None,
            "탈락률": diff(g["탈락률"], prev["탈락률"]) if prev else None,
            "지난_이수율": prev["이수율"] if prev else None,
        }
    return out


# ── 11단계 이상치 플래그 ─────────────────────────────────────────


def course_key(name):
    """과정명에서 기수를 떼어 작년 과정과 맞춘다. '스마트팩토리 운영 4기' → '스마트팩토리 운영'"""
    return re.sub(r"\s*\d+\s*기\s*$", "", str(name)).strip()


def flag_outliers(this, lastyear):
    """작년 동기 대비 이수율 10%p 이상 차이 나는 과정을 표시."""
    if lastyear is None:
        return None

    prev = {}
    for r in lastyear["표"]:
        if not r["_오류"] and isinstance(r["이수율"], float):
            prev[(r["기관명"], course_key(r["과정명"]))] = r

    flags, matched = [], 0
    for r in this["표"]:
        if r["_오류"] or not isinstance(r["이수율"], float):
            continue
        p = prev.get((r["기관명"], course_key(r["과정명"])))
        if not p:
            continue
        matched += 1
        gap = r["이수율"] - p["이수율"]
        if abs(gap) >= OUTLIER_THRESHOLD:
            flags.append(
                {
                    "기관명": r["기관명"],
                    "과정명": r["과정명"],
                    "이번": r["이수율"],
                    "작년": p["이수율"],
                    "작년과정": p["과정명"],
                    "차이": gap * 100,
                    "방향": "급감" if gap < 0 else "급증",
                }
            )
    flags.sort(key=lambda f: abs(f["차이"]), reverse=True)
    return {"플래그": flags, "비교된과정": matched}


# ── 10·15단계 특이사항·일정 정리 ─────────────────────────────────


def group_notes(notes):
    """분류(출결/시설/기타)별로 묶고, 확인필요를 앞세운다."""
    order, groups = [], {}
    for n in notes:
        cat = n["분류"] or "기타"
        if cat not in order:
            order.append(cat)
        groups.setdefault(cat, []).append(n)
    for cat in groups:
        groups[cat].sort(key=lambda n: n["확인필요"].upper() != "Y")
    return [(c, groups[c]) for c in order]


def need_check_count(notes):
    return sum(1 for n in notes if n["확인필요"].upper() == "Y")


# ── 14단계 요약 초안 ─────────────────────────────────────────────


def say_pct(v):
    """요약 문장용 지표 표기. 숫자가 아니면 그 사유를 그대로 문장에 넣는다."""
    if isinstance(v, float):
        return f"{v * 100:.1f}%"
    return v if isinstance(v, str) else "산출 불가(입력 없음)"


def draft_summary(week, this, cmp_, outliers, notes):
    t = this["합계_전체"][0]
    c = this["행수"]
    lines = []

    if t["과정수"] == 0:
        head = f"{week} 는 집계할 수 있는 과정이 없습니다."
        if t["제외"]:
            head += f" 제출된 {t['제외']}개 과정이 모두 입력 오류로 제외되어, 정정 후 다시 집계해야 합니다."
        lines.append(head)
        return "\n".join(lines + ["오류를 정정하기 전에는 전주 대비 비교와 이상치 판단을 할 수 없습니다."])

    head = f"{week} 교육실적을 취합한 결과, 집계 대상 {t['과정수']}개 과정의 전체 이수율은 {say_pct(t['이수율'])}, 탈락률은 {say_pct(t['탈락률'])}입니다."
    if t["제외"]:
        head += f" 입력 오류로 {t['제외']}개 과정이 집계에서 제외되어, 오류 정정 후 수치가 달라질 수 있습니다."
    lines.append(head)

    if cmp_ is None:
        lines.append("직전 주차 데이터가 없어 전주 대비 비교는 생략하고, 이번 주를 기준주로 삼았습니다.")
    else:
        d = cmp_["전체"]["이수율"]
        if d is None:
            lines.append("전주 대비 이수율 증감은 비교 자료가 부족해 산출하지 못했습니다.")
        else:
            mark = "▲" if d > 0 else ("▼" if d < 0 else "―")
            lines.append(
                f"전주 이수율 {cmp_['전체']['지난_이수율']*100:.1f}% 대비 {mark}{abs(d):.1f}%p 변동했습니다."
            )

    if outliers is None:
        lines.append("작년 동기 자료가 없어 이상치 플래그는 이번 주차에 적용하지 않았습니다.")
    elif outliers["플래그"]:
        top = outliers["플래그"][0]
        names = ", ".join(f"{f['기관명']} {f['과정명']}" for f in outliers["플래그"][:3])
        lines.append(
            f"작년 동기 대비 10%p 이상 차이 나는 과정이 {len(outliers['플래그'])}건 확인되었습니다({names})."
            f" 특히 {top['기관명']} {top['과정명']}은 작년 {top['작년']*100:.1f}% → 이번 {top['이번']*100:.1f}%로 {abs(top['차이']):.1f}%p {top['방향']}해 확인이 필요합니다."
        )
    else:
        lines.append(f"작년 동기와 비교한 {outliers['비교된과정']}개 과정 모두 10%p 이내로, 이상치는 없습니다.")

    nc = need_check_count(notes)
    if notes:
        cats = ", ".join(f"{c}({len(v)})" for c, v in group_notes(notes))
        lines.append(f"보고된 특이사항은 {len(notes)}건이며 분류별로 {cats}입니다." + (f" 이 중 {nc}건은 확인이 필요합니다." if nc else ""))
    return "\n".join(lines)


# ── 오늘 할 일 ───────────────────────────────────────────────────

PRIORITY = ("높음", "보통", "낮음")


def today_todos(rep, today=None):
    """리포트 결과에서 '오늘 손대야 하는 것'만 뽑아 한 줄씩 모은다.

    리포트는 현황판·이상치·특이사항·일정으로 흩어져 있어서
    무엇부터 해야 하는지 한눈에 안 보인다. 판단은 새로 하지 않고
    이미 계산된 상태값만 행동 문장으로 옮긴다.
    """
    today = today or date.today().isoformat()
    out = []

    def add(priority, kind, target, what):
        out.append({"우선순위": priority, "구분": kind, "대상": target, "할일": what})

    # 승인 워크플로 — 리포트 확정을 막고 있는 것부터
    for b in rep["현황판"]:
        if b["상태"] == "미제출":
            add("높음", "제출 독촉", b["기관명"], "이번 주차 실적 미제출 — 제출 요청")
        elif b["상태"] == "반려 대상":
            add("높음", "재입력 요청", b["기관명"], f"입력 오류 {b['오류행']}행 — 정정 후 재제출 요청")
        elif b["상태"] == "명단 외":
            add("보통", "명단 확인", b["기관명"], "대상 명단에 없는 기관이 제출함 — 명단 확인")
        elif b["상태"] == "검토대기":
            add("보통", "승인 검토", b["기관명"], "제출 완료 — 승인 여부 검토")

    if rep["이상치"]:
        for f in rep["이상치"]["플래그"]:
            add("높음", "이상치 확인", f["기관명"],
                f"{f['과정명']} 이수율 {f['차이']:+.1f}%p {f['방향']} — 원인 확인")

    for n in rep["특이사항"]:
        if n["확인필요"].upper() == "Y":
            add("보통", "특이사항 확인", n["기관명"], f"{n['과정명'] or '-'} · {n['내용']}")

    for s in rep["주요일정"]:
        if str(s["날짜"])[:10] == today:
            add("높음", "오늘 일정", s["기관명"], f"{s['구분']} — {s['내용']}")

    out.sort(key=lambda t: PRIORITY.index(t["우선순위"]))
    return out


# ── 전체 조립 ────────────────────────────────────────────────────


def open_source(src, only_approved=False):
    """입력 출처를 (집계결과, 특이사항, 주요일정) 으로 연다.

    src 는 엑셀 파일 경로이거나, 저장소에서 읽은 주차 데이터(dict)다.
    어느 쪽이든 같은 검증·집계 규칙(feature1)을 통과시킨다.
    """
    if src is None:
        return None, [], []
    if isinstance(src, dict):  # 저장소 (화면 직접 입력 + 업로드가 쌓인 것)
        rows = storage.to_rows(src, only_approved)
        return process_rows(rows), storage.to_notes(src), storage.to_plans(src)
    return process(src), read_notes(src), read_schedule(src)  # 엑셀 파일 1건


def build_report(this_src, last_src=None, ly_src=None, roster=None, week="이번 주차", only_approved=False, today=None):
    this, notes, schedule = open_source(this_src, only_approved)  # 1~3·5·7단계
    last, _, _ = open_source(last_src, only_approved)
    ly, _, _ = open_source(ly_src, only_approved)

    cmp_ = compare_weeks(this, last)  # 8~9단계
    outliers = flag_outliers(this, ly)  # 11단계

    rep = {
        "주차": week,
        "실적": this,
        "현황판": submission_board(this, roster or []),  # 출력 ①
        "비교": cmp_,  # 9단계 (None = 첫 주)
        "이상치": outliers,  # 11단계 (None = 작년 자료 없음)
        "특이사항": notes,
        "특이사항_분류": group_notes(notes),
        "확인필요수": need_check_count(notes),
        "주요일정": sorted(schedule, key=lambda s: s["날짜"]),
        "요약초안": draft_summary(week, this, cmp_, outliers, notes),
        "첫주": last is None,
    }
    rep["오늘할일"] = today_todos(rep, today)  # 나머지 결과에서 파생되므로 마지막에
    rep["기준일"] = today or date.today().isoformat()
    return rep


def totals_row(result):
    return result["합계_전체"][0]


__all__ = [
    "build_report",
    "today_todos",
    "totals_row",
    "COUNT_COLS",
    "OUTLIER_THRESHOLD",
]
