"""핵심 기능 1 — 훈련실적 엑셀 검증·표준화·집계 (정부 「지산맞」 양식)

specs/feature-1-spec.md 명세 구현.
훈련기관이 쓰는 정부 「산업맞춤형 공동훈련센터(지산맞) 훈련실적」 파일을 읽어
형식 확인 → 표준화 → 지표 계산 → 합계 산출 후 통합 실적표·합계·오류 목록을 만든다.

과정별 실적은 두 시트에 나뉘어 있다:
    「양성훈련 현황」 = 양성(채용예정자) 과정
    「향상훈련 현황」 = 향상(재직자) 과정
머리글은 5행쯤에 있고 데이터는 그 아래 「소  계」 행 전까지다. 열은 이름으로 찾으므로
양성/향상의 열 순서 차이나 정부 원본·간소 양식을 모두 읽는다.

사용법:
    python feature1_aggregate.py [엑셀파일경로]
    (경로 생략 시 sample/2026년 A기관.xlsx)
"""

import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

# ── 시트·필드 ────────────────────────────────────────────────────

SHEETS = ("양성훈련 현황", "향상훈련 현황")  # 읽는 두 시트 (구분: 양성/향상)
TARGET_SHEET = "교육실적"  # 기관 연간 목표 훈련인원이 있는 시트

# 앱 내부 필드
TEXT_COLS = ["기관명", "전문기관명", "정기수시", "과정구분", "NCS대분류명", "KECO세분류명", "과정명"]
COUNT_COLS = ["훈련목표인원", "훈련실시인원", "중도탈락자", "훈련중", "훈련수료인원", "취업인원"]
NEED_TEXT = ["기관명", "과정명"]                                    # 없으면 오류
NEED_COUNT = ["훈련목표인원", "훈련실시인원", "중도탈락자", "훈련수료인원"]  # 없으면 오류
OPT_COUNT = ["훈련중", "취업인원"]                                   # 비어도 됨
# 기관 연간 목표(교육실적 시트) — 각 행에 붙여 두고 합계 단계에서 '실시율(목표 대비)'의 분모로 쓴다.
# 훈련목표인원(과정 정원)은 모집률용, 이 목표는 정부 총계의 실시율·수료율(목표 대비) 분모.
TARGET_FIELDS = ["목표_총", "목표_양성", "목표_향상", "목표_수시"]

# 정부 양식 머리글(정규화) → 앱 필드. 이름이 정확히 같을 때만 매칭한다
# ('훈련인원'이 '평균훈련인원' 안에 걸리지 않도록 부분일치는 쓰지 않는다).
HEADER_MAP = {
    "훈련센터명": "기관명",
    "훈련기관명": "전문기관명",
    "정기/수시": "정기수시",
    "과정구분": "과정구분",
    "NCS대분류명": "NCS대분류명",
    "KECO세분류명": "KECO세분류명",
    "훈련과정명": "과정명",
    "훈련인원": "훈련목표인원",
    "실시인원": "훈련실시인원",
    "중도탈락": "중도탈락자",      # 향상 시트
    "중도탈락인원": "중도탈락자",  # 양성 시트
    "훈련중": "훈련중",
    "수료인원": "훈련수료인원",
    "취업인원": "취업인원",
}
# 머리글 행을 찾을 때 반드시 함께 있어야 하는 앱 필드(과정명 + 실시인원)
HEADER_ANCHOR = ("과정명", "훈련실시인원")

BLANK = "(미기재)"        # 분류값이 비어 있을 때 합계에서 쓰는 이름
UNVERIFIED = "검증 필요"  # 오류 행의 지표 (계산해도 신뢰할 수 없음)
UNCOMPUTABLE = "계산 불가"  # 분모가 0이라 계산 자체가 안 되는 경우
NOT_APPLICABLE = "해당없음"  # 취업률처럼 그 구분에 개념이 없는 경우 (향상 과정)


def norm(s):
    """머리글 비교용 — 공백·줄바꿈 제거."""
    return "".join(str(s).split()) if s is not None else ""


# ── 표 출력 보조 (한글 폭 보정) ──────────────────────────────────


def width(s):
    return sum(2 if unicodedata.east_asian_width(c) in "WF" else 1 for c in str(s))


def pad(s, w, align="left"):
    gap = " " * max(0, w - width(s))
    return gap + str(s) if align == "right" else str(s) + gap


def table(headers, rows, aligns=None, rule_before=None):
    """rule_before: 해당 인덱스 행 앞에 구분선을 넣는다 (합계 행 구분용)."""
    aligns = aligns or ["left"] * len(headers)
    cells = [headers] + [[str(c) for c in r] for r in rows]
    widths = [max(width(row[i]) for row in cells) for i in range(len(headers))]
    line = "─┼─".join("─" * w for w in widths)

    out = [" │ ".join(pad(h, w) for h, w in zip(headers, widths)), line]
    for i, r in enumerate(rows):
        if rule_before is not None and i == rule_before:
            out.append(line)
        out.append(" │ ".join(pad(c, w, a) for c, w, a in zip(r, widths, aligns)))
    return "\n".join(out)


def pct(v):
    """지표 표시. 문자열(검증 필요/계산 불가/해당없음)은 그대로 통과시킨다."""
    if isinstance(v, str):
        return v
    return "입력 없음" if v is None else f"{v * 100:.1f}%"


def num(v, raw=None):
    """인원 표시. 읽지 못한 값은 원본을 그대로 보여준다."""
    if v is not None:
        return v
    return f"'{raw}'" if raw not in (None, "") else ""


# ── 1. 읽기 ──────────────────────────────────────────────────────


def _find_header(ws, limit=12):
    """머리글 행을 찾아 (행번호, {필드: 열인덱스}) 반환. 못 찾으면 (None, None)."""
    for r, values in enumerate(ws.iter_rows(min_row=1, max_row=limit, values_only=True), start=1):
        idx = {}
        for col, cell in enumerate(values):
            field = HEADER_MAP.get(norm(cell))
            if field and field not in idx:  # 같은 이름이 또 나오면 첫 열을 쓴다
                idx[field] = col
        if all(a in idx for a in HEADER_ANCHOR):
            return r, idx
    return None, None


def _is_total_row(순번):
    """'소  계'·'합계'·'총계' 같은 마무리 행인지."""
    return "계" in norm(순번)


def read_sheet_rows(ws, 구분, start_n):
    """한 현황 시트를 과정 행 목록으로. 머리글 아래 '소 계' 전까지."""
    header_row, idx = _find_header(ws)
    if header_row is None:
        return [], start_n
    rows, n = [], start_n
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        순번 = values[0] if values else None
        if _is_total_row(순번):
            break  # 소계/총계 행에서 멈춘다
        if all(v is None or str(v).strip() == "" for v in values):
            continue  # 빈 행은 건너뜀
        def get(field):
            c = idx.get(field)
            return values[c] if c is not None and c < len(values) else None
        # 과정명·기관명이 모두 비면 실데이터가 아니라고 보고 건너뛴다
        if not (str(get("과정명") or "").strip() or str(get("기관명") or "").strip()):
            continue
        n += 1
        rows.append({"_행": n, "_시트": ws.title, "구분": 구분,
                     **{f: get(f) for f in TEXT_COLS + COUNT_COLS}})
    return rows, n


def _first_number_after(values, start):
    """values[start+1:] 에서 처음 나오는 숫자를 정수로. 없으면 None."""
    for v in values[start + 1:]:
        iv = to_int(v)
        if iv is not None:
            return iv
    return None


def read_targets(wb):
    """교육실적 시트에서 기관 연간 목표 훈련인원을 읽는다.

    라벨(정기(채용예정자)/정기(재직자향상)/수시) 오른쪽의 첫 숫자를 값으로 본다.
    양성=정기(채용예정자), 향상=정기(재직자향상). 없으면 빈 dict.
    """
    if TARGET_SHEET not in wb.sheetnames:
        return {}
    ws = wb[TARGET_SHEET]
    out = {}
    for values in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        for i, cell in enumerate(values):
            label = norm(cell)
            if label.startswith("지산맞") and "목표훈련인원" in label and "목표_총" not in out:
                out["목표_총"] = _first_number_after(values, i)
            elif label == "정기(채용예정자)":
                out["목표_양성"] = _first_number_after(values, i)
            elif label == "정기(재직자향상)":
                out["목표_향상"] = _first_number_after(values, i)
            elif label == "수시":
                out["목표_수시"] = _first_number_after(values, i)
    return {k: v for k, v in out.items() if v is not None}


def read_rows(path):
    """정부 양식의 양성·향상 현황 시트를 읽어 과정 행 목록으로.

    각 행에는 그 기관의 연간 목표(교육실적 시트)를 붙여 둔다 — 합계 단계에서
    '실시율(목표 대비)'의 분모로 쓰기 위함이다.
    """
    wb = load_workbook(path, data_only=True)
    present = [s for s in SHEETS if s in wb.sheetnames]
    if not present:
        raise SystemExit(
            f"오류: 「양성훈련 현황」·「향상훈련 현황」 시트를 찾을 수 없습니다. "
            f"(시트 목록: {', '.join(wb.sheetnames)})"
        )
    targets = read_targets(wb)
    rows, n = [], 1
    for sheet in present:
        구분 = "양성" if sheet.startswith("양성") else "향상"
        got, n = read_sheet_rows(wb[sheet], 구분, n)
        rows.extend(got)
    if not rows:
        raise SystemExit("오류: 양성·향상 현황 시트에 과정 데이터가 없습니다. 머리글 아래에 실적을 채워 주세요.")
    for r in rows:  # 기관 목표를 모든 과정 행에 복사 (합계에서 기관별로 한 번만 쓴다)
        for k in TARGET_FIELDS:
            r[k] = targets.get(k)
    return rows


# ── 2~3. 형식 확인 + 표준화 ──────────────────────────────────────


def to_int(v):
    """인원 값을 정수로 변환. 불가하면 None."""
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return int(v) if float(v).is_integer() else None
    s = str(v).strip().replace(",", "").replace("명", "")
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f.is_integer() else None


def check_and_normalize(row):
    """한 행을 검사·표준화한다. (표준화된 행, 오류목록) 반환."""
    errors = []
    n = row["_행"]
    out = {"_행": n, "_시트": row.get("_시트"), "구분": row.get("구분") or "", "_원본": {}}

    # 문자 항목: 공백 정리 + 필수값 확인
    for c in TEXT_COLS:
        v = "" if row.get(c) is None else str(row[c]).strip()
        if c in NEED_TEXT and not v:
            errors.append((n, c, "필수값 누락"))
        out[c] = v

    # 인원 항목: 숫자 변환 + 음수 확인 (선택 항목은 비어도 오류 아님)
    for c in COUNT_COLS:
        raw = row.get(c)
        out["_원본"][c] = raw
        blank = raw is None or str(raw).strip() == ""
        if blank:
            if c in NEED_COUNT:
                errors.append((n, c, "필수값 누락"))
            out[c] = None
            continue
        v = to_int(raw)
        if v is None:
            errors.append((n, c, f"숫자로 읽을 수 없음 (값: '{raw}')"))
        elif v < 0:
            errors.append((n, c, f"음수는 올 수 없음 (값: {v})"))
            v = None
        out[c] = v

    # 논리 검사: 값이 둘 다 정상일 때만
    sil, su, tal = out["훈련실시인원"], out["훈련수료인원"], out["중도탈락자"]
    if sil is not None and su is not None and su > sil:
        errors.append((n, "훈련수료인원", f"실시인원({sil})보다 많음 (수료 {su})"))
    if sil is not None and tal is not None and tal > sil:
        errors.append((n, "중도탈락자", f"실시인원({sil})보다 많음 (탈락 {tal})"))

    # 집계 구분: 수시는 양성/향상과 별도 버킷으로 뽑는다 (정부 총계 방식)
    out["집계구분"] = "수시" if out.get("정기수시") == "수시" else (out["구분"] or "기타")
    out["_기관구분"] = f'{out["기관명"] or BLANK} · {out["집계구분"]}'
    for k in TARGET_FIELDS:  # 기관 연간 목표는 그대로 실어 나른다 (검증 대상 아님)
        out[k] = row.get(k)
    return out, errors


# ── 4. 지표 계산 ─────────────────────────────────────────────────


def ratio(numerator, denominator):
    """분모가 0이면 '계산 불가', 값 자체가 없으면 None."""
    if denominator == 0:
        return UNCOMPUTABLE
    if numerator is None or denominator is None:
        return None
    return numerator / denominator


def 취업률(취업, 수료, has_취업):
    """취업 개념이 없는 구분(향상만)이면 '해당없음'."""
    if not has_취업:
        return NOT_APPLICABLE
    return ratio(취업, 수료)


def add_metrics(row):
    """오류 행은 값이 신뢰할 수 없으므로 지표를 계산하지 않는다.

    과정 단위 지표만 여기서 낸다. '실시율(목표 대비)'는 기관 연간 목표가 분모라
    합계 단계(summarize)에서 계산한다.
    """
    if row["_오류"]:
        for k in ("모집률", "수료율", "탈락률", "취업률"):
            row[k] = UNVERIFIED
    else:
        정원, 실시 = row["훈련목표인원"], row["훈련실시인원"]
        수료, 탈락, 취업 = row["훈련수료인원"], row["중도탈락자"], row["취업인원"]
        row["모집률"] = ratio(실시, 정원)          # 정원 대비 실시(모집)
        row["수료율"] = ratio(수료, 실시)          # 실시 대비 수료
        row["탈락률"] = ratio(탈락, 실시)          # 실시 대비 중도탈락
        row["취업률"] = 취업률(취업, 수료, row["구분"] == "양성")
    return row


# ── 5. 합계 산출 ─────────────────────────────────────────────────


def _sum(items, col):
    return sum((i[col] or 0) for i in items)


def group_target(items, target_kind):
    """그룹의 '목표 훈련인원'. 기관별로 한 번만 더한다(행마다 복사돼 있으므로).

    target_kind 'total' = 기관 연간 목표(총), 'bucket' = 그 구분(양성/향상/수시)의 목표.
    연간 목표가 없는 기관(직접 입력 등)은 그 기관의 정원(훈련목표인원) 합으로 대체한다.
    None 이면 목표 기반 지표를 내지 않는다(예: NCS/KECO 그룹).
    """
    if target_kind is None:
        return None
    by_inst = {}
    for r in items:
        by_inst.setdefault(r.get("기관명") or BLANK, []).append(r)
    total = 0
    for rs in by_inst.values():
        field = "목표_총" if target_kind == "total" else f'목표_{rs[0].get("집계구분")}'
        t = next((x.get(field) for x in rs if x.get(field) is not None), None)
        total += t if t is not None else sum((x.get("훈련목표인원") or 0) for x in rs)
    return total


def summarize(clean, dropped, key, target_kind=None):
    """key 기준 합계. clean=집계 대상, dropped=오류로 빠진 행. 순서는 등장 순.

    target_kind 를 주면 '실시율(목표 대비)'·'수료율(목표 대비)'를 기관 연간 목표로 낸다.
    """
    order, groups, drops = [], {}, {}
    for r in clean + dropped:
        name = r.get(key) or BLANK
        if name not in order:
            order.append(name)
    for r in clean:
        groups.setdefault(r.get(key) or BLANK, []).append(r)
    for r in dropped:
        nm = r.get(key) or BLANK
        drops[nm] = drops.get(nm, 0) + 1

    out = []
    for name in order:
        items = groups.get(name, [])
        s = {c: _sum(items, c) for c in COUNT_COLS}
        has_취업 = any(i["구분"] == "양성" for i in items)
        목표 = group_target(items, target_kind)
        out.append({
            "구분": name,
            "과정수": len(items),
            "제외": drops.get(name, 0),
            **s,
            "목표훈련인원": 목표,
            "실시율": ratio(s["훈련실시인원"], 목표) if 목표 is not None else None,       # 목표 대비 실시
            "수료율_목표": ratio(s["훈련수료인원"], 목표) if 목표 is not None else None,  # 목표 대비 수료
            "모집률": ratio(s["훈련실시인원"], s["훈련목표인원"]),  # 정원 대비 실시
            "수료율": ratio(s["훈련수료인원"], s["훈련실시인원"]),  # 실시 대비 수료
            "탈락률": ratio(s["중도탈락자"], s["훈련실시인원"]),
            "취업률": 취업률(s["취업인원"], s["훈련수료인원"], has_취업),
        })
    return out


# ── 6. 결과 반환 ─────────────────────────────────────────────────


def process(path):
    """읽기(1) → 검증·표준화·지표·합계(2~6)."""
    return process_rows(read_rows(path))


def process_rows(raw):
    """읽어들인 행 목록에 동작 2~6을 수행한다.

    입력 출처(엑셀 업로드 / 화면 직접 입력)와 무관하게 같은 규칙을 적용하려고
    읽기(1단계)와 분리했다. 각 행은 _행·구분 과 TEXT/COUNT 필드를 가진 dict.
    """
    rows, errors = [], []
    for r in raw:
        norm_row, errs = check_and_normalize(r)
        norm_row["_오류"] = bool(errs)
        rows.append(add_metrics(norm_row))
        errors.extend(errs)

    clean = [r for r in rows if not r["_오류"]]
    dropped = [r for r in rows if r["_오류"]]

    total = summarize(
        [{**r, "_전체": "전체"} for r in clean],
        [{**r, "_전체": "전체"} for r in dropped],
        "_전체", target_kind="total",
    )
    if not total:  # 행이 없어도 합계 한 줄은 있어야 소비하는 쪽이 안 깨진다
        total = [{"구분": "전체", "과정수": 0, "제외": 0, **{c: 0 for c in COUNT_COLS},
                  "목표훈련인원": None, "실시율": UNCOMPUTABLE, "수료율_목표": UNCOMPUTABLE,
                  "모집률": UNCOMPUTABLE, "수료율": UNCOMPUTABLE, "탈락률": UNCOMPUTABLE,
                  "취업률": UNCOMPUTABLE}]

    return {
        "표": rows,  # 오류 행 포함 (지표는 '검증 필요')
        "합계_전체": total,
        "합계_기관별": summarize(clean, dropped, "기관명", target_kind="total"),
        "합계_구분별": summarize(clean, dropped, "집계구분", target_kind="bucket"),      # 양성/향상/수시
        "합계_기관구분별": summarize(clean, dropped, "_기관구분", target_kind="bucket"),  # 기관 × 구분
        "합계_NCS별": summarize(clean, dropped, "NCS대분류명"),
        "합계_KECO별": summarize(clean, dropped, "KECO세분류명"),
        "오류": errors,
        "행수": {"전체": len(rows), "정상": len(clean), "오류": len(dropped)},
    }


# ── 출력 (CLI) ───────────────────────────────────────────────────


def show(result, path):
    R = "right"
    c = result["행수"]
    print(f"\n입력 파일: {path}")
    print(f"읽은 과정: {c['전체']}개 (정상 {c['정상']} / 오류 {c['오류']})")

    # ① 통합 실적표
    print("\n" + "=" * 60)
    print("■ 통합 실적표")
    print("=" * 60)
    rows = [
        [
            "⚠" if r["_오류"] else "",
            r["_행"],
            r["구분"],
            r["기관명"],
            r["과정명"][:22],
            *(num(r[col], r["_원본"].get(col)) for col in ("훈련목표인원", "훈련실시인원", "훈련수료인원", "중도탈락자")),
            pct(r["모집률"]),
            pct(r["수료율"]),
        ]
        for r in result["표"]
    ]
    t = result["합계_전체"][0]
    rows.append([
        "", "", "", "합계",
        f"정상 {t['과정수']}개" + (f" · 제외 {t['제외']}개" if t["제외"] else ""),
        *(t[col] for col in ("훈련목표인원", "훈련실시인원", "훈련수료인원", "중도탈락자")),
        pct(t["실시율"]), pct(t["수료율"]),
    ])
    print(table(
        ["", "행", "구분", "훈련기관", "과정(정원)명", "정원", "실시", "수료", "탈락", "모집/실시율", "수료율"],
        rows,
        ["left", R, "left", "left", "left", R, R, R, R, R, R],
        rule_before=len(rows) - 1,
    ))
    if c["오류"]:
        print(f"\n⚠ 표시 {c['오류']}개 과정은 오류가 있어 지표 계산·합계에서 제외됨 (오류 목록 참고)")

    # ② 구분별 합계
    for title, key in [
        ("기관별 합계", "합계_기관별"),
        ("양성/향상/수시 합계", "합계_구분별"),
        ("NCS대분류별 합계", "합계_NCS별"),
    ]:
        print("\n" + "=" * 60)
        print(f"■ {title}")
        print("=" * 60)
        rows = [
            [
                g["구분"], g["과정수"], g["제외"] or "",
                g["목표훈련인원"] if g["목표훈련인원"] is not None else g["훈련목표인원"],
                g["훈련실시인원"], g["훈련수료인원"], g["중도탈락자"],
                pct(g["실시율"]), pct(g["수료율"]), pct(g["취업률"]),
            ]
            for g in result[key]
        ]
        print(table(
            ["구분", "과정수", "제외", "목표", "실시", "수료", "탈락", "실시율", "수료율", "취업률"],
            rows,
            ["left", R, R, R, R, R, R, R, R, R],
        ))

    # ③ 오류 목록
    print("\n" + "=" * 60)
    print(f"■ 오류 목록  ({len(result['오류'])}건)")
    print("=" * 60)
    if not result["오류"]:
        print("오류 없음")
    else:
        print(table(["행", "컬럼", "사유"], [[n, col, m] for n, col, m in result["오류"]], [R, "left", "left"]))
    print()


if __name__ == "__main__":
    default = Path(__file__).parent / "sample" / "2026년 A기관.xlsx"
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else default
    if not target.exists():
        raise SystemExit(f"오류: 파일을 찾을 수 없습니다 → {target}")
    show(process(target), target)
